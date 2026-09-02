from __future__ import annotations

from io import BytesIO
import math
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.dpp_consolidation_service import (
    _material_code,
    _material_key,
    _normalize_unit,
)

SOURCE_SHEET = "DPP"


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip().lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _number(value: object, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else default
    try:
        return float(str(value).strip().replace(".", "").replace(",", "."))
    except ValueError:
        return default


def _cell(rows: list[tuple], row: int, column: int) -> object:
    if row < 1 or column < 1 or row > len(rows):
        return None
    current = rows[row - 1]
    return current[column - 1] if column <= len(current) else None


def _find_header_row(rows: list[tuple]) -> int:
    max_scan = min(len(rows), 40)
    for row_number in range(1, max_scan + 1):
        values = {
            _normalize(value)
            for value in rows[row_number - 1][:16]
            if value not in (None, "")
        }
        if {"material", "descricao", "um", "grupo origem"}.issubset(values):
            return row_number
    raise ValueError("DPP base: linha de cabeçalho não encontrada.")


def _headers(rows: list[tuple], header_row: int) -> dict[int, str]:
    return {
        column: str(value).strip()
        for column, value in enumerate(rows[header_row - 1], start=1)
        if value not in (None, "")
    }


def _find_column(headers: dict[int, str], name: str, required: bool = True) -> int | None:
    target = _normalize(name)
    for column, value in headers.items():
        if _normalize(value) == target:
            return column
    if required:
        raise ValueError(f"DPP base: coluna '{name}' não encontrada.")
    return None


def _find_label_row(rows: list[tuple], header_row: int, label: str, contains: bool = False) -> int | None:
    target = _normalize(label)
    for row in range(1, header_row):
        for column in range(1, min(8, len(rows[row - 1])) + 1):
            current = _normalize(_cell(rows, row, column))
            if (contains and target in current) or (not contains and current == target):
                return row
    return None


def _looks_like_model_code(value: object) -> bool:
    if not isinstance(value, str):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/]*-[A-Za-z0-9._/-]+", value.strip()))


def _detect_model_code_row(rows: list[tuple], header_row: int, start: int, end: int) -> int | None:
    best_row = None
    best_score = 0
    for row in range(max(1, header_row - 5), header_row):
        score = sum(1 for column in range(start, end + 1) if _looks_like_model_code(_cell(rows, row, column)))
        if score > best_score:
            best_row = row
            best_score = score
    return best_row


def parse_previous_dpp(content: bytes) -> tuple[dict[str, dict], list[dict], dict]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError("DPP base: a aba 'DPP' não foi encontrada.")

    sheet = workbook[SOURCE_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValueError("DPP base: a aba 'DPP' está vazia.")

    header_row = _find_header_row(rows)
    headers = _headers(rows, header_row)
    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    origin_col = _find_column(headers, "Grupo Origem")
    optional_col = _find_column(headers, "OPC", required=False)
    check_col = _find_column(headers, "Check", required=False)
    price_col = _find_column(headers, "Preço", required=False)

    kit_row = _find_label_row(rows, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(rows, header_row, "REAL")

    model_start = origin_col + 1
    model_end = (check_col - 1) if check_col else origin_col
    model_code_row = _detect_model_code_row(rows, header_row, model_start, model_end) if model_end >= model_start else None

    models: list[dict] = []
    if model_end >= model_start:
        for column in range(model_start, model_end + 1):
            name = _text(_cell(rows, header_row, column))
            if not name:
                continue
            models.append(
                {
                    "name": name,
                    "code": _text(_cell(rows, model_code_row, column)) if model_code_row else None,
                    "previous_kit_pgd": _number(_cell(rows, kit_row, column), 0.0) if kit_row else 0.0,
                    "previous_real": _number(_cell(rows, real_row, column), 0.0) if real_row else 0.0,
                    "column": get_column_letter(column),
                }
            )

    materials: dict[str, dict] = {}
    for excel_row in range(header_row + 1, len(rows) + 1):
        raw_material = _cell(rows, excel_row, material_col)
        material = _material_code(raw_material)
        key = _material_key(raw_material)
        if not material or not key:
            continue

        optional_material = _material_code(_cell(rows, excel_row, optional_col)) if optional_col else None
        price = _number(_cell(rows, excel_row, price_col), 0.0) if price_col else 0.0
        materials[key] = {
            "material": material,
            "material_key": key,
            "description": _text(_cell(rows, excel_row, description_col)),
            "um": _normalize_unit(_cell(rows, excel_row, um_col)),
            "group_origin": _text(_cell(rows, excel_row, origin_col)),
            "optional_material": optional_material,
            "price": price,
            "from_history": True,
            "historical_source": {
                "sheet": SOURCE_SHEET,
                "row": excel_row,
                "material_cell": f"{get_column_letter(material_col)}{excel_row}",
                "reference": f"{SOURCE_SHEET}!{get_column_letter(material_col)}{excel_row}",
            },
        }

    optional_count = sum(1 for item in materials.values() if item.get("optional_material"))
    return materials, models, {
        "sheet": SOURCE_SHEET,
        "header_row": header_row,
        "materials": len(materials),
        "optional_materials": optional_count,
        "models": len(models),
        "kit_row": kit_row,
        "real_row": real_row,
        "model_code_row": model_code_row,
    }
