from __future__ import annotations

from collections.abc import Callable
from copy import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from app.services.dpp_consolidation_service import _material_key
from app.services.dpp_monthly_base import (
    SOURCE_SHEET,
    _find_column,
    _find_header_row,
    _find_label_row,
    _headers,
    _normalize,
)
from app.services.dpp_scenario_service import get_monthly_scenario

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
ProgressCallback = Callable[[int, str], None]

# Estas colunas pertencem à lógica estrutural do DPP. Quando o arquivo-base possui
# fórmula nelas, o ORION preserva a fórmula e a traduz para cada linha do cenário.
# Campos não calculados pelo ORION (Preço, Amount e Coments) não são herdados do
# mês anterior para não misturar dados históricos com o cenário atual.
FORMULA_PREFERRED_HEADERS = {"check", "wiu", "nec", "stk ttl", "saldo"}


def _notify(progress: ProgressCallback | None, value: int, activity: str) -> None:
    if progress is not None:
        progress(value, activity)


def _number(value: object) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _find_any_column(headers: dict[int, str], *names: str) -> int | None:
    normalized = {_normalize(name) for name in names}
    for column, value in headers.items():
        if _normalize(value) in normalized:
            return column
    return None


def _find_stock_source_column(headers: dict[int, str]) -> int | None:
    for column, value in headers.items():
        normalized = _normalize(value)
        if normalized.startswith("stk ") and normalized not in {"stk op", "stk ttl"}:
            return column
    return None


def _find_explosion_column(headers: dict[int, str]) -> int | None:
    for column, value in headers.items():
        if _normalize(value).startswith("explosao"):
            return column
    return None


def _output_name(reference_month: str, extension: str) -> str:
    year, month = (reference_month.split("-") + ["", ""])[:2]
    if year and month:
        return f"DPP_ORION_{year}_{month}{extension}"
    return f"DPP_ORION{extension}"


def _set_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    if hasattr(calculation, "calcMode"):
        calculation.calcMode = "auto"
    if hasattr(calculation, "fullCalcOnLoad"):
        calculation.fullCalcOnLoad = True
    if hasattr(calculation, "forceFullCalc"):
        calculation.forceFullCalc = True


def _copy_row_layout(sheet, source_row: int, target_row: int) -> None:
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel

    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _material_rows(sheet, *, header_row: int, material_col: int) -> list[int]:
    return [
        row
        for row in range(header_row + 1, sheet.max_row + 1)
        if _material_key(sheet.cell(row, material_col).value)
    ]


def _formula_templates(sheet, *, headers: dict[int, str], material_rows: list[int]) -> dict[int, tuple[str, str]]:
    templates: dict[int, tuple[str, str]] = {}
    wanted_columns = {
        column
        for column, header in headers.items()
        if _normalize(header) in FORMULA_PREFERRED_HEADERS
    }
    if not wanted_columns:
        return templates

    for row in material_rows:
        for column in wanted_columns:
            if column in templates:
                continue
            cell = sheet.cell(row, column)
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                templates[column] = (cell.coordinate, value)
        if len(templates) == len(wanted_columns):
            break
    return templates


def _translated_formula(formula: str, *, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        # Fórmulas fora do subconjunto compreendido pelo Translator continuam sendo
        # preservadas em vez de desaparecerem. O Excel fará o recálculo ao abrir.
        return formula


def _restore_formula_templates(sheet, row: int, templates: dict[int, tuple[str, str]]) -> None:
    for column, (origin, formula) in templates.items():
        target = sheet.cell(row, column)
        target.value = _translated_formula(formula, origin=origin, target=target.coordinate)


def _ordered_scenario_materials(sheet, *, scenario: dict, header_row: int, material_col: int) -> list[dict]:
    scenario_materials = [
        material
        for material in (scenario.get("materials") or [])
        if _material_key(material.get("material"))
    ]
    scenario_by_key = {
        _material_key(material.get("material")): material
        for material in scenario_materials
    }
    if len(scenario_by_key) != len(scenario_materials):
        raise ValueError(
            "O cenário ORION contém materiais duplicados após a normalização. "
            "A exportação foi interrompida para não gerar um Excel diferente do Dashboard."
        )

    existing_order: list[str] = []
    seen: set[str] = set()
    for row in _material_rows(sheet, header_row=header_row, material_col=material_col):
        key = _material_key(sheet.cell(row, material_col).value)
        if key and key not in seen:
            existing_order.append(key)
            seen.add(key)

    ordered_keys = [key for key in existing_order if key in scenario_by_key]
    ordered_keys.extend(key for key in scenario_by_key if key not in seen)
    return [scenario_by_key[key] for key in ordered_keys]


def _prepare_exact_material_rows(
    sheet,
    *,
    header_row: int,
    material_col: int,
    target_count: int,
    formula_templates: dict[int, tuple[str, str]],
) -> list[int]:
    existing_material_rows = _material_rows(sheet, header_row=header_row, material_col=material_col)
    existing_end = max(existing_material_rows, default=header_row)
    existing_slots = max(existing_end - header_row, 0)
    style_source_row = existing_material_rows[0] if existing_material_rows else header_row + 1

    if target_count > existing_slots:
        insert_at = existing_end + 1 if existing_end > header_row else header_row + 1
        amount = target_count - existing_slots
        sheet.insert_rows(insert_at, amount)
        for row in range(insert_at, insert_at + amount):
            _copy_row_layout(sheet, style_source_row, row)
            _restore_formula_templates(sheet, row, formula_templates)
    elif target_count < existing_slots:
        delete_at = header_row + target_count + 1
        sheet.delete_rows(delete_at, existing_slots - target_count)

    return list(range(header_row + 1, header_row + target_count + 1))


def _audit_material_projection(sheet, *, rows: list[int], material_col: int, materials: list[dict]) -> None:
    expected_keys = [_material_key(material.get("material")) for material in materials]
    exported_keys = [_material_key(sheet.cell(row, material_col).value) for row in rows]

    if len(exported_keys) != len(expected_keys) or any(not key for key in exported_keys):
        raise ValueError(
            f"Falha de consistência do Excel ORION: o cenário possui {len(expected_keys)} materiais, "
            f"mas a faixa exportada possui {sum(1 for key in exported_keys if key)} materiais preenchidos."
        )
    if exported_keys != expected_keys:
        raise ValueError(
            "Falha de consistência do Excel ORION: a sequência de materiais exportada não corresponde ao cenário do Dashboard."
        )


def _audit_formula_projection(
    sheet,
    *,
    rows: list[int],
    formula_templates: dict[int, tuple[str, str]],
) -> None:
    for column in formula_templates:
        missing = [
            row
            for row in rows
            if not (
                isinstance(sheet.cell(row, column).value, str)
                and sheet.cell(row, column).value.startswith("=")
            )
        ]
        if missing:
            header = sheet.cell(rows[0] - 1, column).value if rows else column
            raise ValueError(
                f"Falha de consistência do Excel ORION: a fórmula estrutural da coluna '{header}' "
                f"não foi propagada para {len(missing)} linha(s)."
            )


def _write_orion_scenario_to_dpp(
    workbook,
    scenario: dict,
    progress: ProgressCallback | None = None,
) -> dict:
    _notify(progress, 18, "Lendo estrutura da aba DPP")
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError("O DPP do mês anterior usado como modelo não possui a aba 'DPP'.")

    sheet = workbook[SOURCE_SHEET]
    source_rows = list(sheet.iter_rows(values_only=True))
    if not source_rows:
        raise ValueError("A aba 'DPP' do arquivo-base está vazia.")

    header_row = _find_header_row(source_rows)
    headers = _headers(source_rows, header_row)

    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check", required=False)
    wiu_col = _find_any_column(headers, "WIU")
    optional_col = _find_column(headers, "OPC", required=False)

    stock_sap_col = _find_stock_source_column(headers)
    explosion_col = _find_explosion_column(headers)
    stock_op_col = _find_any_column(headers, "STK OP", "STK OPC", "STK OPCS")
    stock_total_col = _find_any_column(headers, "STK TTL", "STK TOTAL")
    nec_col = _find_any_column(headers, "NEC", "NECESSIDADE")
    balance_col = _find_any_column(headers, "SALDO", "BALANCE")

    kit_row = _find_label_row(source_rows, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(source_rows, header_row, "REAL")
    if kit_row is None or real_row is None:
        raise ValueError("O layout do DPP do mês anterior não contém as linhas KIT Disponível PGD e REAL esperadas.")

    model_end = (check_col - 1) if check_col else origin_col
    if model_end < origin_col + 1:
        raise ValueError("Não foi possível localizar as colunas de modelos no layout do DPP do mês anterior.")

    _notify(progress, 24, "Mapeando modelos no layout do Excel")
    scenario_models = {
        _normalize(model.get("name")): model
        for model in scenario.get("models", [])
        if model.get("name")
    }
    template_model_columns = [
        column
        for column in range(origin_col + 1, model_end + 1)
        if sheet.cell(header_row, column).value not in (None, "")
    ]
    model_columns: dict[int, dict | None] = {}
    total_model_columns = max(len(template_model_columns), 1)
    last_model_progress = -1
    for index, column in enumerate(template_model_columns, start=1):
        template_name = sheet.cell(header_row, column).value
        model = scenario_models.get(_normalize(template_name))
        model_columns[column] = model
        sheet.cell(kit_row, column).value = _number(model.get("kit_pgd")) if model else 0.0
        sheet.cell(real_row, column).value = _number(model.get("real")) if model else 0.0

        current_progress = 24 + int((index / total_model_columns) * 10)
        if current_progress != last_model_progress:
            _notify(progress, current_progress, f"Preenchendo KIT e REAL · {index}/{len(template_model_columns)} modelos")
            last_model_progress = current_progress

    _notify(progress, 35, "Preparando faixa exata de materiais")
    source_material_rows = _material_rows(sheet, header_row=header_row, material_col=material_col)
    formula_templates = _formula_templates(
        sheet,
        headers=headers,
        material_rows=source_material_rows,
    )
    ordered_materials = _ordered_scenario_materials(
        sheet,
        scenario=scenario,
        header_row=header_row,
        material_col=material_col,
    )
    target_rows = _prepare_exact_material_rows(
        sheet,
        header_row=header_row,
        material_col=material_col,
        target_count=len(ordered_materials),
        formula_templates=formula_templates,
    )
    formula_columns = set(formula_templates)
    last_operational_col = max(headers)

    _notify(progress, 38, "Preenchendo dados calculados pelo ORION")
    total_material_rows = max(len(target_rows), 1)
    last_material_progress = -1
    for index, (row, material) in enumerate(zip(target_rows, ordered_materials, strict=True), start=1):
        # A linha anterior é somente molde visual. Todos os valores operacionais são
        # limpos antes de projetar o cenário atual para impedir vazamento do mês anterior.
        for column in range(material_col, last_operational_col + 1):
            sheet.cell(row, column).value = None
        _restore_formula_templates(sheet, row, formula_templates)

        sheet.cell(row, material_col).value = material.get("material")
        sheet.cell(row, description_col).value = material.get("description")
        sheet.cell(row, um_col).value = material.get("um")
        sheet.cell(row, origin_col).value = material.get("group_origin")

        consumption = material.get("consumption_by_model") or {}
        normalized_consumption = {_normalize(name): _number(value) for name, value in consumption.items()}
        for column, model in model_columns.items():
            model_name = _normalize(model.get("name")) if model else _normalize(sheet.cell(header_row, column).value)
            sheet.cell(row, column).value = normalized_consumption.get(model_name, 0.0)

        # Check é um campo do DPP/WIU. O status interno OK/INVESTIGAR não deve ser
        # exportado nessa coluna, pois é uma classificação diferente.
        if check_col and check_col not in formula_columns:
            sheet.cell(row, check_col).value = material.get("check") or None
        if wiu_col and wiu_col not in formula_columns:
            sheet.cell(row, wiu_col).value = "WIU" if material.get("in_current_wiu") else None
        if optional_col:
            sheet.cell(row, optional_col).value = material.get("optional_material")
        if stock_sap_col:
            sheet.cell(row, stock_sap_col).value = _number(material.get("stock_sap_effective", material.get("stock_sap")))
        if explosion_col:
            sheet.cell(row, explosion_col).value = _number(material.get("explosion"))
        if stock_op_col:
            sheet.cell(row, stock_op_col).value = _number(material.get("stock_op"))
        if stock_total_col and stock_total_col not in formula_columns:
            sheet.cell(row, stock_total_col).value = _number(material.get("stock_total"))
        if nec_col and nec_col not in formula_columns:
            sheet.cell(row, nec_col).value = _number(material.get("nec"))
        if balance_col and balance_col not in formula_columns:
            sheet.cell(row, balance_col).value = _number(material.get("balance"))

        current_progress = 38 + int((index / total_material_rows) * 50)
        if current_progress != last_material_progress:
            _notify(progress, current_progress, f"Preenchendo materiais · {index}/{len(target_rows)}")
            last_material_progress = current_progress

    _notify(progress, 89, "Validando consistência do cenário exportado")
    _audit_material_projection(
        sheet,
        rows=target_rows,
        material_col=material_col,
        materials=ordered_materials,
    )
    _audit_formula_projection(
        sheet,
        rows=target_rows,
        formula_templates=formula_templates,
    )

    _notify(progress, 90, "Configurando recálculo das fórmulas")
    _set_recalculation(workbook)
    return {
        "header_row": header_row,
        "material_col": material_col,
        "material_count": len(ordered_materials),
        "material_keys": [_material_key(material.get("material")) for material in ordered_materials],
        "formula_columns": sorted(formula_columns),
    }


def _validate_serialized_export(content: bytes, audit: dict) -> None:
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError("Falha de consistência do Excel ORION: a aba DPP desapareceu após a serialização.")
        sheet = workbook[SOURCE_SHEET]
        start_row = int(audit["header_row"]) + 1
        count = int(audit["material_count"])
        material_col = int(audit["material_col"])
        exported_keys = [
            _material_key(sheet.cell(row, material_col).value)
            for row in range(start_row, start_row + count)
        ]
        if exported_keys != audit["material_keys"]:
            raise ValueError(
                "Falha de consistência do Excel ORION após salvar: os materiais do arquivo não correspondem ao cenário do Dashboard."
            )
        for column in audit.get("formula_columns") or []:
            for row in range(start_row, start_row + count):
                value = sheet.cell(row, int(column)).value
                if not (isinstance(value, str) and value.startswith("=")):
                    raise ValueError(
                        "Falha de consistência do Excel ORION após salvar: uma fórmula estrutural não foi preservada."
                    )
    finally:
        workbook.close()


def export_monthly_scenario_excel(
    *,
    scenario_id: str,
    template_content: bytes,
    template_filename: str,
    progress: ProgressCallback | None = None,
) -> tuple[bytes, str, str]:
    _notify(progress, 4, "Validando cenário ORION")
    scenario = get_monthly_scenario(scenario_id)
    if scenario is None:
        raise ValueError("Cenário ORION não encontrado ou expirado. Gere o cenário mensal novamente.")

    _notify(progress, 8, "Validando DPP do mês anterior")
    extension = Path(template_filename or "dpp.xlsx").suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Use como modelo o DPP do mês anterior em formato .xlsx ou .xlsm.")
    if not template_content:
        raise ValueError("O DPP do mês anterior usado como modelo está vazio.")

    keep_vba = extension == ".xlsm"
    _notify(progress, 12, "Abrindo workbook do mês anterior")
    workbook = load_workbook(BytesIO(template_content), data_only=False, keep_vba=keep_vba)
    _notify(progress, 16, "Workbook carregado")
    try:
        audit = _write_orion_scenario_to_dpp(workbook, scenario, progress=progress)
        _notify(progress, 94, "Serializando workbook para Excel")
        output = BytesIO()
        workbook.save(output)
        content = output.getvalue()
        _notify(progress, 97, "Conferindo arquivo serializado")
        _validate_serialized_export(content, audit)
        _notify(progress, 99, "Finalizando arquivo para download")
    finally:
        workbook.close()

    media_type = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if keep_vba
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return content, _output_name(scenario.get("reference_month") or "", extension), media_type
