from __future__ import annotations

from collections import Counter, OrderedDict
from decimal import Decimal, InvalidOperation
from io import BytesIO
import math
from pathlib import Path
import re
import unicodedata

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
WIU_SOURCE_SHEET = "POWER"
WIU_MATRIX_SHEET = "WIU JULHO"
EXPLOSION_SOURCE_SHEET = "CONSOLIDADO"
STOCK_SOURCE_SHEET = "CONSOLIDADO"

# Camada criada agora, mas deliberadamente desligada até a regra ser validada com a analista.
UNIT_CONVERSION_ENABLED = False
UNIT_CONVERSION_FACTORS = {
    ("KG", "G"): 1000.0,
    ("M", "CM"): 100.0,
    ("L", "ML"): 1000.0,
}
OPEN_MAX_ENTRIES_PER_MATERIAL = 25


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip().lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _number(value: object, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else default

    text = str(value).strip().replace(" ", "")
    if not text:
        return default
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return default
    return number if math.isfinite(number) else default


def _material_code(value: object) -> str | None:
    """Converte códigos vindos do Excel/SAP para uma chave textual estável.

    Evita a falha clássica de PROCV entre 14731008000020 (número)
    e "14731008000020" (texto), sem remover zeros à esquerda de códigos
    que já chegam como texto.
    """
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")

    text = str(value).strip()
    if not text:
        return None

    numeric_decimal = re.fullmatch(r"([0-9]+)\.0+", text)
    if numeric_decimal:
        return numeric_decimal.group(1)

    if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?[eE][+-]?[0-9]+", text):
        try:
            decimal_value = Decimal(text)
            if decimal_value == decimal_value.to_integral():
                return format(decimal_value.quantize(Decimal("1")), "f")
        except InvalidOperation:
            pass

    return text


def _material_key(value: object) -> str | None:
    code = _material_code(value)
    return code.upper() if code else None


def _was_material_normalized(raw: object, normalized_code: str | None) -> bool:
    if normalized_code is None:
        return False
    if not isinstance(raw, str):
        return True
    return raw.strip() != normalized_code


def _normalize_unit(value: object) -> str | None:
    text = _text(value)
    return text.upper() if text else None


def _unit_conversion(source_unit: object, target_unit: object) -> dict:
    source = _normalize_unit(source_unit)
    target = _normalize_unit(target_unit)
    if not source or not target:
        return {
            "source_unit": source,
            "target_unit": target,
            "mismatch": False,
            "supported": False,
            "factor": None,
            "enabled": UNIT_CONVERSION_ENABLED,
            "applied": False,
            "status": "UNAVAILABLE",
        }

    if source == target:
        return {
            "source_unit": source,
            "target_unit": target,
            "mismatch": False,
            "supported": True,
            "factor": 1.0,
            "enabled": UNIT_CONVERSION_ENABLED,
            "applied": False,
            "status": "SAME_UNIT",
        }

    factor = UNIT_CONVERSION_FACTORS.get((source, target))
    return {
        "source_unit": source,
        "target_unit": target,
        "mismatch": True,
        "supported": factor is not None,
        "factor": factor,
        "enabled": UNIT_CONVERSION_ENABLED,
        "applied": bool(UNIT_CONVERSION_ENABLED and factor is not None),
        "status": "CONVERSION_AVAILABLE_DISABLED" if factor is not None else "UNSUPPORTED",
    }


def _validate_upload(file: UploadFile, label: str) -> None:
    filename = file.filename or ""
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"{label}: envie um arquivo .xlsx ou .xlsm.")


def _headers(row: tuple) -> dict[int, str]:
    return {
        index: str(value).strip()
        for index, value in enumerate(row, start=1)
        if value not in (None, "")
    }


def _find_column(headers: dict[int, str], name: str) -> int:
    target = _normalize(name)
    for column, value in headers.items():
        if _normalize(value) == target:
            return column
    raise ValueError(f"Coluna obrigatória '{name}' não encontrada.")


def _find_column_startswith(headers: dict[int, str], prefix: str) -> int:
    target = _normalize(prefix)
    for column, value in headers.items():
        if _normalize(value).startswith(target):
            return column
    raise ValueError(f"Coluna iniciada por '{prefix}' não encontrada.")


def _value(row: tuple, column: int) -> object:
    if column < 1 or column > len(row):
        return None
    return row[column - 1]


def _read_sheet(content: bytes, sheet_name: str) -> list[tuple]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"A aba '{sheet_name}' não foi encontrada.")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValueError(f"A aba '{sheet_name}' está vazia.")
    return rows


def _read_model_order(content: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if WIU_MATRIX_SHEET not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook[WIU_MATRIX_SHEET]
    first_two = []
    for row in sheet.iter_rows(min_row=1, max_row=2, values_only=True):
        first_two.append(tuple(row))
    workbook.close()
    if len(first_two) < 2:
        return []

    code_row, name_row = first_two
    models = []
    for column in range(5, max(len(code_row), len(name_row)) + 1):
        name = _text(_value(name_row, column))
        code = _text(_value(code_row, column))
        if not name or not code:
            continue
        models.append(
            {
                "name": name,
                "code": code,
                "column": get_column_letter(column),
            }
        )
    return models


def _parse_wiu(content: bytes) -> tuple[list[dict], list[dict], dict]:
    rows = _read_sheet(content, WIU_SOURCE_SHEET)
    headers = _headers(rows[0])

    code_col = _find_column(headers, "Código")
    model_col = _find_column(headers, "Modelo")
    material_col = _find_column(headers, "Componente")
    description_col = _find_column(headers, "Descrição")
    usage_col = _find_column(headers, "Uso BOM")
    um_col = _find_column(headers, "Unidade de Medida")
    origin_col = _find_column(headers, "Grupo Origem")

    model_order = _read_model_order(content)
    model_meta: OrderedDict[str, dict] = OrderedDict(
        (item["name"], dict(item)) for item in model_order
    )
    materials: OrderedDict[str, dict] = OrderedDict()
    imported_source_rows = 0
    national_source_rows = 0
    normalized_codes = 0

    for excel_row, row in enumerate(rows[1:], start=2):
        raw_material = _value(row, material_col)
        material = _material_code(raw_material)
        material_key = _material_key(raw_material)
        model = _text(_value(row, model_col))
        origin = _text(_value(row, origin_col))
        if not material or not material_key or not model:
            continue

        if _normalize(origin) != "importado":
            national_source_rows += 1
            continue
        imported_source_rows += 1
        if _was_material_normalized(raw_material, material):
            normalized_codes += 1

        model_code = _text(_value(row, code_col))
        if model not in model_meta:
            model_meta[model] = {
                "name": model,
                "code": model_code,
                "column": None,
            }
        elif not model_meta[model].get("code") and model_code:
            model_meta[model]["code"] = model_code

        current = materials.setdefault(
            material_key,
            {
                "material": material,
                "material_key": material_key,
                "description": _text(_value(row, description_col)),
                "um": _normalize_unit(_value(row, um_col)),
                "group_origin": origin or "Importado",
                "consumption_by_model": {},
                "wiu_first_row": excel_row,
                "wiu_occurrences": 0,
            },
        )
        current["wiu_occurrences"] += 1
        if not current.get("description"):
            current["description"] = _text(_value(row, description_col))
        if not current.get("um"):
            current["um"] = _normalize_unit(_value(row, um_col))

        usage = _number(_value(row, usage_col), 0.0)
        if usage:
            current["consumption_by_model"][model] = (
                current["consumption_by_model"].get(model, 0.0) + usage
            )

    model_names = list(model_meta.keys())
    for material in materials.values():
        used_models = [
            model for model in model_names
            if abs(material["consumption_by_model"].get(model, 0.0)) > 1e-12
        ]
        material["used_models"] = used_models
        material["check"] = "// ".join(used_models)
        material["source"] = {
            "sheet": WIU_SOURCE_SHEET,
            "first_row": material.pop("wiu_first_row"),
            "occurrences": material.pop("wiu_occurrences"),
            "material_column": get_column_letter(material_col),
        }

    material_count_by_model = {name: 0 for name in model_names}
    for material in materials.values():
        for model in material["used_models"]:
            material_count_by_model[model] += 1

    models = []
    for item in model_meta.values():
        models.append(
            {
                **item,
                "material_count": material_count_by_model.get(item["name"], 0),
            }
        )

    diagnostics = {
        "sheet": WIU_SOURCE_SHEET,
        "source_rows": max(len(rows) - 1, 0),
        "imported_source_rows": imported_source_rows,
        "excluded_non_imported_rows": national_source_rows,
        "imported_materials": len(materials),
        "models": len(models),
        "material_codes_normalized": normalized_codes,
    }
    return list(materials.values()), models, diagnostics


def _parse_explosion(content: bytes) -> tuple[dict[str, dict], dict]:
    rows = _read_sheet(content, EXPLOSION_SOURCE_SHEET)
    headers = _headers(rows[0])
    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    explosion_col = _find_column_startswith(headers, "EXPLOSÃO")

    result: dict[str, dict] = {}
    normalized_codes = 0
    for excel_row, row in enumerate(rows[1:], start=2):
        raw_material = _value(row, material_col)
        material = _material_code(raw_material)
        key = _material_key(raw_material)
        if not material or not key:
            continue
        if _was_material_normalized(raw_material, material):
            normalized_codes += 1
        result[key] = {
            "material": material,
            "value": _number(_value(row, explosion_col), 0.0),
            "description": _text(_value(row, description_col)),
            "um": _normalize_unit(_value(row, um_col)),
            "source": {
                "sheet": EXPLOSION_SOURCE_SHEET,
                "row": excel_row,
                "cell": f"{get_column_letter(explosion_col)}{excel_row}",
            },
        }

    return result, {
        "sheet": EXPLOSION_SOURCE_SHEET,
        "rows": len(result),
        "value_header": headers[explosion_col],
        "material_codes_normalized": normalized_codes,
    }


def _parse_stock(content: bytes) -> tuple[dict[str, dict], dict]:
    rows = _read_sheet(content, STOCK_SOURCE_SHEET)
    headers = _headers(rows[0])
    material_col = _find_column(headers, "Material")
    stock_col = _find_column_startswith(headers, "STK TTL")
    try:
        um_col = _find_column(headers, "UOM")
    except ValueError:
        um_col = None

    result: dict[str, dict] = {}
    normalized_codes = 0
    for excel_row, row in enumerate(rows[1:], start=2):
        raw_material = _value(row, material_col)
        material = _material_code(raw_material)
        key = _material_key(raw_material)
        if not material or not key:
            continue
        if _was_material_normalized(raw_material, material):
            normalized_codes += 1
        result[key] = {
            "material": material,
            "value": _number(_value(row, stock_col), 0.0),
            "um": _normalize_unit(_value(row, um_col)) if um_col else None,
            "source": {
                "sheet": STOCK_SOURCE_SHEET,
                "row": excel_row,
                "cell": f"{get_column_letter(stock_col)}{excel_row}",
            },
        }

    return result, {
        "sheet": STOCK_SOURCE_SHEET,
        "rows": len(result),
        "value_header": headers[stock_col],
        "material_codes_normalized": normalized_codes,
    }


def _find_open_sheet_and_header(content: bytes):
    required = {
        "status", "po", "pi", "lote", "modelo", "material",
        "texto breve", "qtd.do pedido", "uom",
    }
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=20, values_only=True),
            start=1,
        ):
            headers = _headers(tuple(row))
            normalized_headers = {_normalize(value) for value in headers.values()}
            if required.issubset(normalized_headers):
                return workbook, sheet, row_number, headers
    workbook.close()
    raise ValueError(
        "OPEN: não foi possível localizar uma aba com STATUS, PO, PI, LOTE, MODELO, "
        "Material, Texto breve, Qtd.do pedido e UOM."
    )


def _parse_open(content: bytes) -> tuple[dict[str, dict], dict]:
    workbook, sheet, header_row, headers = _find_open_sheet_and_header(content)
    status_col = _find_column(headers, "STATUS")
    po_col = _find_column(headers, "PO")
    pi_col = _find_column(headers, "PI")
    lot_col = _find_column(headers, "LOTE")
    model_col = _find_column(headers, "MODELO")
    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Texto breve")
    quantity_col = _find_column(headers, "Qtd.do pedido")
    um_col = _find_column(headers, "UOM")

    pending_by_material: dict[str, dict] = {}
    status_counts: Counter[str] = Counter()
    source_rows = 0
    pending_rows = 0
    normalized_codes = 0

    for excel_row, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        source_rows += 1
        raw_material = _value(row, material_col)
        material = _material_code(raw_material)
        key = _material_key(raw_material)
        if not material or not key:
            continue

        if _was_material_normalized(raw_material, material):
            normalized_codes += 1

        status = _text(_value(row, status_col)) or "Sem status"
        status_counts[status] += 1
        if "pendente" not in _normalize(status):
            continue

        pending_rows += 1
        current = pending_by_material.setdefault(
            key,
            {
                "material": material,
                "description": _text(_value(row, description_col)),
                "pending_records": 0,
                "pending_quantity": 0.0,
                "entries": {},
            },
        )
        current["pending_records"] += 1
        quantity = _number(_value(row, quantity_col), 0.0)
        current["pending_quantity"] += quantity

        entry_key = (
            _text(_value(row, pi_col)) or "",
            _text(_value(row, po_col)) or "",
            _text(_value(row, lot_col)) or "",
            _text(_value(row, model_col)) or "",
            _normalize_unit(_value(row, um_col)) or "",
        )
        entry = current["entries"].setdefault(
            entry_key,
            {
                "pi": entry_key[0] or None,
                "po": entry_key[1] or None,
                "lot": entry_key[2] or None,
                "model": entry_key[3] or None,
                "um": entry_key[4] or None,
                "quantity": 0.0,
                "source_rows": [],
            },
        )
        entry["quantity"] += quantity
        if len(entry["source_rows"]) < 5:
            entry["source_rows"].append(excel_row)

    workbook.close()

    result: dict[str, dict] = {}
    for key, current in pending_by_material.items():
        entries = list(current["entries"].values())
        entries.sort(key=lambda item: (item["pi"] or "", item["po"] or "", item["lot"] or ""))
        total_entries = len(entries)
        result[key] = {
            "material": current["material"],
            "description": current["description"],
            "pending_records": current["pending_records"],
            "pending_quantity": current["pending_quantity"],
            "entries_total": total_entries,
            "entries": entries[:OPEN_MAX_ENTRIES_PER_MATERIAL],
            "truncated": total_entries > OPEN_MAX_ENTRIES_PER_MATERIAL,
            "source": {
                "sheet": sheet.title,
                "header_row": header_row,
            },
        }

    return result, {
        "sheet": sheet.title,
        "header_row": header_row,
        "source_rows": source_rows,
        "pending_rows": pending_rows,
        "pending_materials": len(result),
        "status_counts": dict(status_counts),
        "material_codes_normalized": normalized_codes,
        "max_entries_per_material": OPEN_MAX_ENTRIES_PER_MATERIAL,
    }


async def consolidate_dpp_sources(
    wiu: UploadFile,
    explosion: UploadFile,
    stock: UploadFile,
    open_orders: UploadFile | None = None,
) -> dict:
    _validate_upload(wiu, "WIU")
    _validate_upload(explosion, "Explosão")
    _validate_upload(stock, "STK SAP")
    if open_orders is not None:
        _validate_upload(open_orders, "OPEN")

    wiu_content = await wiu.read()
    explosion_content = await explosion.read()
    stock_content = await stock.read()
    open_content = await open_orders.read() if open_orders is not None else None

    if not wiu_content:
        raise ValueError("O arquivo WIU está vazio.")
    if not explosion_content:
        raise ValueError("O arquivo de Explosão está vazio.")
    if not stock_content:
        raise ValueError("O arquivo de STK SAP está vazio.")
    if open_orders is not None and not open_content:
        raise ValueError("O arquivo OPEN está vazio.")

    materials, models, wiu_diagnostics = _parse_wiu(wiu_content)
    explosion_map, explosion_diagnostics = _parse_explosion(explosion_content)
    stock_map, stock_diagnostics = _parse_stock(stock_content)

    open_map: dict[str, dict] = {}
    open_diagnostics = None
    if open_content:
        open_map, open_diagnostics = _parse_open(open_content)

    explosion_matches = 0
    stock_matches = 0
    open_pending_matches = 0
    unit_mismatches = 0
    convertible_unit_mismatches = 0

    explosion_only_materials = set(explosion_map)
    stock_only_materials = set(stock_map)
    open_only_materials = set(open_map)

    for material in materials:
        key = material["material_key"]

        explosion_entry = explosion_map.get(key)
        if explosion_entry:
            explosion_matches += 1
            explosion_only_materials.discard(key)
        material["explosion"] = explosion_entry["value"] if explosion_entry else 0.0
        material["explosion_um"] = explosion_entry["um"] if explosion_entry else material["um"]
        material["explosion_source"] = explosion_entry["source"] if explosion_entry else None

        stock_entry = stock_map.get(key)
        if stock_entry:
            stock_matches += 1
            stock_only_materials.discard(key)
        material["stock_sap"] = stock_entry["value"] if stock_entry else 0.0
        material["stock_um"] = stock_entry["um"] if stock_entry else None
        material["stock_source"] = stock_entry["source"] if stock_entry else None

        conversion = _unit_conversion(material["stock_um"], material["um"])
        material["unit_conversion"] = conversion
        if conversion["mismatch"]:
            unit_mismatches += 1
            if conversion["supported"]:
                convertible_unit_mismatches += 1

        stock_effective = material["stock_sap"]
        if conversion["applied"] and conversion["factor"] is not None:
            stock_effective *= conversion["factor"]

        material["stock_sap_effective"] = stock_effective
        material["available_base"] = stock_effective + material["explosion"]
        material["available_base_mode"] = (
            "UNIT_CONVERSION_APPLIED"
            if conversion["applied"]
            else "RAW_SOURCE_VALUES"
        )

        open_entry = open_map.get(key)
        if open_entry:
            open_pending_matches += 1
            open_only_materials.discard(key)
            material["open_investigation"] = {
                "pending_records": open_entry["pending_records"],
                "pending_quantity": open_entry["pending_quantity"],
                "entries_total": open_entry["entries_total"],
                "entries": open_entry["entries"],
                "truncated": open_entry["truncated"],
                "source": open_entry["source"],
            }
        else:
            material["open_investigation"] = {
                "pending_records": 0,
                "pending_quantity": 0.0,
                "entries_total": 0,
                "entries": [],
                "truncated": False,
                "source": None,
            }

        material["nec"] = None
        material["balance"] = None
        material["status"] = "BASE_FISICA_PRONTA"

    materials.sort(key=lambda item: item["material"])

    source_status = [
        {
            "id": "wiu",
            "label": "WIU",
            "required": True,
            "loaded": True,
            "filename": wiu.filename,
            "detail": f"{len(materials)} materiais importados · {len(models)} modelos",
        },
        {
            "id": "explosion",
            "label": "Explosão de Placas",
            "required": True,
            "loaded": True,
            "filename": explosion.filename,
            "detail": f"{explosion_matches} materiais do WIU com registro de explosão",
        },
        {
            "id": "stock",
            "label": "STK SAP (1º dia do mês)",
            "required": True,
            "loaded": True,
            "filename": stock.filename,
            "detail": f"{stock_matches} materiais do WIU encontrados no snapshot SAP",
        },
        {
            "id": "open",
            "label": "OPEN",
            "required": False,
            "loaded": open_content is not None,
            "filename": open_orders.filename if open_orders is not None else None,
            "detail": (
                f"{open_diagnostics['pending_rows']} linhas pendentes · "
                f"{open_pending_matches} materiais da base com evidência OPEN"
                if open_diagnostics
                else "Opcional: usado apenas como evidência de investigação; não altera o estoque"
            ),
        },
    ]

    pending = [
        "KIT PGD / REAL para calcular NEC e SALDO de produção",
        "OPC/STK OP para fechar integralmente materiais opcionais quando aplicável",
        "Confirmar com a analista a regra de conversão KG→G, M→CM e L→ML antes de ativá-la",
    ]

    total_normalized_codes = (
        wiu_diagnostics["material_codes_normalized"]
        + explosion_diagnostics["material_codes_normalized"]
        + stock_diagnostics["material_codes_normalized"]
        + (open_diagnostics["material_codes_normalized"] if open_diagnostics else 0)
    )

    return {
        "mode": "source_consolidation",
        "status": "PARCIAL",
        "scope": (
            "Consolidação construída a partir das fontes mensais. WIU, Explosão e STK SAP formam "
            "a base física obrigatória. OPEN é uma fonte auxiliar de investigação e nunca é somado "
            "ao estoque. Os códigos de material são normalizados para texto antes dos cruzamentos. "
            "A camada de conversão de UM existe, mas permanece desligada até validação da regra."
        ),
        "sources": source_status,
        "summary": {
            "materials": len(materials),
            "models": len(models),
            "explosion_matches": explosion_matches,
            "explosion_without_wiu": len(explosion_only_materials),
            "stock_loaded": True,
            "stock_matches": stock_matches,
            "stock_without_wiu": len(stock_only_materials),
            "open_loaded": open_content is not None,
            "open_pending_materials": open_pending_matches,
            "open_pending_rows": open_diagnostics["pending_rows"] if open_diagnostics else 0,
            "open_without_wiu": len(open_only_materials) if open_content is not None else 0,
            "unit_mismatches": unit_mismatches,
            "convertible_unit_mismatches": convertible_unit_mismatches,
            "material_codes_normalized": total_normalized_codes,
        },
        "capabilities": {
            "material_model_matrix": True,
            "check_from_wiu": True,
            "explosion": True,
            "stock_sap": True,
            "available_base": True,
            "open_investigation": open_content is not None,
            "material_code_normalization": True,
            "unit_conversion_layer": True,
            "unit_conversion_enabled": UNIT_CONVERSION_ENABLED,
            "nec": False,
            "balance": False,
        },
        "unit_conversion": {
            "enabled": UNIT_CONVERSION_ENABLED,
            "supported_pairs": [
                {"from": source, "to": target, "factor": factor}
                for (source, target), factor in UNIT_CONVERSION_FACTORS.items()
            ],
            "note": (
                "Conversões ainda não são aplicadas aos valores. O ORION apenas detecta a diferença "
                "de UM e informa o fator potencial até validação com a analista."
            ),
        },
        "pending": pending,
        "diagnostics": {
            "wiu": wiu_diagnostics,
            "explosion": explosion_diagnostics,
            "stock": stock_diagnostics,
            "open": open_diagnostics,
        },
        "models": models,
        "materials": materials,
    }
