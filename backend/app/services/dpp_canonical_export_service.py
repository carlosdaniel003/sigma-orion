from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.dpp_consolidation_service import _material_key
from app.services.dpp_export_service import (
    SUPPORTED_EXTENSIONS,
    _output_name,
    _write_orion_scenario_to_dpp,
)
from app.services.dpp_monthly_base import (
    SOURCE_SHEET,
    _find_column,
    _find_header_row,
    _find_label_row,
    _headers,
    _normalize,
)
from app.services.dpp_projection_service import (
    BALANCE_NEGATIVE_FIELD,
    BALANCE_POSITIVE_FIELD,
    STOCK_TOTAL_COMPONENT_FIELDS,
    build_orion_projection,
    column_values_equal,
)
from app.services.dpp_scenario_service import get_monthly_scenario

FORMULA_FIELDS = {"check", "nec", "stock_op", "stock_total", "balance", "amount"}
CONSOLIDADO_SHEET = "CONSOLIDADO"


def _set_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    if hasattr(calculation, "calcMode"):
        calculation.calcMode = "auto"
    if hasattr(calculation, "fullCalcOnLoad"):
        calculation.fullCalcOnLoad = True
    if hasattr(calculation, "forceFullCalc"):
        calculation.forceFullCalc = True


def _field_columns(projection: dict) -> dict[str, int]:
    result: dict[str, int] = {}
    for column, spec in projection["specs"].items():
        field = spec.get("field")
        if field and field != "model" and spec.get("supported"):
            result[field] = column
    return result


def _excel_number(value: object) -> str:
    number = float(value or 0.0)
    if number.is_integer():
        return str(int(number))
    return format(number, ".15g")


def _formula_for_field(
    field: str,
    *,
    row: int,
    projection: dict,
    field_columns: dict[str, int],
    real_row: int,
    header_row: int,
    material: dict | None = None,
) -> str | None:
    model_start = get_column_letter(int(projection["model_start"]))
    model_end = get_column_letter(int(projection["model_end"]))

    if field == "check":
        # O DPP Final grava funções futuras no formato OOXML com prefixos _xlfn.
        # Usar a mesma codificação evita que dois Excels com a mesma regra apareçam
        # diferentes em uma auditoria célula a célula do arquivo serializado.
        return (
            f'=_xlfn.TEXTJOIN("// ", TRUE, _xlfn._xlws.FILTER(${model_start}${header_row}:${model_end}${header_row}, '
            f'{model_start}{row}:{model_end}{row}>0, ""))'
        )

    if field == "nec":
        return f"=SUMPRODUCT(${model_start}${real_row}:${model_end}${real_row},{model_start}{row}:{model_end}{row})"

    if field == "stock_op":
        sources = list((material or {}).get("stock_op_sources") or [])
        if not sources:
            return None
        if len(sources) == 1:
            optional_column = field_columns.get("optional_material")
            if optional_column is None:
                return None
            optional_ref = f"{get_column_letter(optional_column)}{row}"
            # O DPP Final usa [N]CONSOLIDADO porque aponta para um arquivo externo.
            # [N] é um índice de externalLink do XLSX, não parte da regra de negócio.
            # No ORION a mesma tabela é incorporada ao workbook para a fórmula continuar
            # recalculável e auditável sem depender do caminho corporativo R:\\.
            return f"=VLOOKUP({optional_ref},{CONSOLIDADO_SHEET}!$A:$F,6,0)"
        return "=" + "+".join(_excel_number(item.get("value")) for item in sources)

    if field == "stock_total":
        component_columns = [field_columns.get(name) for name in STOCK_TOTAL_COMPONENT_FIELDS]
        if any(column is None for column in component_columns):
            return None
        refs = [f"{get_column_letter(int(column))}{row}" for column in component_columns]
        return "=" + "+".join(refs)

    if field == "balance":
        positive = field_columns.get(BALANCE_POSITIVE_FIELD)
        negative = field_columns.get(BALANCE_NEGATIVE_FIELD)
        if positive is None or negative is None:
            return None
        return f"={get_column_letter(positive)}{row}-{get_column_letter(negative)}{row}"

    if field == "amount":
        price = field_columns.get("price")
        balance = field_columns.get("balance")
        if price is None or balance is None:
            return None
        return f"={get_column_letter(price)}{row}*{get_column_letter(balance)}{row}"

    return None


def _workbook_context(workbook, scenario: dict) -> dict:
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError("Falha de consistência do Excel ORION: a aba DPP não foi encontrada.")

    sheet = workbook[SOURCE_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Falha de consistência do Excel ORION: a aba DPP está vazia.")

    header_row = _find_header_row(rows)
    headers = _headers(rows, header_row)
    material_col = _find_column(headers, "Material")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    real_row = _find_label_row(rows, header_row, "REAL")
    kit_row = _find_label_row(rows, header_row, "KIT Disponivel PGD", contains=True)
    if real_row is None or kit_row is None:
        raise ValueError("Falha de consistência do Excel ORION: linhas KIT Disponível PGD/REAL não encontradas.")

    projection = build_orion_projection(
        scenario=scenario,
        headers=headers,
        material_col=material_col,
        origin_col=origin_col,
        check_col=check_col,
    )

    row_by_key: dict[str, int] = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        key = _material_key(sheet.cell(row, material_col).value)
        if key:
            row_by_key[key] = row

    return {
        "sheet": sheet,
        "header_row": header_row,
        "headers": headers,
        "material_col": material_col,
        "origin_col": origin_col,
        "check_col": check_col,
        "real_row": real_row,
        "kit_row": kit_row,
        "projection": projection,
        "row_by_key": row_by_key,
        "field_columns": _field_columns(projection),
    }


def _prepare_consolidado_sheet(workbook, scenario: dict) -> None:
    sources_by_material: dict[str, float] = {}
    for material in scenario.get("materials") or []:
        for source in material.get("stock_op_sources") or []:
            code = str(source.get("material") or "").strip()
            if code:
                sources_by_material[code] = float(source.get("value") or 0.0)

    if not sources_by_material:
        return

    if CONSOLIDADO_SHEET in workbook.sheetnames:
        sheet = workbook[CONSOLIDADO_SHEET]
        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(CONSOLIDADO_SHEET)

    sheet.sheet_state = "hidden"
    sheet.cell(1, 1).value = "Material"
    sheet.cell(1, 6).value = "STK OP"
    for row, (material, value) in enumerate(sorted(sources_by_material.items()), start=2):
        sheet.cell(row, 1).value = material
        sheet.cell(row, 6).value = value


def _apply_gap_formulas(context: dict, scenario: dict) -> None:
    gap_row = context["real_row"] + 1
    if gap_row >= context["header_row"]:
        return

    sheet = context["sheet"]
    projection = context["projection"]
    models = {
        _normalize(model.get("name")): model
        for model in scenario.get("models") or []
        if model.get("name")
    }
    for column in range(projection["model_start"], projection["model_end"] + 1):
        cell = sheet.cell(gap_row, column)
        model = models.get(_normalize(sheet.cell(context["header_row"], column).value))
        difference = float(model.get("difference_real_vs_kit") or 0.0) if model else 0.0
        if abs(difference) > 1e-4:
            letter = get_column_letter(column)
            cell.value = f"={letter}{context['real_row']}-{letter}{context['kit_row']}"
        else:
            cell.value = None


def _apply_summary_formulas(workbook, context: dict) -> None:
    if "Resumo de Análise" not in workbook.sheetnames or "DPP BALANCE" not in workbook.sheetnames:
        return

    sheet = workbook["Resumo de Análise"]
    projection = context["projection"]
    model_start = get_column_letter(int(projection["model_start"]))
    model_end = get_column_letter(int(projection["model_end"]))
    model_name_row = max(context["real_row"] - 1, 1)

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 2).value in (None, ""):
            continue
        sheet.cell(row, 3).value = (
            f'=XLOOKUP(B{row},\'DPP BALANCE\'!${model_start}${model_name_row}:${model_end}${model_name_row},'
            f'\'DPP BALANCE\'!${model_start}${context["kit_row"]}:${model_end}${context["kit_row"]},"ND")'
        )
        sheet.cell(row, 4).value = (
            f'=XLOOKUP(B{row},DPP!${model_start}${model_name_row}:${model_end}${model_name_row},'
            f'DPP!${model_start}${context["real_row"]}:${model_end}${context["real_row"]},"ND")'
        )
        sheet.cell(row, 5).value = f"=D{row}-C{row}"


def _apply_canonical_projection(workbook, scenario: dict, progress=None) -> dict:
    context = _workbook_context(workbook, scenario)
    sheet = context["sheet"]
    projection = context["projection"]
    row_by_key = context["row_by_key"]
    field_columns = context["field_columns"]

    expected_keys = {row["key"] for row in projection["rows"]}
    actual_keys = set(row_by_key)
    if actual_keys != expected_keys:
        missing = len(expected_keys - actual_keys)
        extra = len(actual_keys - expected_keys)
        raise ValueError(
            "Falha de consistência do Excel ORION: a projeção física não corresponde ao cenário "
            f"({missing} material(is) ausente(s), {extra} excedente(s))."
        )

    models = {
        _normalize(model.get("name")): model
        for model in (scenario.get("models") or [])
        if model.get("name")
    }
    for column in range(projection["model_start"], projection["model_end"] + 1):
        name = _normalize(sheet.cell(context["header_row"], column).value)
        model = models.get(name)
        sheet.cell(context["kit_row"], column).value = float(model.get("kit_pgd") or 0.0) if model else 0.0
        sheet.cell(context["real_row"], column).value = float(model.get("real") or 0.0) if model else 0.0

    _prepare_consolidado_sheet(workbook, scenario)
    _apply_gap_formulas(context, scenario)

    canonical_columns = [
        column
        for column in projection["columns"]
        if projection["specs"][column].get("field") != "model"
    ]
    total_rows = max(len(projection["rows"]), 1)
    last_progress = -1

    for index, projected_row in enumerate(projection["rows"], start=1):
        excel_row = row_by_key[projected_row["key"]]
        for column in canonical_columns:
            spec = projection["specs"][column]
            cell = sheet.cell(excel_row, column)

            if not spec["supported"]:
                cell.value = None
                continue

            field = spec.get("field")
            if field in FORMULA_FIELDS:
                formula = _formula_for_field(
                    field,
                    row=excel_row,
                    projection=projection,
                    field_columns=field_columns,
                    real_row=context["real_row"],
                    header_row=context["header_row"],
                    material=projected_row["material"],
                )
                if formula is not None:
                    cell.value = formula
                else:
                    cell.value = projected_row["values"].get(column)
            else:
                cell.value = projected_row["values"].get(column)

        if progress is not None:
            current = 78 + int((index / total_rows) * 10)
            if current != last_progress:
                progress(current, f"Aplicando projeção canônica · {index}/{len(projection['rows'])} materiais")
                last_progress = current

    _apply_summary_formulas(workbook, context)
    _set_recalculation(workbook)
    return context


def _value_at(row_values: tuple, column: int):
    index = column - 1
    return row_values[index] if 0 <= index < len(row_values) else None


def _validate_canonical_projection(content: bytes, scenario: dict, progress=None) -> None:
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError("Falha de consistência do Excel ORION: a aba DPP não foi encontrada após salvar.")

        sheet = workbook[SOURCE_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Falha de consistência do Excel ORION: a aba DPP está vazia após salvar.")

        header_row = _find_header_row(rows)
        headers = _headers(rows, header_row)
        material_col = _find_column(headers, "Material")
        origin_col = _find_column(headers, "Grupo Origem")
        check_col = _find_column(headers, "Check")
        real_row = _find_label_row(rows, header_row, "REAL")
        kit_row = _find_label_row(rows, header_row, "KIT Disponivel PGD", contains=True)
        if real_row is None or kit_row is None:
            raise ValueError("Falha de consistência do Excel ORION: linhas KIT Disponível PGD/REAL não encontradas após salvar.")

        projection = build_orion_projection(
            scenario=scenario,
            headers=headers,
            material_col=material_col,
            origin_col=origin_col,
            check_col=check_col,
        )
        field_columns = _field_columns(projection)

        row_values_by_key: dict[str, tuple[int, tuple]] = {}
        for excel_row, row_values in enumerate(rows, start=1):
            if excel_row <= header_row:
                continue
            key = _material_key(_value_at(row_values, material_col))
            if key:
                row_values_by_key[key] = (excel_row, row_values)

        expected_keys = {row["key"] for row in projection["rows"]}
        if set(row_values_by_key) != expected_keys or len(row_values_by_key) != len(projection["rows"]):
            raise ValueError(
                "Falha de consistência do Excel ORION após salvar: quantidade/lista de materiais difere do Dashboard."
            )

        kit_values = rows[kit_row - 1]
        real_values = rows[real_row - 1]
        header_values = rows[header_row - 1]
        models = {
            _normalize(model.get("name")): model
            for model in (scenario.get("models") or [])
            if model.get("name")
        }
        for column in range(projection["model_start"], projection["model_end"] + 1):
            model = models.get(_normalize(_value_at(header_values, column)))
            expected_kit = float(model.get("kit_pgd") or 0.0) if model else 0.0
            expected_real = float(model.get("real") or 0.0) if model else 0.0
            if abs(float(_value_at(kit_values, column) or 0.0) - expected_kit) > 1e-4:
                raise ValueError("Falha de consistência do Excel ORION: KIT PGD divergiu do cenário após salvar.")
            if abs(float(_value_at(real_values, column) or 0.0) - expected_real) > 1e-4:
                raise ValueError("Falha de consistência do Excel ORION: REAL divergiu do cenário após salvar.")

        total_rows = max(len(projection["rows"]), 1)
        last_progress = -1
        for index, projected_row in enumerate(projection["rows"], start=1):
            excel_row, row_values = row_values_by_key[projected_row["key"]]
            for column in projection["columns"]:
                spec = projection["specs"][column]
                actual = _value_at(row_values, column)

                if not spec["supported"]:
                    if actual not in (None, ""):
                        raise ValueError(
                            "Falha de consistência do Excel ORION: uma coluna não calculada recebeu dado residual."
                        )
                    continue

                field = spec.get("field")
                if field in FORMULA_FIELDS:
                    expected_formula = _formula_for_field(
                        field,
                        row=excel_row,
                        projection=projection,
                        field_columns=field_columns,
                        real_row=real_row,
                        header_row=header_row,
                        material=projected_row["material"],
                    )
                    if expected_formula is not None:
                        if actual != expected_formula:
                            raise ValueError(
                                f"Falha de consistência do Excel ORION: fórmula DPP de '{field}' não foi preservada."
                            )
                        continue

                expected = projected_row["values"].get(column)
                if not column_values_equal(actual, expected, spec.get("comparison")):
                    header = headers.get(column, str(column))
                    raise ValueError(
                        "Falha de consistência do Excel ORION: "
                        f"a coluna '{header}' não corresponde ao cenário do Dashboard."
                    )

            if progress is not None:
                current = 96 + int((index / total_rows) * 3)
                if current != last_progress:
                    progress(current, f"Auditando Excel ORION · {index}/{len(projection['rows'])} materiais")
                    last_progress = current
    finally:
        workbook.close()


def export_monthly_scenario_excel(
    *,
    scenario_id: str,
    template_content: bytes,
    template_filename: str,
    progress=None,
) -> tuple[bytes, str, str]:
    scenario = get_monthly_scenario(scenario_id)
    if scenario is None:
        raise ValueError("Cenário ORION não encontrado ou expirado. Gere o cenário mensal novamente.")

    extension = Path(template_filename or "dpp.xlsx").suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Use como modelo o DPP do mês anterior em formato .xlsx ou .xlsm.")
    if not template_content:
        raise ValueError("O DPP do mês anterior usado como modelo está vazio.")

    keep_vba = extension == ".xlsm"
    if progress is not None:
        progress(3, "Abrindo workbook do mês anterior")

    workbook = load_workbook(BytesIO(template_content), data_only=False, keep_vba=keep_vba)
    try:
        def template_progress(value: int, activity: str) -> None:
            if progress is None:
                return
            mapped = 5 + int((min(max(float(value), 0.0), 100.0) / 100.0) * 71)
            progress(mapped, activity)

        _write_orion_scenario_to_dpp(workbook, scenario, progress=template_progress)

        if progress is not None:
            progress(78, "Aplicando fórmulas do DPP Final ao Cenário ORION")
        _apply_canonical_projection(workbook, scenario, progress=progress)

        if progress is not None:
            progress(90, "Serializando Excel ORION")
        output = BytesIO()
        workbook.save(output)
        canonical_content = output.getvalue()
    finally:
        workbook.close()

    if progress is not None:
        progress(96, "Auditando fórmulas e valores do Excel ORION")
    _validate_canonical_projection(canonical_content, scenario, progress=progress)
    if progress is not None:
        progress(99, "Excel ORION consistente com o Dashboard e fórmulas DPP")

    media_type = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if keep_vba
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return canonical_content, _output_name(scenario.get("reference_month") or "", extension), media_type