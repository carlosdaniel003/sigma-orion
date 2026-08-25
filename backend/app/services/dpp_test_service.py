from __future__ import annotations

from io import BytesIO
import math

from fastapi import UploadFile
from openpyxl import load_workbook

from app.services.dpp_consolidation_service import _material_key, _normalize_unit, _validate_upload
from app.services.dpp_monthly_base import parse_previous_dpp
from app.services.dpp_monthly_service import generate_monthly_dpp
from app.services.dpp_optional_service import optional_material_keyset
from app.services.dpp_scenario_service import recalculate_monthly_scenario
from app.services.dpp_service import (
    SOURCE_SHEET,
    _as_text,
    _cell,
    _detect_model_code_row,
    _find_column,
    _find_explosion_column,
    _find_header_row,
    _find_label_row,
    _find_stock_source_column,
    _headers,
    _normalize,
    _number,
)

REL_TOL = 1e-9
ABS_TOL = 1e-4
MAX_MISMATCH_SAMPLES = 200
MAX_LEGACY_SAMPLES = 200
MAX_HUMAN_SAMPLES = 200


def _same_number(left: object, right: object) -> bool:
    if left is None and right is None:
        return True
    try:
        return math.isclose(float(left or 0.0), float(right or 0.0), rel_tol=REL_TOL, abs_tol=ABS_TOL)
    except (TypeError, ValueError):
        return False


def _same_text(left: object, right: object) -> bool:
    return _normalize(left) == _normalize(right)


def _same_optional_materials(left: object, right: object) -> bool:
    return optional_material_keyset(left) == optional_material_keyset(right)


def _counter() -> dict[str, int]:
    return {
        "checked": 0,
        "matches": 0,
        "mismatches": 0,
        "legacy_corrections": 0,
        "human_interventions": 0,
    }


def _register(counter: dict[str, int], matches: bool, classification: str | None = None) -> None:
    counter["checked"] += 1
    if matches:
        counter["matches"] += 1
    elif classification == "LEGACY_CORRECTION":
        counter["legacy_corrections"] += 1
    elif classification == "HUMAN_INTERVENTION":
        counter["human_interventions"] += 1
    else:
        counter["mismatches"] += 1


def _formula_state(rows: list[tuple], row: int, column: int) -> dict:
    raw = _cell(rows, row, column)
    is_formula = isinstance(raw, str) and raw.lstrip().startswith("=")
    return {
        "is_empty": raw in (None, ""),
        "has_formula": is_formula,
        "formula": raw if is_formula else None,
    }


def _parse_expected_dpp(content: bytes) -> dict:
    values_workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    formulas_workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    if SOURCE_SHEET not in values_workbook.sheetnames or SOURCE_SHEET not in formulas_workbook.sheetnames:
        values_workbook.close()
        formulas_workbook.close()
        raise ValueError("DPP esperado: a aba 'DPP' não foi encontrada.")

    values_sheet = values_workbook[SOURCE_SHEET]
    formulas_sheet = formulas_workbook[SOURCE_SHEET]
    rows = list(values_sheet.iter_rows(values_only=True))
    formula_rows = list(formulas_sheet.iter_rows(values_only=True))
    values_workbook.close()
    formulas_workbook.close()
    if not rows:
        raise ValueError("DPP esperado: a aba 'DPP' está vazia.")

    max_column = max(len(row) for row in rows)
    header_row = _find_header_row(rows, max_column)
    headers = _headers(rows, max_column, header_row)

    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    nec_col = _find_column(headers, "NEC")
    stock_col = _find_stock_source_column(headers)
    explosion_col = _find_explosion_column(headers)
    optional_col = _find_column(headers, "OPC")
    optional_stock_col = _find_column(headers, "STK OP")
    total_stock_col = _find_column(headers, "STK TTL")
    balance_col = _find_column(headers, "SALDO")

    model_start = origin_col + 1
    model_end = check_col - 1
    kit_row = _find_label_row(rows, max_column, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(rows, max_column, header_row, "REAL")
    model_code_row = _detect_model_code_row(rows, header_row, model_start, model_end)

    models: dict[str, dict] = {}
    model_name_by_column: dict[int, str] = {}
    for column in range(model_start, model_end + 1):
        name = _as_text(_cell(rows, header_row, column))
        if not name:
            continue
        key = _normalize(name)
        model_name_by_column[column] = name
        models[key] = {
            "name": name,
            "code": _as_text(_cell(rows, model_code_row, column)) if model_code_row else None,
            "kit_pgd": _number(_cell(rows, kit_row, column), 0.0) if kit_row else 0.0,
            "real": _number(_cell(rows, real_row, column), 0.0) if real_row else 0.0,
        }

    materials: dict[str, dict] = {}
    for excel_row in range(header_row + 1, len(rows) + 1):
        raw_material = _cell(rows, excel_row, material_col)
        key = _material_key(raw_material)
        if not key:
            continue

        consumption_by_model: dict[str, float] = {}
        for column, model_name in model_name_by_column.items():
            value = _number(_cell(rows, excel_row, column), 0.0) or 0.0
            if abs(value) > 1e-12:
                consumption_by_model[model_name] = float(value)

        materials[key] = {
            "material": _as_text(raw_material),
            "material_was_numeric": isinstance(raw_material, (int, float)) and not isinstance(raw_material, bool),
            "description": _as_text(_cell(rows, excel_row, description_col)),
            "um": _normalize_unit(_cell(rows, excel_row, um_col)),
            "group_origin": _as_text(_cell(rows, excel_row, origin_col)),
            "optional_material": _as_text(_cell(rows, excel_row, optional_col)),
            "consumption_by_model": consumption_by_model,
            "stock_sap": _number(_cell(rows, excel_row, stock_col), 0.0) or 0.0,
            "explosion": _number(_cell(rows, excel_row, explosion_col), 0.0) or 0.0,
            "stock_op": _number(_cell(rows, excel_row, optional_stock_col), 0.0) or 0.0,
            "stock_total": _number(_cell(rows, excel_row, total_stock_col), 0.0) or 0.0,
            "nec": _number(_cell(rows, excel_row, nec_col), 0.0) or 0.0,
            "balance": _number(_cell(rows, excel_row, balance_col), 0.0) or 0.0,
            "formula_state": {
                "stock_total": _formula_state(formula_rows, excel_row, total_stock_col),
                "balance": _formula_state(formula_rows, excel_row, balance_col),
            },
            "source": {"sheet": SOURCE_SHEET, "row": excel_row, "reference": f"{SOURCE_SHEET}!A{excel_row}"},
        }

    return {
        "models": models,
        "materials": materials,
        "header_row": header_row,
    }


def _append_sample(
    samples: list[dict],
    *,
    limit: int,
    scope: str,
    key: str,
    field: str,
    generated: object,
    expected: object,
    classification: str,
    reason: str | None = None,
) -> None:
    if len(samples) >= limit:
        return
    samples.append(
        {
            "scope": scope,
            "key": key,
            "field": field,
            "generated": generated,
            "expected": expected,
            "classification": classification,
            "reason": reason,
        }
    )


def _legacy_stock_correction(current: dict, reference: dict) -> bool:
    if not reference.get("material_was_numeric"):
        return False
    if not current.get("stock_source"):
        return False
    if not _same_number(reference.get("stock_sap"), 0.0):
        return False
    return not _same_number(current.get("stock_sap"), reference.get("stock_sap"))


def _missing_formula_legacy(field: str, reference: dict) -> bool:
    if field not in ("stock_total", "balance"):
        return False
    state = reference.get("formula_state", {}).get(field, {})
    return bool(state.get("is_empty"))


def _legacy_numeric_difference(field: str, current: dict, reference: dict) -> bool:
    if _missing_formula_legacy(field, reference):
        return True

    if not _legacy_stock_correction(current, reference):
        return False

    if field == "stock_sap":
        return True

    stock_delta = float(current.get("stock_sap") or 0.0) - float(reference.get("stock_sap") or 0.0)
    if field == "stock_total":
        if not _same_number(current.get("explosion"), reference.get("explosion")):
            return False
        if not _same_number(current.get("stock_op"), reference.get("stock_op")):
            return False
        total_delta = float(current.get("stock_total") or 0.0) - float(reference.get("stock_total") or 0.0)
        return _same_number(total_delta, stock_delta)

    if field == "balance":
        if not _same_number(current.get("nec"), reference.get("nec")):
            return False
        if not _legacy_numeric_difference("stock_total", current, reference):
            return False
        balance_delta = float(current.get("balance") or 0.0) - float(reference.get("balance") or 0.0)
        return _same_number(balance_delta, stock_delta)

    return False


def _human_opc_change(
    material_key: str,
    current: dict,
    reference: dict,
    previous_materials: dict[str, dict] | None,
) -> bool:
    if previous_materials is None:
        return False
    previous = previous_materials.get(material_key) or {}
    previous_optional = previous.get("optional_material")
    expected_optional = reference.get("optional_material")
    generated_optional = current.get("optional_material")

    if _same_optional_materials(expected_optional, previous_optional):
        return False
    return _same_optional_materials(generated_optional, previous_optional)


def _human_opc_numeric_difference(
    field: str,
    material_key: str,
    current: dict,
    reference: dict,
    previous_materials: dict[str, dict] | None,
) -> bool:
    if not _human_opc_change(material_key, current, reference, previous_materials):
        return False

    if field == "stock_op":
        return not _same_number(current.get("stock_op"), reference.get("stock_op"))

    if field == "stock_total":
        if not _same_number(current.get("explosion"), reference.get("explosion")):
            return False
        stock_delta = 0.0
        if _legacy_stock_correction(current, reference):
            stock_delta = float(current.get("stock_sap") or 0.0) - float(reference.get("stock_sap") or 0.0)
        elif not _same_number(current.get("stock_sap"), reference.get("stock_sap")):
            return False
        opc_delta = float(current.get("stock_op") or 0.0) - float(reference.get("stock_op") or 0.0)
        total_delta = float(current.get("stock_total") or 0.0) - float(reference.get("stock_total") or 0.0)
        return _same_number(total_delta, stock_delta + opc_delta)

    if field == "balance":
        if not _same_number(current.get("nec"), reference.get("nec")):
            return False
        if not _human_opc_numeric_difference("stock_total", material_key, current, reference, previous_materials):
            return False
        balance_delta = float(current.get("balance") or 0.0) - float(reference.get("balance") or 0.0)
        total_delta = float(current.get("stock_total") or 0.0) - float(reference.get("stock_total") or 0.0)
        return _same_number(balance_delta, total_delta)

    return False


def _legacy_reason(field: str, reference: dict) -> str:
    if _missing_formula_legacy(field, reference):
        label = "STK TTL" if field == "stock_total" else "SALDO"
        return (
            f"A célula {label} está vazia e sem fórmula no DPP histórico; "
            "o ORION mantém o cálculo determinístico em vez de reproduzir a ausência da fórmula."
        )
    return (
        "O DPP histórico armazenou o código do material como número e registrou STK zero; "
        "o ORION normalizou a chave número × texto e encontrou o estoque no STK SAP."
    )


def _human_reason(field: str) -> str:
    if field == "optional_material":
        return (
            "O OPC do DPP consolidado foi criado, removido ou reassociado depois da base do mês anterior; "
            "o ORION preservou o OPC disponível na base histórica."
        )
    return (
        "A diferença é consequência de uma alteração manual de OPC feita durante o mês; "
        "o campo derivado foi calculado pelo ORION usando o OPC disponível na base histórica."
    )


def _compare(*, generated: dict, expected: dict, previous_materials: dict[str, dict] | None = None) -> dict:
    checks = {
        name: _counter()
        for name in (
            "materials",
            "models",
            "matrix",
            "description",
            "um",
            "origin",
            "optional_material",
            "kit_pgd",
            "real_reference",
            "stock_sap",
            "explosion",
            "stock_op",
            "stock_total",
            "nec",
            "balance",
        )
    }
    mismatches: list[dict] = []
    legacy_corrections: list[dict] = []
    human_interventions: list[dict] = []

    generated_models = {_normalize(item.get("name")): item for item in generated.get("models", []) if item.get("name")}
    expected_models = expected["models"]
    all_model_keys = sorted(set(generated_models) | set(expected_models))
    common_model_keys = sorted(set(generated_models) & set(expected_models))

    for key in all_model_keys:
        present_generated = key in generated_models
        present_expected = key in expected_models
        matches = present_generated and present_expected
        _register(checks["models"], matches)
        if not matches:
            _append_sample(
                mismatches,
                limit=MAX_MISMATCH_SAMPLES,
                scope="model",
                key=generated_models.get(key, expected_models.get(key, {})).get("name", key),
                field="presence",
                generated=present_generated,
                expected=present_expected,
                classification="ORION_DIVERGENCE",
            )

    for key in common_model_keys:
        current = generated_models[key]
        reference = expected_models[key]
        for field in ("kit_pgd", "real"):
            check_name = "kit_pgd" if field == "kit_pgd" else "real_reference"
            matches = _same_number(current.get(field), reference.get(field))
            _register(checks[check_name], matches)
            if not matches:
                _append_sample(
                    mismatches,
                    limit=MAX_MISMATCH_SAMPLES,
                    scope="model",
                    key=reference["name"],
                    field=field,
                    generated=current.get(field),
                    expected=reference.get(field),
                    classification="ORION_DIVERGENCE",
                )

    generated_materials = {
        _material_key(item.get("material")): item
        for item in generated.get("materials", [])
        if _material_key(item.get("material"))
    }
    expected_materials = expected["materials"]
    all_material_keys = sorted(set(generated_materials) | set(expected_materials))
    common_material_keys = sorted(set(generated_materials) & set(expected_materials))

    for key in all_material_keys:
        present_generated = key in generated_materials
        present_expected = key in expected_materials
        matches = present_generated and present_expected
        _register(checks["materials"], matches)
        if not matches:
            item = generated_materials.get(key) or expected_materials.get(key) or {}
            _append_sample(
                mismatches,
                limit=MAX_MISMATCH_SAMPLES,
                scope="material",
                key=item.get("material") or key,
                field="presence",
                generated=present_generated,
                expected=present_expected,
                classification="ORION_DIVERGENCE",
            )

    text_fields = (
        ("description", "description", _same_text),
        ("um", "um", _same_text),
        ("origin", "group_origin", _same_text),
    )
    numeric_fields = (
        ("stock_sap", "stock_sap"),
        ("explosion", "explosion"),
        ("stock_op", "stock_op"),
        ("stock_total", "stock_total"),
        ("nec", "nec"),
        ("balance", "balance"),
    )

    for key in common_material_keys:
        current = generated_materials[key]
        reference = expected_materials[key]
        display_key = reference.get("material") or current.get("material") or key

        for check_name, field, comparator in text_fields:
            matches = comparator(current.get(field), reference.get(field))
            _register(checks[check_name], matches)
            if not matches:
                _append_sample(
                    mismatches,
                    limit=MAX_MISMATCH_SAMPLES,
                    scope="material",
                    key=display_key,
                    field=field,
                    generated=current.get(field),
                    expected=reference.get(field),
                    classification="ORION_DIVERGENCE",
                )

        opc_matches = _same_optional_materials(current.get("optional_material"), reference.get("optional_material"))
        opc_classification = None
        if not opc_matches and _human_opc_change(key, current, reference, previous_materials):
            opc_classification = "HUMAN_INTERVENTION"
        _register(checks["optional_material"], opc_matches, opc_classification)
        if not opc_matches:
            target = human_interventions if opc_classification == "HUMAN_INTERVENTION" else mismatches
            _append_sample(
                target,
                limit=MAX_HUMAN_SAMPLES if opc_classification == "HUMAN_INTERVENTION" else MAX_MISMATCH_SAMPLES,
                scope="material",
                key=display_key,
                field="optional_material",
                generated=current.get("optional_material"),
                expected=reference.get("optional_material"),
                classification=opc_classification or "ORION_DIVERGENCE",
                reason=_human_reason("optional_material") if opc_classification else None,
            )

        for check_name, field in numeric_fields:
            matches = _same_number(current.get(field), reference.get(field))
            classification = None
            if not matches:
                if _missing_formula_legacy(field, reference):
                    classification = "LEGACY_CORRECTION"
                elif _human_opc_numeric_difference(field, key, current, reference, previous_materials):
                    classification = "HUMAN_INTERVENTION"
                elif _legacy_numeric_difference(field, current, reference):
                    classification = "LEGACY_CORRECTION"

            _register(checks[check_name], matches, classification)
            if matches:
                continue

            if classification == "LEGACY_CORRECTION":
                _append_sample(
                    legacy_corrections,
                    limit=MAX_LEGACY_SAMPLES,
                    scope="material",
                    key=display_key,
                    field=field,
                    generated=current.get(field),
                    expected=reference.get(field),
                    classification=classification,
                    reason=_legacy_reason(field, reference),
                )
            elif classification == "HUMAN_INTERVENTION":
                _append_sample(
                    human_interventions,
                    limit=MAX_HUMAN_SAMPLES,
                    scope="material",
                    key=display_key,
                    field=field,
                    generated=current.get(field),
                    expected=reference.get(field),
                    classification=classification,
                    reason=_human_reason(field),
                )
            else:
                _append_sample(
                    mismatches,
                    limit=MAX_MISMATCH_SAMPLES,
                    scope="material",
                    key=display_key,
                    field=field,
                    generated=current.get(field),
                    expected=reference.get(field),
                    classification="ORION_DIVERGENCE",
                )

        generated_matrix = current.get("consumption_by_model", {})
        expected_matrix = reference.get("consumption_by_model", {})
        for model_key in common_model_keys:
            generated_name = generated_models[model_key]["name"]
            expected_name = expected_models[model_key]["name"]
            generated_value = generated_matrix.get(generated_name, 0.0)
            expected_value = expected_matrix.get(expected_name, 0.0)
            matches = _same_number(generated_value, expected_value)
            _register(checks["matrix"], matches)
            if not matches:
                _append_sample(
                    mismatches,
                    limit=MAX_MISMATCH_SAMPLES,
                    scope="matrix",
                    key=f"{display_key} × {expected_name}",
                    field="uso_bom",
                    generated=generated_value,
                    expected=expected_value,
                    classification="ORION_DIVERGENCE",
                )

    critical_checks = (
        "materials",
        "models",
        "matrix",
        "description",
        "um",
        "origin",
        "optional_material",
        "kit_pgd",
        "real_reference",
        "stock_sap",
        "explosion",
        "stock_op",
        "stock_total",
        "nec",
        "balance",
    )
    passed = all(checks[name]["mismatches"] == 0 for name in critical_checks)
    if generated.get("summary", {}).get("pgd_unresolved_positive", 0):
        passed = False

    mismatch_total = sum(item["mismatches"] for item in checks.values())
    legacy_total = sum(item["legacy_corrections"] for item in checks.values())
    human_total = sum(item["human_interventions"] for item in checks.values())

    status = "APROVADO" if passed else "DIVERGENCIAS"
    if passed and legacy_total and human_total:
        status = "APROVADO_COM_INTERVENCOES_E_CORRECOES_LEGADO"
    elif passed and human_total:
        status = "APROVADO_COM_INTERVENCOES_HUMANAS"
    elif passed and legacy_total:
        status = "APROVADO_COM_CORRECOES_LEGADO"

    return {
        "pass": passed,
        "status": status,
        "checks": checks,
        "mismatches": mismatches,
        "legacy_corrections": legacy_corrections,
        "human_interventions": human_interventions,
        "mismatch_samples_truncated": mismatch_total > len(mismatches),
        "legacy_samples_truncated": legacy_total > len(legacy_corrections),
        "human_samples_truncated": human_total > len(human_interventions),
        "mismatch_total": mismatch_total,
        "legacy_corrections_total": legacy_total,
        "human_interventions_total": human_total,
    }


async def test_monthly_dpp_reconstruction(
    *,
    base_dpp: UploadFile,
    expected_dpp: UploadFile,
    wiu: UploadFile,
    explosion: UploadFile,
    stock: UploadFile,
    pgd: UploadFile,
    reference_month: str,
    open_orders: UploadFile | None = None,
) -> dict:
    for file, label in (
        (base_dpp, "DPP mês anterior"),
        (expected_dpp, "DPP consolidado esperado"),
        (wiu, "WIU"),
        (explosion, "Explosão"),
        (stock, "STK SAP"),
        (pgd, "PGD"),
    ):
        _validate_upload(file, label)

    base_content = await base_dpp.read()
    if not base_content:
        raise ValueError("O DPP do mês anterior está vazio.")
    previous_materials, _, _ = parse_previous_dpp(base_content)
    await base_dpp.seek(0)

    expected_content = await expected_dpp.read()
    if not expected_content:
        raise ValueError("O DPP consolidado esperado está vazio.")
    expected = _parse_expected_dpp(expected_content)

    generated_initial = await generate_monthly_dpp(
        base_dpp=base_dpp,
        wiu=wiu,
        explosion=explosion,
        stock=stock,
        pgd=pgd,
        reference_month=reference_month,
        open_orders=open_orders,
    )

    expected_models = expected["models"]
    generated_models = {_normalize(item.get("name")): item for item in generated_initial.get("models", [])}
    reference_real = {
        generated_models[key]["name"]: float(item.get("real") or 0.0)
        for key, item in expected_models.items()
        if key in generated_models
    }
    generated = recalculate_monthly_scenario(
        scenario_id=generated_initial["scenario_id"],
        real_by_model=reference_real,
    )

    comparison = _compare(
        generated=generated,
        expected=expected,
        previous_materials=previous_materials,
    )
    return {
        "mode": "monthly_dpp_reconstruction_test",
        "reference_month": reference_month,
        "scenario_id": generated["scenario_id"],
        "pass": comparison["pass"],
        "status": comparison["status"],
        "note": (
            "O REAL do DPP consolidado esperado é aplicado ao cenário gerado somente para validar "
            "a reconstrução e o motor de cálculo. Este teste não valida o solver automático de REAL."
        ),
        "requirements": {
            "historical_base_required": True,
            "reason": (
                "Materiais e OPCs são acumulativos; por isso a validação integral exige o DPP do mês anterior. "
                "WIU, Explosão, STK e PGD sozinhos não reproduzem a base histórica completa."
            ),
        },
        "summary": {
            "generated_materials": len(generated.get("materials", [])),
            "expected_materials": len(expected["materials"]),
            "generated_models": len(generated.get("models", [])),
            "expected_models": len(expected["models"]),
            "pgd_unresolved_positive": generated.get("summary", {}).get("pgd_unresolved_positive", 0),
            "reference_real_models_applied": len(reference_real),
            "legacy_corrections_total": comparison["legacy_corrections_total"],
            "human_interventions_total": comparison["human_interventions_total"],
            "orion_mismatches_total": comparison["mismatch_total"],
        },
        "checks": comparison["checks"],
        "mismatches": comparison["mismatches"],
        "legacy_corrections": comparison["legacy_corrections"],
        "human_interventions": comparison["human_interventions"],
        "mismatch_samples_truncated": comparison["mismatch_samples_truncated"],
        "legacy_samples_truncated": comparison["legacy_samples_truncated"],
        "human_samples_truncated": comparison["human_samples_truncated"],
        "sources": {
            "base_dpp": base_dpp.filename,
            "expected_dpp": expected_dpp.filename,
            "wiu": wiu.filename,
            "explosion": explosion.filename,
            "stock": stock.filename,
            "pgd": pgd.filename,
            "open": open_orders.filename if open_orders is not None else None,
        },
    }
