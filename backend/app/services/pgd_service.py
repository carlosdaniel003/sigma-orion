from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import math
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip().lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _number(value: object, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else default
    text = str(value).strip().replace(" ", "")
    if not text:
        return default
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return default
    return number if math.isfinite(number) else default


def _parse_reference_month(reference_month: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", reference_month.strip())
    if not match:
        raise ValueError("Mês de referência inválido. Use o formato YYYY-MM.")
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        raise ValueError("Mês de referência inválido.")
    return year, month


def _month_from_cell(value: object, epoch) -> tuple[int, int] | None:
    current = value
    if isinstance(current, (int, float)) and not isinstance(current, bool):
        try:
            current = from_excel(current, epoch)
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(current, datetime):
        return current.year, current.month
    if isinstance(current, date):
        return current.year, current.month
    if isinstance(current, str):
        text = current.strip()
        for pattern in (r"(\d{2})/(\d{2})/(\d{4})", r"(\d{4})-(\d{2})-(\d{2})"):
            match = re.fullmatch(pattern, text)
            if match:
                if pattern.startswith("(\\d{2})"):
                    return int(match.group(3)), int(match.group(2))
                return int(match.group(1)), int(match.group(2))
    return None


def _looks_like_product_code(value: object) -> bool:
    text = _text(value)
    if not text:
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/]*-[A-Za-z0-9._/-]+", text))


def parse_pgd(content: bytes, reference_month: str) -> tuple[list[dict], dict]:
    target = _parse_reference_month(reference_month)
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)

    models: list[dict] = []
    sheets_scanned = 0
    negative_clamped = 0
    sheets_without_month: list[str] = []

    for sheet in workbook.worksheets:
        if not _normalize(sheet.title).startswith("pgd "):
            continue
        sheets_scanned += 1
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        production_header_row = None
        for row_number in range(1, min(12, len(rows)) + 1):
            row = rows[row_number - 1]
            if len(row) >= 2 and _normalize(row[1]) == "production":
                production_header_row = row_number
                break

        if production_header_row is None:
            continue

        header = rows[production_header_row - 1]
        target_col = None
        for column, value in enumerate(header, start=1):
            if _month_from_cell(value, workbook.epoch) == target:
                target_col = column
                break

        if target_col is None:
            sheets_without_month.append(sheet.title)
            continue

        row_number = production_header_row + 1
        while row_number <= len(rows):
            row = rows[row_number - 1]
            code = _text(row[0]) if len(row) >= 1 else None
            model = _text(row[1]) if len(row) >= 2 else None
            if not code or not model or not _looks_like_product_code(code):
                row_number += 1
                continue

            kit_row = None
            for candidate in range(row_number + 1, min(row_number + 10, len(rows)) + 1):
                candidate_row = rows[candidate - 1]
                label = candidate_row[1] if len(candidate_row) >= 2 else None
                if _normalize(label) == "kit disponivel":
                    kit_row = candidate
                    break

            if kit_row is None:
                row_number += 1
                continue

            kit_values = rows[kit_row - 1]
            raw_value = kit_values[target_col - 1] if target_col <= len(kit_values) else None
            raw_kit = _number(raw_value, 0.0)
            kit_pgd = max(raw_kit, 0.0)
            if raw_kit < 0:
                negative_clamped += 1

            models.append(
                {
                    "name": model,
                    "code": code,
                    "kit_pgd": kit_pgd,
                    "kit_pgd_raw": raw_kit,
                    "source": {
                        "sheet": sheet.title,
                        "model_row": row_number,
                        "kit_row": kit_row,
                        "cell": f"{get_column_letter(target_col)}{kit_row}",
                        "reference": f"{sheet.title}!{get_column_letter(target_col)}{kit_row}",
                    },
                }
            )
            row_number = kit_row + 1

    workbook.close()
    return models, {
        "reference_month": reference_month,
        "sheets_scanned": sheets_scanned,
        "models_found": len(models),
        "positive_models": sum(1 for item in models if item["kit_pgd"] > 0),
        "negative_values_clamped_to_zero": negative_clamped,
        "sheets_without_reference_month": sheets_without_month,
    }
