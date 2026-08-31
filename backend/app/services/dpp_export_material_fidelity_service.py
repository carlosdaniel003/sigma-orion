from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from app.services.dpp_consolidation_service import _material_key
from app.services.dpp_monthly_base import SOURCE_SHEET, _find_column, _find_header_row, _headers


def _scenario_material_keys(scenario: dict) -> list[str]:
    materials = list(scenario.get("materials") or [])
    keys = [_material_key(material.get("material")) for material in materials]
    invalid = [index for index, key in enumerate(keys) if not key]
    if invalid:
        raise ValueError(
            "Falha de consistência do Excel ORION: o Dashboard contém material sem código válido "
            f"em {len(invalid)} item(ns)."
        )
    if len(set(keys)) != len(keys):
        raise ValueError(
            "Falha de consistência do Excel ORION: o Dashboard contém materiais duplicados após normalização."
        )
    return [str(key) for key in keys]


def _sheet_layout(sheet) -> tuple[int, int]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Falha de consistência do Excel ORION: a aba DPP está vazia.")
    header_row = _find_header_row(rows)
    headers = _headers(rows, header_row)
    material_col = _find_column(headers, "Material")
    return header_row, material_col


def _material_rows_and_keys(sheet, *, header_row: int, material_col: int) -> tuple[list[int], list[str]]:
    rows: list[int] = []
    keys: list[str] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        key = _material_key(sheet.cell(row, material_col).value)
        if not key:
            continue
        rows.append(row)
        keys.append(str(key))
    return rows, keys


def _expand_filter_and_tables(sheet, *, header_row: int, material_col: int, last_material_row: int) -> dict:
    audit = {
        "auto_filter": None,
        "tables": {},
    }

    filter_ref = getattr(sheet.auto_filter, "ref", None)
    if filter_ref:
        min_col, min_row, max_col, _max_row = range_boundaries(filter_ref)
        if min_row <= header_row and min_col <= material_col <= max_col:
            new_ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{last_material_row}"
            )
            sheet.auto_filter.ref = new_ref
            audit["auto_filter"] = new_ref

    for table in sheet.tables.values():
        min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        if min_row <= header_row and min_col <= material_col <= max_col:
            new_ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{last_material_row}"
            )
            table.ref = new_ref
            if getattr(table, "autoFilter", None) is not None:
                table.autoFilter.ref = new_ref
            audit["tables"][table.name] = new_ref

    return audit


def _assert_material_contract(sheet, scenario: dict) -> dict:
    expected_keys = _scenario_material_keys(scenario)
    expected_count = len(expected_keys)
    expected_set = set(expected_keys)
    header_row, material_col = _sheet_layout(sheet)
    material_rows, exported_keys = _material_rows_and_keys(
        sheet,
        header_row=header_row,
        material_col=material_col,
    )

    exported_count = len(exported_keys)
    exported_set = set(exported_keys)
    missing = sorted(expected_set - exported_set)
    extra = sorted(exported_set - expected_set)

    if exported_count != expected_count or missing or extra:
        raise ValueError(
            "Falha de consistência do Excel ORION: quantidade/lista de materiais diverge do Dashboard "
            f"(Dashboard={expected_count}, Excel={exported_count}, ausentes={len(missing)}, excedentes={len(extra)})."
        )

    first_material_row = header_row + 1
    last_material_row = header_row + expected_count
    expected_rows = list(range(first_material_row, last_material_row + 1))
    if material_rows != expected_rows:
        raise ValueError(
            "Falha de consistência do Excel ORION: os materiais existem, mas não formam uma faixa contínua "
            f"de {first_material_row} a {last_material_row}."
        )

    if len(exported_set) != exported_count:
        raise ValueError(
            "Falha de consistência do Excel ORION: existem materiais duplicados na coluna Material do arquivo exportado."
        )

    return {
        "dashboard_materials": expected_count,
        "excel_materials": exported_count,
        "header_row": header_row,
        "material_col": material_col,
        "first_material_row": first_material_row,
        "last_material_row": last_material_row,
    }


def finalize_and_audit_material_extent(content: bytes, scenario: dict) -> tuple[bytes, dict]:
    """Garante que Dashboard, coluna Material e estruturas de filtro/tabela tenham a mesma extensão."""
    if not content:
        raise ValueError("Falha de consistência do Excel ORION: conteúdo XLSX vazio.")

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError("Falha de consistência do Excel ORION: a aba DPP não foi encontrada.")
        sheet = workbook[SOURCE_SHEET]
        audit = _assert_material_contract(sheet, scenario)
        structure = _expand_filter_and_tables(
            sheet,
            header_row=audit["header_row"],
            material_col=audit["material_col"],
            last_material_row=audit["last_material_row"],
        )
        output = BytesIO()
        workbook.save(output)
        finalized = output.getvalue()
    finally:
        workbook.close()

    audit_workbook = load_workbook(BytesIO(finalized), data_only=False, read_only=False)
    try:
        sheet = audit_workbook[SOURCE_SHEET]
        final_audit = _assert_material_contract(sheet, scenario)

        filter_ref = getattr(sheet.auto_filter, "ref", None)
        if structure["auto_filter"] and filter_ref != structure["auto_filter"]:
            raise ValueError(
                "Falha de consistência do Excel ORION: o AutoFilter não alcança todos os materiais após salvar."
            )

        for table_name, expected_ref in structure["tables"].items():
            table = sheet.tables.get(table_name)
            if table is None or table.ref != expected_ref:
                raise ValueError(
                    f"Falha de consistência do Excel ORION: a tabela '{table_name}' não alcança todos os materiais."
                )

        final_audit["auto_filter_ref"] = filter_ref
        final_audit["table_refs"] = {
            table_name: sheet.tables[table_name].ref
            for table_name in structure["tables"]
            if table_name in sheet.tables
        }
        final_audit["status"] = "DASHBOARD_EXCEL_MATERIALS_OK"
        return finalized, final_audit
    finally:
        audit_workbook.close()
