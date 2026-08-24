from __future__ import annotations

from io import BytesIO
import math
from pathlib import Path
import re
import unicodedata

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SUPPORTED_DPP_EXTENSIONS = {".xlsx", ".xlsm"}
VALIDATION_REL_TOL = 1e-9
VALIDATION_ABS_TOL = 1e-4


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip().lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def _number(value: object, default: float | None = 0.0) -> float | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else default

    text = str(value).strip().replace(" ", "")
    if not text:
        return default

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return default
    return number if math.isfinite(number) else default


def _as_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _matches(left: float | None, right: float | None) -> bool | None:
    if left is None or right is None:
        return None
    return math.isclose(left, right, rel_tol=VALIDATION_REL_TOL, abs_tol=VALIDATION_ABS_TOL)


def _cell(rows: list[tuple], row: int | None, column: int) -> object:
    if row is None or row < 1 or column < 1 or row > len(rows):
        return None
    current = rows[row - 1]
    if column > len(current):
        return None
    return current[column - 1]


def _find_header_row(rows: list[tuple], max_column: int) -> int:
    max_scan = min(len(rows), 40)
    for row in range(1, max_scan + 1):
        values = {_normalize(_cell(rows, row, column)) for column in range(1, min(max_column, 12) + 1)}
        if {"material", "descricao", "um", "grupo origem"}.issubset(values):
            return row
    raise ValueError("Não foi possível localizar a linha de cabeçalho do DPP.")


def _find_label_row(
    rows: list[tuple],
    max_column: int,
    header_row: int,
    label: str,
    contains: bool = False,
) -> int | None:
    target = _normalize(label)
    for row in range(1, header_row):
        for column in range(1, min(max_column, 8) + 1):
            current = _normalize(_cell(rows, row, column))
            if (contains and target in current) or (not contains and current == target):
                return row
    return None


def _headers(rows: list[tuple], max_column: int, header_row: int) -> dict[int, str]:
    result = {}
    for column in range(1, max_column + 1):
        value = _cell(rows, header_row, column)
        if value not in (None, ""):
            result[column] = str(value).strip()
    return result


def _find_column(headers: dict[int, str], name: str) -> int:
    target = _normalize(name)
    for column, value in headers.items():
        if _normalize(value) == target:
            return column
    raise ValueError(f"Coluna obrigatória '{name}' não encontrada no DPP.")


def _find_stock_source_column(headers: dict[int, str]) -> int:
    for column, value in headers.items():
        normalized = _normalize(value)
        if normalized.startswith("stk ") and normalized not in {"stk op", "stk ttl"}:
            return column
    raise ValueError("Coluna de estoque base (STK <data>) não encontrada no DPP.")


def _find_explosion_column(headers: dict[int, str]) -> int:
    for column, value in headers.items():
        if _normalize(value).startswith("explosao"):
            return column
    raise ValueError("Coluna de EXPLOSÃO não encontrada no DPP.")


def _find_comments_column(headers: dict[int, str]) -> int | None:
    accepted = {"coments", "comments", "comentarios", "comentario"}
    for column, value in headers.items():
        if _normalize(value) in accepted:
            return column
    return None


def _looks_like_model_code(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/]*-[A-Za-z0-9._/-]+", text))


def _detect_model_code_row(
    rows: list[tuple],
    header_row: int,
    model_start: int,
    model_end: int,
) -> int | None:
    candidates = range(max(1, header_row - 4), header_row)
    best_row = None
    best_score = 0
    for row in candidates:
        score = sum(
            1
            for column in range(model_start, model_end + 1)
            if _looks_like_model_code(_cell(rows, row, column))
        )
        if score > best_score:
            best_row = row
            best_score = score
    return best_row


def _status_for_model(rows: list[tuple], kit_row: int | None, column: int) -> str | None:
    if not kit_row or kit_row <= 1:
        return None
    value = _as_text(_cell(rows, kit_row - 1, column))
    if value and _normalize(value) in {"ok", "ng"}:
        return value.upper()
    return value


def _validation_counter() -> dict[str, int]:
    return {"checked": 0, "matches": 0, "mismatches": 0}


def _register_validation(counter: dict[str, int], result: bool | None) -> None:
    if result is None:
        return
    counter["checked"] += 1
    if result:
        counter["matches"] += 1
    else:
        counter["mismatches"] += 1


async def analyze_dpp_file(file: UploadFile, divergence_limit: int = 100) -> dict:
    filename = file.filename or "dpp.xlsx"
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_DPP_EXTENSIONS:
        raise ValueError("Envie um DPP em formato .xlsx ou .xlsm.")

    content = await file.read()
    if not content:
        raise ValueError("O arquivo do DPP está vazio.")

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if "DPP" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("A aba 'DPP' não foi encontrada no arquivo enviado.")

    sheet = workbook["DPP"]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValueError("A aba 'DPP' está vazia.")

    max_column = max(len(row) for row in rows)
    header_row = _find_header_row(rows, max_column)
    headers = _headers(rows, max_column, header_row)

    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    wiu_col = _find_column(headers, "WIU")
    nec_col = _find_column(headers, "NEC")
    stock_col = _find_stock_source_column(headers)
    explosion_col = _find_explosion_column(headers)
    optional_col = _find_column(headers, "OPC")
    optional_stock_col = _find_column(headers, "STK OP")
    total_stock_col = _find_column(headers, "STK TTL")
    balance_col = _find_column(headers, "SALDO")
    price_col = _find_column(headers, "Preço")
    amount_col = _find_column(headers, "Amount")
    comments_col = _find_comments_column(headers)

    model_start = origin_col + 1
    model_end = check_col - 1
    if model_end < model_start:
        raise ValueError("Não foi possível identificar o bloco de modelos do DPP.")

    real_row = _find_label_row(rows, max_column, header_row, "REAL")
    kit_row = _find_label_row(rows, max_column, header_row, "KIT Disponivel PGD", contains=True)
    if real_row is None:
        raise ValueError("A linha REAL não foi encontrada acima do cabeçalho do DPP.")

    model_code_row = _detect_model_code_row(rows, header_row, model_start, model_end)

    models = []
    real_by_column: dict[int, float] = {}
    model_name_by_column: dict[int, str] = {}
    for column in range(model_start, model_end + 1):
        name = _as_text(_cell(rows, header_row, column))
        if not name:
            continue
        real = _number(_cell(rows, real_row, column), 0.0) or 0.0
        kit = _number(_cell(rows, kit_row, column), 0.0) if kit_row else None
        real_by_column[column] = real
        model_name_by_column[column] = name
        models.append(
            {
                "column": get_column_letter(column),
                "name": name,
                "code": _as_text(_cell(rows, model_code_row, column)) if model_code_row else None,
                "kit_pgd": kit,
                "real": real,
                "difference_real_vs_kit": real - kit if kit is not None else None,
                "status": _status_for_model(rows, kit_row, column),
            }
        )

    validation = {
        "nec": _validation_counter(),
        "stk_total": _validation_counter(),
        "saldo": _validation_counter(),
        "amount": _validation_counter(),
    }

    total_materials = 0
    materials_to_investigate = 0
    optional_material_links = 0
    check_wiu_equal = 0
    check_wiu_different = 0
    materials: list[dict] = []
    divergences: list[dict] = []

    for row in range(header_row + 1, len(rows) + 1):
        material = _as_text(_cell(rows, row, material_col))
        if not material:
            continue

        total_materials += 1
        necessity_python = 0.0
        used_models: list[str] = []
        for column in range(model_start, model_end + 1):
            consumption = _number(_cell(rows, row, column), 0.0) or 0.0
            if consumption == 0:
                continue
            necessity_python += real_by_column.get(column, 0.0) * consumption
            model_name = model_name_by_column.get(column)
            if model_name:
                used_models.append(model_name)

        stock = _number(_cell(rows, row, stock_col), 0.0) or 0.0
        explosion = _number(_cell(rows, row, explosion_col), 0.0) or 0.0
        optional_material = _as_text(_cell(rows, row, optional_col))
        optional_stock = _number(_cell(rows, row, optional_stock_col), 0.0) or 0.0
        stock_total_python = stock + explosion + optional_stock
        balance_python = stock_total_python - necessity_python
        price = _number(_cell(rows, row, price_col), None)
        amount_python = price * balance_python if price is not None else None

        nec_excel = _number(_cell(rows, row, nec_col), None)
        stock_total_excel = _number(_cell(rows, row, total_stock_col), None)
        balance_excel = _number(_cell(rows, row, balance_col), None)
        amount_excel = _number(_cell(rows, row, amount_col), None)

        nec_match = _matches(necessity_python, nec_excel)
        stock_total_match = _matches(stock_total_python, stock_total_excel)
        balance_match = _matches(balance_python, balance_excel)
        amount_match = _matches(amount_python, amount_excel)
        _register_validation(validation["nec"], nec_match)
        _register_validation(validation["stk_total"], stock_total_match)
        _register_validation(validation["saldo"], balance_match)
        _register_validation(validation["amount"], amount_match)

        check_value = _as_text(_cell(rows, row, check_col))
        wiu_value = _as_text(_cell(rows, row, wiu_col))
        if check_value or wiu_value:
            if _normalize(check_value) == _normalize(wiu_value):
                check_wiu_equal += 1
            else:
                check_wiu_different += 1

        if optional_material:
            optional_material_links += 1

        status = "INVESTIGAR" if balance_python < -VALIDATION_ABS_TOL else "OK"
        validation_ok = all(
            result is not False
            for result in (nec_match, stock_total_match, balance_match, amount_match)
        )
        materials.append(
            {
                "material": material,
                "description": _as_text(_cell(rows, row, description_col)),
                "um": _as_text(_cell(rows, row, um_col)),
                "group_origin": _as_text(_cell(rows, row, origin_col)),
                "used_models": used_models,
                "optional_material": optional_material,
                "nec": necessity_python,
                "stock_total": stock_total_python,
                "balance": balance_python,
                "status": status,
                "validation_ok": validation_ok,
            }
        )

        if status == "INVESTIGAR":
            materials_to_investigate += 1
            divergences.append(
                {
                    "material": material,
                    "description": _as_text(_cell(rows, row, description_col)),
                    "um": _as_text(_cell(rows, row, um_col)),
                    "group_origin": _as_text(_cell(rows, row, origin_col)),
                    "used_models": used_models,
                    "check": check_value,
                    "wiu": wiu_value,
                    "optional_material": optional_material,
                    "comments": _as_text(_cell(rows, row, comments_col)) if comments_col else None,
                    "excel": {
                        "nec": nec_excel,
                        "stock": stock,
                        "explosion": explosion,
                        "optional_stock": optional_stock,
                        "stock_total": stock_total_excel,
                        "balance": balance_excel,
                        "price": price,
                        "amount": amount_excel,
                    },
                    "python": {
                        "nec": necessity_python,
                        "stock_total": stock_total_python,
                        "balance": balance_python,
                        "amount": amount_python,
                        "status": "INVESTIGAR",
                    },
                    "validation": {
                        "nec": nec_match,
                        "stock_total": stock_total_match,
                        "balance": balance_match,
                        "amount": amount_match,
                    },
                }
            )

    materials.sort(key=lambda item: (item["status"] != "INVESTIGAR", item["material"]))
    divergences.sort(key=lambda item: item["python"]["balance"])
    safe_limit = min(max(divergence_limit, 1), 500)

    return {
        "filename": filename,
        "sheet": "DPP",
        "scope": (
            "Análise determinística somente do DPP já preenchido. WIU, explosão, PGD e BOM "
            "são tratados nesta etapa apenas como valores já consolidados no próprio DPP."
        ),
        "structure": {
            "header_row": header_row,
            "kit_pgd_row": kit_row,
            "real_row": real_row,
            "model_code_row": model_code_row,
            "model_start_column": get_column_letter(model_start),
            "model_end_column": get_column_letter(model_end),
            "model_count": len(models),
            "stock_source_header": headers[stock_col],
            "explosion_header": headers[explosion_col],
        },
        "models": models,
        "materials": materials,
        "summary": {
            "total_materials": total_materials,
            "materials_to_investigate": materials_to_investigate,
            "optional_material_links": optional_material_links,
            "check_wiu_equal": check_wiu_equal,
            "check_wiu_different": check_wiu_different,
            "validation": validation,
        },
        "rules_applied": [
            "NEC = soma(REAL do modelo × consumo do material no modelo)",
            "STK TTL = STK base + EXPLOSÃO + STK OP",
            "SALDO = STK TTL - NEC",
            "Amount = Preço × SALDO quando Preço estiver disponível",
            "SALDO negativo = INVESTIGAR; não implica compra automática",
            "OPC representa código de material opcional e STK OP seu estoque considerado",
        ],
        "divergences": divergences[:safe_limit],
        "divergence_limit": safe_limit,
    }