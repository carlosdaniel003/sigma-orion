from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from app.services.dpp_consolidation_service import _material_key
from app.services.dpp_monthly_base import (
    SOURCE_SHEET,
    _find_column,
    _find_header_row,
    _find_label_row,
    _headers,
    _normalize,
)

FORMULA_TOLERANCE = 1e-4
MONTH_NAMES_PT = {
    "01": "JANEIRO",
    "02": "FEVEREIRO",
    "03": "MARÇO",
    "04": "ABRIL",
    "05": "MAIO",
    "06": "JUNHO",
    "07": "JULHO",
    "08": "AGOSTO",
    "09": "SETEMBRO",
    "10": "OUTUBRO",
    "11": "NOVEMBRO",
    "12": "DEZEMBRO",
}


def _number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _excel_number(value: object) -> str:
    number = _number(value)
    if number.is_integer():
        return str(int(number))
    return format(number, ".15g")


def _formula_text(value: object) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _find_any_column(headers: dict[int, str], *names: str) -> int | None:
    targets = {_normalize(name) for name in names}
    for column, header in headers.items():
        if _normalize(header) in targets:
            return column
    return None


def _find_stock_source_column(headers: dict[int, str]) -> int | None:
    for column, header in headers.items():
        normalized = _normalize(header)
        if normalized.startswith("stk ") and normalized not in {"stk op", "stk opc", "stk opcs", "stk ttl", "stk total"}:
            return column
    return None


def _find_explosion_column(headers: dict[int, str]) -> int | None:
    for column, header in headers.items():
        if _normalize(header).startswith("explosao"):
            return column
    return None


def _month_label(reference_month: str) -> str:
    parts = str(reference_month or "").split("-")
    month = parts[1] if len(parts) >= 2 else ""
    name = MONTH_NAMES_PT.get(month)
    if not name:
        raise ValueError("Falha de consistência do Excel ORION: mês de referência inválido.")
    return f"KIT Disponivel PGD ({name})"


def _set_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    if hasattr(calculation, "calcMode"):
        calculation.calcMode = "auto"
    if hasattr(calculation, "fullCalcOnLoad"):
        calculation.fullCalcOnLoad = True
    if hasattr(calculation, "forceFullCalc"):
        calculation.forceFullCalc = True


def _layout(sheet) -> dict:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Falha de consistência do Excel ORION: a aba DPP está vazia.")

    header_row = _find_header_row(rows)
    headers = _headers(rows, header_row)
    material_col = _find_column(headers, "Material")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    optional_col = _find_column(headers, "OPC", required=False)
    nec_col = _find_any_column(headers, "NEC", "NECESSIDADE")
    stock_op_col = _find_any_column(headers, "STK OP", "STK OPC", "STK OPCS")
    stock_total_col = _find_any_column(headers, "STK TTL", "STK TOTAL")
    balance_col = _find_any_column(headers, "SALDO", "BALANCE")
    price_col = _find_column(headers, "Preço", required=False)
    amount_col = _find_column(headers, "Amount", required=False)
    stock_source_col = _find_stock_source_column(headers)
    explosion_col = _find_explosion_column(headers)

    kit_row = _find_label_row(rows, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(rows, header_row, "REAL")
    if kit_row is None or real_row is None:
        raise ValueError("Falha de consistência do Excel ORION: linhas KIT Disponível PGD/REAL não encontradas.")

    label_cell = None
    for row in range(1, header_row):
        for column in range(1, min(8, sheet.max_column) + 1):
            cell = sheet.cell(row, column)
            if "kit disponivel pgd" in _normalize(cell.value):
                label_cell = cell
                break
        if label_cell is not None:
            break
    if label_cell is None:
        raise ValueError("Falha de consistência do Excel ORION: rótulo KIT Disponível PGD não encontrado.")

    model_start = origin_col + 1
    model_end = check_col - 1
    gap_row = real_row + 1
    if model_end < model_start or gap_row >= header_row:
        raise ValueError("Falha de consistência do Excel ORION: estrutura de modelos/linha de diferença inválida.")

    required = {
        "nec": nec_col,
        "stock_source": stock_source_col,
        "explosion": explosion_col,
        "stock_op": stock_op_col,
        "stock_total": stock_total_col,
        "balance": balance_col,
    }
    missing = [name for name, column in required.items() if column is None]
    if missing:
        raise ValueError(
            "Falha de consistência do Excel ORION: colunas necessárias às fórmulas DPP não encontradas: "
            + ", ".join(missing)
        )

    return {
        "header_row": header_row,
        "headers": headers,
        "material_col": material_col,
        "origin_col": origin_col,
        "check_col": check_col,
        "optional_col": optional_col,
        "nec_col": nec_col,
        "stock_source_col": stock_source_col,
        "explosion_col": explosion_col,
        "stock_op_col": stock_op_col,
        "stock_total_col": stock_total_col,
        "balance_col": balance_col,
        "price_col": price_col,
        "amount_col": amount_col,
        "kit_row": kit_row,
        "real_row": real_row,
        "gap_row": gap_row,
        "model_start": model_start,
        "model_end": model_end,
        "label_cell": label_cell,
    }


def _scenario_materials(scenario: dict) -> dict[str, dict]:
    return {
        _material_key(material.get("material")): material
        for material in scenario.get("materials") or []
        if _material_key(material.get("material"))
    }


def _expected_formulas(layout: dict, row: int, material: dict) -> dict[str, str | None]:
    model_start = get_column_letter(layout["model_start"])
    model_end = get_column_letter(layout["model_end"])
    check = (
        f'=_xlfn.TEXTJOIN("// ", TRUE, _xlfn._xlws.FILTER(${model_start}${layout["header_row"]}:'
        f'${model_end}${layout["header_row"]}, {model_start}{row}:{model_end}{row}>0, ""))'
    )
    nec = (
        f"=SUMPRODUCT(${model_start}${layout['real_row']}:${model_end}${layout['real_row']},"
        f"{model_start}{row}:{model_end}{row})"
    )

    stock_op = None
    sources = list(material.get("stock_op_sources") or [])
    if sources:
        if len(sources) == 1 and layout["optional_col"] is not None:
            optional = f"{get_column_letter(layout['optional_col'])}{row}"
            stock_op = f"=VLOOKUP({optional},CONSOLIDADO!$A:$F,6,0)"
        elif len(sources) > 1:
            stock_op = "=" + "+".join(_excel_number(source.get("value")) for source in sources)

    stock_total = (
        "="
        + f"{get_column_letter(layout['stock_source_col'])}{row}"
        + "+"
        + f"{get_column_letter(layout['explosion_col'])}{row}"
        + "+"
        + f"{get_column_letter(layout['stock_op_col'])}{row}"
    )
    balance = (
        f"={get_column_letter(layout['stock_total_col'])}{row}-"
        f"{get_column_letter(layout['nec_col'])}{row}"
    )
    amount = None
    if layout["price_col"] is not None and layout["amount_col"] is not None:
        amount = (
            f"={get_column_letter(layout['price_col'])}{row}*"
            f"{get_column_letter(layout['balance_col'])}{row}"
        )

    return {
        "check": check,
        "nec": nec,
        "stock_op": stock_op,
        "stock_total": stock_total,
        "balance": balance,
        "amount": amount,
    }


def _apply_summary_formulas(workbook, layout: dict) -> None:
    if "Resumo de Análise" not in workbook.sheetnames or "DPP BALANCE" not in workbook.sheetnames:
        return

    sheet = workbook["Resumo de Análise"]
    model_start = get_column_letter(layout["model_start"])
    model_end = get_column_letter(layout["model_end"])
    model_name_row = layout["real_row"] - 1
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 2).value in (None, ""):
            continue
        sheet.cell(row, 3).value = (
            f'=XLOOKUP(B{row},\'DPP BALANCE\'!${model_start}${model_name_row}:${model_end}${model_name_row},'
            f'\'DPP BALANCE\'!${model_start}${layout["kit_row"]}:${model_end}${layout["kit_row"]},"ND")'
        )
        sheet.cell(row, 4).value = (
            f'=XLOOKUP(B{row},DPP!${model_start}${model_name_row}:${model_end}${model_name_row},'
            f'DPP!${model_start}${layout["real_row"]}:${model_end}${layout["real_row"]},"ND")'
        )
        sheet.cell(row, 5).value = f"=D{row}-C{row}"


def _apply_formula_contract(workbook, scenario: dict) -> None:
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError("Falha de consistência do Excel ORION: a aba DPP não foi encontrada.")

    sheet = workbook[SOURCE_SHEET]
    layout = _layout(sheet)
    layout["label_cell"].value = _month_label(scenario.get("reference_month") or "")
    by_key = _scenario_materials(scenario)

    for column in range(layout["model_start"], layout["model_end"] + 1):
        kit = _number(sheet.cell(layout["kit_row"], column).value)
        real = _number(sheet.cell(layout["real_row"], column).value)
        gap = sheet.cell(layout["gap_row"], column)
        if abs(real - kit) > FORMULA_TOLERANCE:
            letter = get_column_letter(column)
            gap.value = f"={letter}{layout['real_row']}-{letter}{layout['kit_row']}"
        else:
            gap.value = None

    for row in range(layout["header_row"] + 1, sheet.max_row + 1):
        key = _material_key(sheet.cell(row, layout["material_col"]).value)
        if not key:
            continue
        material = by_key.get(key)
        if material is None:
            raise ValueError(
                f"Falha de consistência do Excel ORION: material da linha {row} não existe no Cenário ORION."
            )
        formulas = _expected_formulas(layout, row, material)
        check_cell = sheet.cell(row, layout["check_col"])
        check_cell.value = ArrayFormula(ref=check_cell.coordinate, text=formulas["check"])
        sheet.cell(row, layout["nec_col"]).value = formulas["nec"]
        if formulas["stock_op"] is not None:
            sheet.cell(row, layout["stock_op_col"]).value = formulas["stock_op"]
        sheet.cell(row, layout["stock_total_col"]).value = formulas["stock_total"]
        sheet.cell(row, layout["balance_col"]).value = formulas["balance"]
        if formulas["amount"] is not None:
            sheet.cell(row, layout["amount_col"]).value = formulas["amount"]

    _apply_summary_formulas(workbook, layout)
    _set_recalculation(workbook)


def _audit_formula_contract(workbook, scenario: dict) -> None:
    sheet = workbook[SOURCE_SHEET]
    layout = _layout(sheet)
    expected_label = _month_label(scenario.get("reference_month") or "")
    if layout["label_cell"].value != expected_label:
        raise ValueError("Falha de consistência do Excel ORION: mês do cabeçalho DPP está incorreto.")

    by_key = _scenario_materials(scenario)
    audited_materials = 0
    formula_counts = {
        "gap": 0,
        "check": 0,
        "nec": 0,
        "stock_op": 0,
        "stock_total": 0,
        "balance": 0,
        "amount": 0,
    }

    for column in range(layout["model_start"], layout["model_end"] + 1):
        kit = _number(sheet.cell(layout["kit_row"], column).value)
        real = _number(sheet.cell(layout["real_row"], column).value)
        actual = _formula_text(sheet.cell(layout["gap_row"], column).value)
        if abs(real - kit) > FORMULA_TOLERANCE:
            letter = get_column_letter(column)
            expected = f"={letter}{layout['real_row']}-{letter}{layout['kit_row']}"
            if actual != expected:
                raise ValueError(
                    f"Falha de consistência do Excel ORION: fórmula de diferença REAL × KIT ausente/incorreta em {letter}{layout['gap_row']}."
                )
            formula_counts["gap"] += 1
        elif actual is not None:
            raise ValueError(
                "Falha de consistência do Excel ORION: existe fórmula de diferença onde KIT e REAL são iguais."
            )

    for row in range(layout["header_row"] + 1, sheet.max_row + 1):
        key = _material_key(sheet.cell(row, layout["material_col"]).value)
        if not key:
            continue
        material = by_key.get(key)
        if material is None:
            raise ValueError(f"Falha de consistência do Excel ORION: material inesperado na linha {row}.")
        formulas = _expected_formulas(layout, row, material)
        checks = {
            "check": (layout["check_col"], formulas["check"]),
            "nec": (layout["nec_col"], formulas["nec"]),
            "stock_total": (layout["stock_total_col"], formulas["stock_total"]),
            "balance": (layout["balance_col"], formulas["balance"]),
        }
        if formulas["stock_op"] is not None:
            checks["stock_op"] = (layout["stock_op_col"], formulas["stock_op"])
        if formulas["amount"] is not None:
            checks["amount"] = (layout["amount_col"], formulas["amount"])

        for family, (column, expected) in checks.items():
            actual = _formula_text(sheet.cell(row, column).value)
            if actual != expected:
                coordinate = f"{get_column_letter(column)}{row}"
                raise ValueError(
                    f"Falha de consistência do Excel ORION: fórmula '{family}' ausente/incorreta em {coordinate}."
                )
            formula_counts[family] += 1
        audited_materials += 1

    if audited_materials != len(by_key):
        raise ValueError(
            "Falha de consistência do Excel ORION: nem todos os materiais do cenário foram auditados no arquivo final."
        )

    if "Resumo de Análise" in workbook.sheetnames and "DPP BALANCE" in workbook.sheetnames:
        summary = workbook["Resumo de Análise"]
        model_start = get_column_letter(layout["model_start"])
        model_end = get_column_letter(layout["model_end"])
        model_name_row = layout["real_row"] - 1
        for row in range(2, summary.max_row + 1):
            if summary.cell(row, 2).value in (None, ""):
                continue
            expected_c = (
                f'=XLOOKUP(B{row},\'DPP BALANCE\'!${model_start}${model_name_row}:${model_end}${model_name_row},'
                f'\'DPP BALANCE\'!${model_start}${layout["kit_row"]}:${model_end}${layout["kit_row"]},"ND")'
            )
            expected_d = (
                f'=XLOOKUP(B{row},DPP!${model_start}${model_name_row}:${model_end}${model_name_row},'
                f'DPP!${model_start}${layout["real_row"]}:${model_end}${layout["real_row"]},"ND")'
            )
            expected_e = f"=D{row}-C{row}"
            if _formula_text(summary.cell(row, 3).value) != expected_c:
                raise ValueError(f"Falha de consistência do Excel ORION: fórmula do Resumo ausente/incorreta em C{row}.")
            if _formula_text(summary.cell(row, 4).value) != expected_d:
                raise ValueError(f"Falha de consistência do Excel ORION: fórmula do Resumo ausente/incorreta em D{row}.")
            if _formula_text(summary.cell(row, 5).value) != expected_e:
                raise ValueError(f"Falha de consistência do Excel ORION: fórmula do Resumo ausente/incorreta em E{row}.")


def finalize_and_audit_dpp_formulas(content: bytes, scenario: dict) -> bytes:
    """Reaplica e audita as fórmulas no mesmo XLSX que será entregue ao usuário.

    Este é o último estágio do job de exportação. Nenhum download é liberado se uma
    fórmula determinística esperada desaparecer durante as etapas anteriores ou durante
    a serialização do workbook.
    """
    if not content:
        raise ValueError("Falha de consistência do Excel ORION: conteúdo XLSX vazio.")

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        _apply_formula_contract(workbook, scenario)
        output = BytesIO()
        workbook.save(output)
        finalized = output.getvalue()
    finally:
        workbook.close()

    audit = load_workbook(BytesIO(finalized), data_only=False)
    try:
        _audit_formula_contract(audit, scenario)
    finally:
        audit.close()

    return finalized
