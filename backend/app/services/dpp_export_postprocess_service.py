from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.dpp_monthly_base import (
    SOURCE_SHEET,
    _find_column,
    _find_header_row,
    _find_label_row,
    _headers,
    _normalize,
)

GAP_TOLERANCE = 1e-4
MONTH_NAMES_PT = {
    "01": "JANEIRO",
    "02": "FEVEREIRO",
    "03": "MARÇO",
    "04": "ABRIL",
    "05": "MAIO",
    "06": "JUNHO",
    "07": "JULHO",
    "08": "AGOSTO",
    "09": "SETEMBRO",
    "10": "OUTUBRO",
    "11": "NOVEMBRO",
    "12": "DEZEMBRO",
}


def _month_label(reference_month: str) -> str:
    parts = str(reference_month or "").split("-")
    month = parts[1] if len(parts) >= 2 else ""
    name = MONTH_NAMES_PT.get(month)
    if not name:
        raise ValueError(
            "Falha de consistência do Excel ORION: mês de referência inválido para atualizar o cabeçalho do DPP."
        )
    return f"KIT Disponivel PGD ({name})"


def _number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _set_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    if hasattr(calculation, "calcMode"):
        calculation.calcMode = "auto"
    if hasattr(calculation, "fullCalcOnLoad"):
        calculation.fullCalcOnLoad = True
    if hasattr(calculation, "forceFullCalc"):
        calculation.forceFullCalc = True


def _layout(sheet) -> dict:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Falha de consistência do Excel ORION: a aba DPP está vazia.")

    header_row = _find_header_row(rows)
    headers = _headers(rows, header_row)
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    kit_row = _find_label_row(rows, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(rows, header_row, "REAL")
    if kit_row is None or real_row is None:
        raise ValueError(
            "Falha de consistência do Excel ORION: linhas KIT Disponível PGD/REAL não encontradas."
        )

    label_cell = None
    for row in range(1, header_row):
        for column in range(1, min(8, sheet.max_column) + 1):
            value = sheet.cell(row, column).value
            if "kit disponivel pgd" in _normalize(value):
                label_cell = sheet.cell(row, column)
                break
        if label_cell is not None:
            break
    if label_cell is None:
        raise ValueError(
            "Falha de consistência do Excel ORION: célula do rótulo KIT Disponível PGD não localizada."
        )

    model_start = origin_col + 1
    model_end = check_col - 1
    if model_end < model_start:
        raise ValueError("Falha de consistência do Excel ORION: faixa de modelos inválida.")

    gap_row = real_row + 1
    if gap_row >= header_row:
        raise ValueError(
            "Falha de consistência do Excel ORION: não há linha disponível para a diferença REAL × KIT."
        )

    return {
        "header_row": header_row,
        "kit_row": kit_row,
        "real_row": real_row,
        "gap_row": gap_row,
        "model_start": model_start,
        "model_end": model_end,
        "label_cell": label_cell,
    }


def enforce_final_dpp_header_and_gap(content: bytes, reference_month: str) -> bytes:
    """Garante no XLSX final o cabeçalho mensal e a fórmula REAL − KIT do DPP Final.

    A decisão de criar a fórmula usa os valores KIT e REAL já serializados no próprio
    workbook. Assim a fidelidade do Excel não depende de metadados intermediários do
    cenário: se REAL != KIT, a linha imediatamente abaixo de REAL recebe exatamente
    `=<REAL>-<KIT>`; quando são iguais, permanece vazia como no DPP Final de referência.
    """
    if not content:
        raise ValueError("Falha de consistência do Excel ORION: conteúdo XLSX vazio.")

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError("Falha de consistência do Excel ORION: a aba DPP não foi encontrada.")

        sheet = workbook[SOURCE_SHEET]
        layout = _layout(sheet)
        expected_label = _month_label(reference_month)
        layout["label_cell"].value = expected_label

        for column in range(layout["model_start"], layout["model_end"] + 1):
            kit = _number(sheet.cell(layout["kit_row"], column).value)
            real = _number(sheet.cell(layout["real_row"], column).value)
            gap_cell = sheet.cell(layout["gap_row"], column)
            if abs(real - kit) > GAP_TOLERANCE:
                letter = get_column_letter(column)
                gap_cell.value = f"={letter}{layout['real_row']}-{letter}{layout['kit_row']}"
            else:
                gap_cell.value = None

        _set_recalculation(workbook)
        output = BytesIO()
        workbook.save(output)
        finalized = output.getvalue()
    finally:
        workbook.close()

    audit = load_workbook(BytesIO(finalized), data_only=False, read_only=True)
    try:
        sheet = audit[SOURCE_SHEET]
        layout = _layout(sheet)
        if layout["label_cell"].value != _month_label(reference_month):
            raise ValueError(
                "Falha de consistência do Excel ORION: cabeçalho mensal do DPP não foi preservado após salvar."
            )

        for column in range(layout["model_start"], layout["model_end"] + 1):
            kit = _number(sheet.cell(layout["kit_row"], column).value)
            real = _number(sheet.cell(layout["real_row"], column).value)
            actual = sheet.cell(layout["gap_row"], column).value
            if abs(real - kit) > GAP_TOLERANCE:
                letter = get_column_letter(column)
                expected = f"={letter}{layout['real_row']}-{letter}{layout['kit_row']}"
                if actual != expected:
                    raise ValueError(
                        "Falha de consistência do Excel ORION: fórmula de diferença REAL × KIT não foi preservada."
                    )
            elif actual not in (None, ""):
                raise ValueError(
                    "Falha de consistência do Excel ORION: diferença REAL × KIT recebeu fórmula apesar de KIT e REAL serem iguais."
                )
    finally:
        audit.close()

    return finalized
