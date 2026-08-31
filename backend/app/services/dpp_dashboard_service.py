from __future__ import annotations

from collections import OrderedDict
from io import BytesIO
from pathlib import Path
from threading import Lock
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import load_workbook

from app.services.dpp_consolidation_service import _material_key
from app.services.dpp_nec_divergence_service import explain_nec_divergence
from app.services.dpp_projection_service import (
    aggregate,
    build_orion_projection,
    column_values_equal,
    critical_rule_metadata,
    is_critical_material,
)
from app.services.dpp_scenario_service import get_latest_monthly_scenario
from app.services.dpp_service import (
    SOURCE_SHEET,
    SUPPORTED_DPP_EXTENSIONS,
    VALIDATION_ABS_TOL,
    _as_text,
    _cell,
    _find_column,
    _find_header_row,
    _find_label_row,
    _headers,
    _normalize,
    _number,
)

MAX_FINAL_COLUMN_SNAPSHOTS = 4
_FINAL_COLUMN_SNAPSHOTS: OrderedDict[str, dict] = OrderedDict()
_FINAL_COLUMN_SNAPSHOTS_LOCK = Lock()


def _optional_column(headers: dict[int, str], *names: str, startswith: str | None = None) -> int | None:
    accepted = {_normalize(name) for name in names}
    for column, header in headers.items():
        normalized = _normalize(header)
        if normalized in accepted or (startswith and normalized.startswith(_normalize(startswith))):
            return column
    return None


def _register_column_snapshot(analysis_id: str, snapshot: dict) -> None:
    with _FINAL_COLUMN_SNAPSHOTS_LOCK:
        _FINAL_COLUMN_SNAPSHOTS[analysis_id] = snapshot
        _FINAL_COLUMN_SNAPSHOTS.move_to_end(analysis_id)
        while len(_FINAL_COLUMN_SNAPSHOTS) > MAX_FINAL_COLUMN_SNAPSHOTS:
            _FINAL_COLUMN_SNAPSHOTS.popitem(last=False)


def _column_snapshot(analysis_id: str) -> dict | None:
    with _FINAL_COLUMN_SNAPSHOTS_LOCK:
        snapshot = _FINAL_COLUMN_SNAPSHOTS.get(analysis_id)
        if snapshot is not None:
            _FINAL_COLUMN_SNAPSHOTS.move_to_end(analysis_id)
        return snapshot


def _column_rule(header: str, spec: dict) -> dict:
    normalized = _normalize(header)
    field = spec.get("field")
    comparison = spec.get("comparison")

    if field == "model":
        definition = f"Compara o consumo do material na coluna do modelo {header}."
        origin = "ORION: matriz Material × Modelo montada a partir da base mensal; DPP Final: célula consolidada da mesma coluna."
    elif field == "nec":
        definition = "NEC = Σ(REAL do modelo × consumo do material)."
        origin = "O ORION recalcula a necessidade usando o REAL do cenário e a matriz Material × Modelo."
    elif field == "stock_total":
        definition = "STK TTL = STK + EXPLOSÃO + STK OP."
        origin = "O ORION soma os três componentes canônicos de estoque do material."
    elif field == "balance":
        definition = "SALDO = STK TTL − NEC."
        origin = "O ORION subtrai a necessidade calculada do estoque total calculado."
    elif field == "stock_sap_effective":
        definition = "Compara o estoque SAP efetivo do material."
        origin = "ORION: arquivo STK mensal; DPP Final: coluna STK consolidada."
    elif field == "explosion":
        definition = "Compara a quantidade de EXPLOSÃO do material."
        origin = "ORION: arquivo de Explosão mensal; DPP Final: coluna EXPLOSÃO consolidada."
    elif field == "stock_op":
        definition = "Compara o STK OP associado ao material."
        origin = "ORION: valor operacional consolidado no cenário; DPP Final: coluna STK OP."
    elif field == "in_current_wiu":
        definition = "WIU é comparado por presença: preenchido versus não preenchido."
        origin = "ORION: presença do material na WIU mensal; DPP Final: preenchimento da coluna WIU."
    elif field == "check":
        definition = "Compara o valor operacional da coluna Check; o status interno do ORION não é usado como substituto."
        origin = "ORION: campo Check derivado/mantido no DPP; DPP Final: valor consolidado da coluna Check."
    elif field == "optional_material":
        definition = "Compara o código OPC associado ao material."
        origin = "ORION: material opcional identificado no cenário; DPP Final: coluna OPC consolidada."
    elif normalized == "material":
        definition = "O item diverge quando a chave normalizada do material existe apenas em um dos cenários."
        origin = "As bases são alinhadas pela chave Material normalizada."
    else:
        definition = f"Compara o valor da coluna {header} para o mesmo material nos dois cenários."
        origin = "ORION: projeção canônica do cenário; DPP Final: valor lido diretamente do arquivo consolidado."

    if comparison == "numeric":
        criterion = f"Divergente quando |DPP Final − ORION| > {VALIDATION_ABS_TOL:g}."
    elif comparison == "presence":
        criterion = "Divergente quando um lado está preenchido e o outro não."
    else:
        criterion = "Divergente quando os valores normalizados não são iguais."

    return {
        "definition": definition,
        "origin": origin,
        "criterion": criterion,
        "comparison": comparison or "text",
        "tolerance": VALIDATION_ABS_TOL if comparison == "numeric" else None,
    }


def _difference_reason(header: str, spec: dict, final_value: object, orion_value: object, presence: str | None = None) -> str:
    if presence == "final_only":
        return "O material existe no DPP Final, mas não existe na projeção do Cenário ORION."
    if presence == "orion_only":
        return "O material existe no Cenário ORION, mas não existe no DPP Final."

    field = spec.get("field")
    comparison = spec.get("comparison")
    if comparison == "numeric":
        delta = (_number(final_value, 0.0) or 0.0) - (_number(orion_value, 0.0) or 0.0)
        if field == "nec":
            return f"O NEC recalculado pelo ORION difere do consolidado final em {delta:.6g}. A regra usada é NEC = Σ(REAL × consumo)."
        if field == "stock_total":
            return f"O STK TTL do ORION difere do consolidado final em {delta:.6g}. A regra usada é STK + EXPLOSÃO + STK OP."
        if field == "balance":
            return f"O SALDO do ORION difere do consolidado final em {delta:.6g}. A regra usada é STK TTL − NEC."
        if field == "model":
            return f"O consumo do material para {header} difere em {delta:.6g} entre a matriz ORION e o DPP Final."
        return f"O valor numérico difere em {delta:.6g}, acima da tolerância operacional de {VALIDATION_ABS_TOL:g}."
    if comparison == "presence":
        return "A presença do item é diferente: um cenário considera a coluna preenchida e o outro não."
    return "Os valores textuais normalizados são diferentes para o mesmo material."


def _build_column_comparison(
    *,
    rows: list[tuple],
    headers: dict[int, str],
    header_row: int,
    material_col: int,
    origin_col: int,
    check_col: int,
    scenario: dict,
    analysis_id: str | None = None,
) -> dict:
    projection = build_orion_projection(
        scenario=scenario, headers=headers, material_col=material_col, origin_col=origin_col, check_col=check_col,
    )
    final_material_rows: list[tuple[str, int]] = []
    final_row_by_material: dict[str, int] = {}
    for excel_row in range(header_row + 1, len(rows) + 1):
        key = _material_key(_cell(rows, excel_row, material_col))
        if not key:
            continue
        final_material_rows.append((key, excel_row))
        final_row_by_material[key] = excel_row

    columns: list[dict] = []
    divergent_columns = 0
    comparable_columns = 0
    for column in projection["columns"]:
        header = headers[column]
        spec = projection["specs"][column]
        kind = spec["kind"]
        final_total = aggregate([_cell(rows, excel_row, column) for _, excel_row in final_material_rows], kind)
        item = {
            "name": header, "column": column, "kind": kind,
            "aggregation_label": "itens preenchidos" if kind == "count" else "soma",
            "final_total": final_total, "orion_total": None, "delta": None,
            "difference_count": None, "supported": bool(spec["supported"]),
            "drilldown_available": False,
        }
        if not spec["supported"]:
            item["note"] = "Ainda não calculado pelo cenário ORION."
            columns.append(item)
            continue

        comparable_columns += 1
        orion_total = projection["totals"][column]
        delta = final_total - orion_total
        differences = 0
        for material_key in set(final_row_by_material) | set(projection["by_key"]):
            final_row = final_row_by_material.get(material_key)
            projected_row = projection["by_key"].get(material_key)
            if final_row is None or projected_row is None:
                differences += 1
                continue
            if not column_values_equal(_cell(rows, final_row, column), projected_row["values"].get(column), spec.get("comparison")):
                differences += 1
        item.update(
            orion_total=orion_total,
            delta=delta,
            difference_count=differences,
            drilldown_available=bool(analysis_id and differences > 0),
        )
        if differences > 0 or abs(float(delta)) > VALIDATION_ABS_TOL:
            divergent_columns += 1
        columns.append(item)

    if analysis_id:
        description_col = _optional_column(headers, "Descrição", "Descricao")
        max_column = max(len(row) for row in rows)
        real_row = _find_label_row(rows, max_column, header_row, "REAL")
        _register_column_snapshot(analysis_id, {
            "rows": rows,
            "headers": headers,
            "material_col": material_col,
            "description_col": description_col,
            "projection": projection,
            "final_material_rows": final_material_rows,
            "final_row_by_material": final_row_by_material,
            "real_row": real_row,
            "scenario_models": list(scenario.get("models") or []),
        })

    return {
        "analysis_id": analysis_id,
        "scenario_id": projection["scenario_id"], "reference_month": projection["reference_month"],
        "basis": "canonical_orion_projection", "columns_total": len(columns),
        "comparable_columns": comparable_columns, "divergent_columns": divergent_columns,
        "unsupported_columns": len(columns) - comparable_columns,
        "final_materials": len(final_material_rows), "orion_materials": len(projection["rows"]), "columns": columns,
    }


def get_column_divergences(analysis_id: str, column: int, offset: int = 0, limit: int = 25) -> dict | None:
    snapshot = _column_snapshot(analysis_id)
    if snapshot is None:
        return None

    projection = snapshot["projection"]
    spec = projection["specs"].get(column)
    header = snapshot["headers"].get(column)
    if spec is None or header is None:
        raise ValueError("A coluna solicitada não pertence ao comparativo analisado.")
    if not spec.get("supported"):
        raise ValueError("A coluna solicitada ainda não é calculada pelo Cenário ORION.")

    final_row_by_material = snapshot["final_row_by_material"]
    final_order = [key for key, _ in snapshot["final_material_rows"]]
    orion_only = [row["key"] for row in projection["rows"] if row["key"] not in final_row_by_material]
    ordered_keys = final_order + orion_only
    rows = snapshot["rows"]
    material_col = snapshot["material_col"]
    description_col = snapshot.get("description_col")

    total = 0
    items: list[dict] = []
    page_end = offset + limit

    for material_key in ordered_keys:
        final_row = final_row_by_material.get(material_key)
        projected_row = projection["by_key"].get(material_key)
        presence = None
        if final_row is None:
            divergent = True
            presence = "orion_only"
            final_value = None
            orion_value = projected_row["values"].get(column) if projected_row else None
        elif projected_row is None:
            divergent = True
            presence = "final_only"
            final_value = _cell(rows, final_row, column)
            orion_value = None
        else:
            final_value = _cell(rows, final_row, column)
            orion_value = projected_row["values"].get(column)
            divergent = not column_values_equal(final_value, orion_value, spec.get("comparison"))

        if not divergent:
            continue

        if total >= offset and total < page_end:
            material = None
            description = None
            if projected_row:
                material = projected_row["material"].get("material")
                description = projected_row["material"].get("description")
            if final_row is not None:
                material = material or _as_text(_cell(rows, final_row, material_col))
                if description_col:
                    description = description or _as_text(_cell(rows, final_row, description_col))

            delta = None
            if spec.get("comparison") == "numeric" and presence is None:
                delta = (_number(final_value, 0.0) or 0.0) - (_number(orion_value, 0.0) or 0.0)

            reason = _difference_reason(header, spec, final_value, orion_value, presence)
            if (
                spec.get("field") == "nec"
                and presence is None
                and final_row is not None
                and projected_row is not None
            ):
                reason = explain_nec_divergence(
                    rows=rows,
                    headers=snapshot["headers"],
                    real_row=snapshot.get("real_row"),
                    final_row=final_row,
                    projected_row=projected_row,
                    projection=projection,
                    scenario_models=snapshot.get("scenario_models") or [],
                    final_value=final_value,
                    orion_value=orion_value,
                ) or reason

            items.append({
                "material": material or material_key,
                "description": description or "",
                "orion_value": orion_value,
                "final_value": final_value,
                "delta": delta,
                "reason": reason,
            })
        total += 1

    return {
        "analysis_id": analysis_id,
        "column": column,
        "name": header,
        "rule": _column_rule(header, spec),
        "total": total,
        "offset": offset,
        "limit": limit,
        "returned": len(items),
        "has_previous": offset > 0,
        "has_next": offset + len(items) < total,
        "items": items,
    }


def summarize_final_dpp_content(
    content: bytes,
    filename: str = "dpp.xlsx",
    scenario: dict | None = None,
    analysis_id: str | None = None,
) -> dict:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_DPP_EXTENSIONS:
        raise ValueError("Envie um DPP final em formato .xlsx ou .xlsm.")
    if not content:
        raise ValueError("O arquivo do DPP final está vazio.")

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        workbook.close()
        raise ValueError("A aba 'DPP' não foi encontrada no DPP final.")
    sheet = workbook[SOURCE_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValueError("A aba 'DPP' do arquivo final está vazia.")

    max_column = max(len(row) for row in rows)
    header_row = _find_header_row(rows, max_column)
    headers = _headers(rows, max_column, header_row)

    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    optional_col = _find_column(headers, "OPC")
    balance_col = _find_column(headers, "SALDO")
    nec_col = _optional_column(headers, "NEC", "NECESSIDADE")
    stock_source_col = next((c for c, h in headers.items() if _normalize(h).startswith("stk ") and _normalize(h) not in {"stk op", "stk ttl"}), None)
    explosion_col = _optional_column(headers, startswith="explosao")
    stock_op_col = _optional_column(headers, "STK OP", "STK OPC", "STK OPCS")
    stock_total_col = _optional_column(headers, "STK TTL", "STK TOTAL")

    model_start = origin_col + 1
    model_end = check_col - 1
    if model_end < model_start:
        raise ValueError("Não foi possível identificar os modelos do DPP final.")

    kit_row = _find_label_row(rows, max_column, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(rows, max_column, header_row, "REAL")
    if real_row is None:
        raise ValueError("A linha REAL não foi encontrada no DPP final.")

    pgd_total = 0.0
    real_total = 0.0
    model_states: list[dict] = []
    for column in range(model_start, model_end + 1):
        model_name = _as_text(_cell(rows, header_row, column))
        if not model_name:
            continue
        pgd = max(_number(_cell(rows, kit_row, column), 0.0) or 0.0, 0.0) if kit_row else 0.0
        real = max(_number(_cell(rows, real_row, column), 0.0) or 0.0, 0.0)
        delta = real - pgd
        pgd_total += pgd
        real_total += real
        model_states.append({
            "name": model_name, "column": column, "pgd": pgd, "real": real, "delta": delta,
            "active": real > 1e-9, "changed": abs(delta) > 1e-9,
        })

    total_materials = 0
    opc_count = 0
    risk_model_names: set[str] = set()
    critical_material_codes: set[str] = set()
    shared_critical_material_codes: set[str] = set()
    material_details: list[dict] = []

    for excel_row in range(header_row + 1, len(rows) + 1):
        material = _as_text(_cell(rows, excel_row, material_col))
        if not material:
            continue
        total_materials += 1
        optional_material = _as_text(_cell(rows, excel_row, optional_col))
        if optional_material:
            opc_count += 1

        unit = _as_text(_cell(rows, excel_row, um_col))
        balance = _number(_cell(rows, excel_row, balance_col), None)
        critical = is_critical_material(unit, balance)
        affected_models = []
        for model in model_states:
            if not model["active"]:
                continue
            usage = _number(_cell(rows, excel_row, model["column"]), 0.0) or 0.0
            if usage > 1e-9:
                affected_models.append(model["name"])
                if critical:
                    risk_model_names.add(model["name"])

        shared = critical and len(affected_models) > 1
        if critical:
            critical_material_codes.add(material)
        if shared:
            shared_critical_material_codes.add(material)

        material_details.append({
            "material": material,
            "description": _as_text(_cell(rows, excel_row, description_col)),
            "um": unit,
            "group_origin": _as_text(_cell(rows, excel_row, origin_col)),
            "nec": _number(_cell(rows, excel_row, nec_col), None) if nec_col else None,
            "stock": _number(_cell(rows, excel_row, stock_source_col), None) if stock_source_col else None,
            "explosion": _number(_cell(rows, excel_row, explosion_col), None) if explosion_col else None,
            "stock_op": _number(_cell(rows, excel_row, stock_op_col), None) if stock_op_col else None,
            "stock_total": _number(_cell(rows, excel_row, stock_total_col), None) if stock_total_col else None,
            "balance": balance,
            "optional_material": optional_material,
            "critical": critical,
            "shared_critical": shared,
            "affected_models": affected_models,
        })

    for model in model_states:
        model["at_risk"] = model["name"] in risk_model_names
        model.pop("column", None)

    active_models = sum(1 for model in model_states if model["active"])
    risk_models = len(risk_model_names)
    safe_models = max(active_models - risk_models, 0)
    material_coverage = (safe_models / active_models * 100.0) if active_models else 0.0
    pgd_exposed = sum(model["pgd"] for model in model_states if model["name"] in risk_model_names)
    changed_models = sum(1 for model in model_states if model["changed"])
    below_pgd_models = sum(1 for model in model_states if model["delta"] < -1e-9)
    above_pgd_models = sum(1 for model in model_states if model["delta"] > 1e-9)

    column_comparison = None
    if scenario is not None:
        column_comparison = _build_column_comparison(
            rows=rows, headers=headers, header_row=header_row, material_col=material_col,
            origin_col=origin_col, check_col=check_col, scenario=scenario, analysis_id=analysis_id,
        )

    return {
        "analysis_id": analysis_id,
        "filename": filename,
        "status": "DPP_FINAL",
        "models": model_states,
        "material_details": material_details,
        "critical_rule": critical_rule_metadata(),
        "critical_materials": sorted(critical_material_codes),
        "shared_critical_materials": sorted(shared_critical_material_codes),
        "column_comparison": column_comparison,
        "summary": {
            "pgd_total": pgd_total, "real_total": real_total, "model_count": len(model_states),
            "active_models": active_models, "changed_models": changed_models,
            "below_pgd_models": below_pgd_models, "above_pgd_models": above_pgd_models,
            "total_materials": total_materials, "critical_materials": len(critical_material_codes),
            "opc_count": opc_count, "risk_models": risk_models, "safe_models": safe_models,
            "material_coverage": material_coverage, "pgd_exposed": pgd_exposed,
            "shared_critical": len(shared_critical_material_codes),
        },
    }


async def summarize_final_dpp(file: UploadFile) -> dict:
    filename = file.filename or "dpp.xlsx"
    content = await file.read()
    analysis_id = uuid4().hex
    return summarize_final_dpp_content(
        content,
        filename,
        scenario=get_latest_monthly_scenario(),
        analysis_id=analysis_id,
    )
