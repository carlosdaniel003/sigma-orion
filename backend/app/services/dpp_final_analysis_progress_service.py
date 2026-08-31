from __future__ import annotations

from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from threading import Lock
from time import monotonic
from uuid import uuid4

from openpyxl import load_workbook

from app.services.dpp_consolidation_service import _material_key
from app.services.dpp_dashboard_service import _register_column_snapshot
from app.services.dpp_projection_service import (
    aggregate,
    build_orion_projection,
    column_values_equal,
    critical_rule_metadata,
    is_critical_material,
)
from app.services.dpp_scenario_service import get_latest_monthly_scenario
from app.services.dpp_service import (
    SOURCE_SHEET,
    SUPPORTED_DPP_EXTENSIONS,
    VALIDATION_ABS_TOL,
    _as_text,
    _cell,
    _find_column,
    _find_header_row,
    _find_label_row,
    _headers,
    _normalize,
    _number,
)

MAX_JOBS = 8
_JOBS: OrderedDict[str, dict] = OrderedDict()
_LOCK = Lock()
_EXECUTOR = ThreadPoolExecutor(max_workers=2, thread_name_prefix="orion-final-dpp")


def _now() -> float:
    return monotonic()


def _trim_jobs() -> None:
    while len(_JOBS) > MAX_JOBS:
        _JOBS.popitem(last=False)


def _snapshot(job: dict) -> dict:
    created_at = float(job.get("created_at") or _now())
    finished_at = job.get("finished_at")
    end = float(finished_at) if finished_at is not None else _now()
    return {
        "job_id": job["job_id"],
        "kind": job["kind"],
        "status": job["status"],
        "progress": job["progress"],
        "activity": job["activity"],
        "elapsed_seconds": max(end - created_at, 0.0),
        "error": job.get("error"),
        "result": job.get("result") if job["status"] == "completed" else None,
    }


def _create_job() -> str:
    job_id = uuid4().hex
    with _LOCK:
        _JOBS[job_id] = {
            "job_id": job_id,
            "kind": "final_dpp_analysis",
            "status": "queued",
            "progress": 0,
            "activity": "Aguardando análise do DPP Final",
            "created_at": _now(),
            "finished_at": None,
            "error": None,
            "result": None,
        }
        _JOBS.move_to_end(job_id)
        _trim_jobs()
    return job_id


def _update_job(job_id: str, progress: int, activity: str) -> None:
    with _LOCK:
        job = _JOBS.get(job_id)
        if job is None or job["status"] in {"completed", "failed"}:
            return
        job["status"] = "running"
        job["progress"] = max(int(job.get("progress", 0)), min(max(int(progress), 0), 99))
        job["activity"] = activity
        _JOBS.move_to_end(job_id)


def _complete_job(job_id: str, result: dict) -> None:
    with _LOCK:
        job = _JOBS.get(job_id)
        if job is None:
            return
        job["status"] = "completed"
        job["progress"] = 100
        job["activity"] = "Análise do DPP Final concluída"
        job["result"] = result
        job["error"] = None
        job["finished_at"] = _now()
        _JOBS.move_to_end(job_id)


def _fail_job(job_id: str, error: Exception) -> None:
    with _LOCK:
        job = _JOBS.get(job_id)
        if job is None:
            return
        job["status"] = "failed"
        job["activity"] = "Análise do DPP Final interrompida"
        job["error"] = str(error) or error.__class__.__name__
        job["finished_at"] = _now()
        _JOBS.move_to_end(job_id)


def _optional_column(headers: dict[int, str], *names: str, startswith: str | None = None) -> int | None:
    accepted = {_normalize(name) for name in names}
    for column, header in headers.items():
        normalized = _normalize(header)
        if normalized in accepted or (startswith and normalized.startswith(_normalize(startswith))):
            return column
    return None


def _build_column_comparison_with_progress(
    *,
    rows: list[tuple],
    headers: dict[int, str],
    header_row: int,
    material_col: int,
    origin_col: int,
    check_col: int,
    scenario: dict,
    analysis_id: str,
    progress,
) -> dict:
    progress(62, "Preparando projeção canônica para o comparativo completo")
    projection = build_orion_projection(
        scenario=scenario,
        headers=headers,
        material_col=material_col,
        origin_col=origin_col,
        check_col=check_col,
    )

    final_material_rows: list[tuple[str, int]] = []
    final_row_by_material: dict[str, int] = {}
    for excel_row in range(header_row + 1, len(rows) + 1):
        key = _material_key(_cell(rows, excel_row, material_col))
        if not key:
            continue
        final_material_rows.append((key, excel_row))
        final_row_by_material[key] = excel_row

    columns: list[dict] = []
    divergent_columns = 0
    comparable_columns = 0
    projection_columns = list(projection["columns"])
    total_columns = max(len(projection_columns), 1)
    material_keys = set(final_row_by_material) | set(projection["by_key"])

    for index, column in enumerate(projection_columns, start=1):
        header = headers[column]
        spec = projection["specs"][column]
        kind = spec["kind"]
        final_total = aggregate([_cell(rows, excel_row, column) for _, excel_row in final_material_rows], kind)
        item = {
            "name": header,
            "column": column,
            "kind": kind,
            "aggregation_label": "itens preenchidos" if kind == "count" else "soma",
            "final_total": final_total,
            "orion_total": None,
            "delta": None,
            "difference_count": None,
            "supported": bool(spec["supported"]),
            "drilldown_available": False,
        }

        if not spec["supported"]:
            item["note"] = "Ainda não calculado pelo cenário ORION."
            columns.append(item)
        else:
            comparable_columns += 1
            orion_total = projection["totals"][column]
            delta = final_total - orion_total
            differences = 0
            for material_key in material_keys:
                final_row = final_row_by_material.get(material_key)
                projected_row = projection["by_key"].get(material_key)
                if final_row is None or projected_row is None:
                    differences += 1
                    continue
                if not column_values_equal(
                    _cell(rows, final_row, column),
                    projected_row["values"].get(column),
                    spec.get("comparison"),
                ):
                    differences += 1

            item.update(
                orion_total=orion_total,
                delta=delta,
                difference_count=differences,
                drilldown_available=differences > 0,
            )
            if differences > 0 or abs(float(delta)) > VALIDATION_ABS_TOL:
                divergent_columns += 1
            columns.append(item)

        current = 62 + int((index / total_columns) * 36)
        progress(current, f"Comparando colunas do DPP · {index}/{len(projection_columns)}")

    description_col = _optional_column(headers, "Descrição", "Descricao")
    _register_column_snapshot(
        analysis_id,
        {
            "rows": rows,
            "headers": headers,
            "material_col": material_col,
            "description_col": description_col,
            "projection": projection,
            "final_material_rows": final_material_rows,
            "final_row_by_material": final_row_by_material,
        },
    )

    return {
        "analysis_id": analysis_id,
        "scenario_id": projection["scenario_id"],
        "reference_month": projection["reference_month"],
        "basis": "canonical_orion_projection",
        "columns_total": len(columns),
        "comparable_columns": comparable_columns,
        "divergent_columns": divergent_columns,
        "unsupported_columns": len(columns) - comparable_columns,
        "final_materials": len(final_material_rows),
        "orion_materials": len(projection["rows"]),
        "columns": columns,
    }


def _summarize_final_dpp_content_with_progress(
    *,
    content: bytes,
    filename: str,
    scenario: dict | None,
    analysis_id: str,
    progress,
) -> dict:
    progress(3, "Validando arquivo do DPP Final")
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_DPP_EXTENSIONS:
        raise ValueError("Envie um DPP final em formato .xlsx ou .xlsm.")
    if not content:
        raise ValueError("O arquivo do DPP final está vazio.")

    progress(6, "Abrindo workbook do DPP Final")
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError("A aba 'DPP' não foi encontrada no DPP final.")
        sheet = workbook[SOURCE_SHEET]
        estimated_rows = max(int(sheet.max_row or 0), 1)
        rows: list[tuple] = []
        last_progress = -1
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            rows.append(tuple(row))
            current = 6 + int((min(index, estimated_rows) / estimated_rows) * 14)
            if current != last_progress:
                progress(current, f"Lendo aba DPP · {index}/{estimated_rows} linhas")
                last_progress = current
    finally:
        workbook.close()

    if not rows:
        raise ValueError("A aba 'DPP' do arquivo final está vazia.")

    progress(21, "Identificando estrutura e colunas do DPP Final")
    max_column = max(len(row) for row in rows)
    header_row = _find_header_row(rows, max_column)
    headers = _headers(rows, max_column, header_row)

    material_col = _find_column(headers, "Material")
    description_col = _find_column(headers, "Descrição")
    um_col = _find_column(headers, "UM")
    origin_col = _find_column(headers, "Grupo Origem")
    check_col = _find_column(headers, "Check")
    optional_col = _find_column(headers, "OPC")
    balance_col = _find_column(headers, "SALDO")
    nec_col = _optional_column(headers, "NEC", "NECESSIDADE")
    stock_source_col = next(
        (
            column
            for column, header in headers.items()
            if _normalize(header).startswith("stk ") and _normalize(header) not in {"stk op", "stk ttl"}
        ),
        None,
    )
    explosion_col = _optional_column(headers, startswith="explosao")
    stock_op_col = _optional_column(headers, "STK OP", "STK OPC", "STK OPCS")
    stock_total_col = _optional_column(headers, "STK TTL", "STK TOTAL")

    model_start = origin_col + 1
    model_end = check_col - 1
    if model_end < model_start:
        raise ValueError("Não foi possível identificar os modelos do DPP final.")

    kit_row = _find_label_row(rows, max_column, header_row, "KIT Disponivel PGD", contains=True)
    real_row = _find_label_row(rows, max_column, header_row, "REAL")
    if real_row is None:
        raise ValueError("A linha REAL não foi encontrada no DPP final.")

    pgd_total = 0.0
    real_total = 0.0
    model_states: list[dict] = []
    model_columns = list(range(model_start, model_end + 1))
    total_model_columns = max(len(model_columns), 1)
    for index, column in enumerate(model_columns, start=1):
        model_name = _as_text(_cell(rows, header_row, column))
        if model_name:
            pgd = max(_number(_cell(rows, kit_row, column), 0.0) or 0.0, 0.0) if kit_row else 0.0
            real = max(_number(_cell(rows, real_row, column), 0.0) or 0.0, 0.0)
            delta = real - pgd
            pgd_total += pgd
            real_total += real
            model_states.append(
                {
                    "name": model_name,
                    "column": column,
                    "pgd": pgd,
                    "real": real,
                    "delta": delta,
                    "active": real > 1e-9,
                    "changed": abs(delta) > 1e-9,
                }
            )
        current = 22 + int((index / total_model_columns) * 8)
        progress(current, f"Analisando modelos · {index}/{len(model_columns)}")

    total_materials = 0
    opc_count = 0
    risk_model_names: set[str] = set()
    critical_material_codes: set[str] = set()
    shared_critical_material_codes: set[str] = set()
    material_details: list[dict] = []
    material_row_count = max(len(rows) - header_row, 1)
    processed_material_rows = 0
    last_material_progress = -1

    for excel_row in range(header_row + 1, len(rows) + 1):
        processed_material_rows += 1
        material = _as_text(_cell(rows, excel_row, material_col))
        if material:
            total_materials += 1
            optional_material = _as_text(_cell(rows, excel_row, optional_col))
            if optional_material:
                opc_count += 1

            unit = _as_text(_cell(rows, excel_row, um_col))
            balance = _number(_cell(rows, excel_row, balance_col), None)
            critical = is_critical_material(unit, balance)
            affected_models = []
            for model in model_states:
                if not model["active"]:
                    continue
                usage = _number(_cell(rows, excel_row, model["column"]), 0.0) or 0.0
                if usage > 1e-9:
                    affected_models.append(model["name"])
                    if critical:
                        risk_model_names.add(model["name"])

            shared = critical and len(affected_models) > 1
            if critical:
                critical_material_codes.add(material)
            if shared:
                shared_critical_material_codes.add(material)

            material_details.append(
                {
                    "material": material,
                    "description": _as_text(_cell(rows, excel_row, description_col)),
                    "um": unit,
                    "group_origin": _as_text(_cell(rows, excel_row, origin_col)),
                    "nec": _number(_cell(rows, excel_row, nec_col), None) if nec_col else None,
                    "stock": _number(_cell(rows, excel_row, stock_source_col), None) if stock_source_col else None,
                    "explosion": _number(_cell(rows, excel_row, explosion_col), None) if explosion_col else None,
                    "stock_op": _number(_cell(rows, excel_row, stock_op_col), None) if stock_op_col else None,
                    "stock_total": _number(_cell(rows, excel_row, stock_total_col), None) if stock_total_col else None,
                    "balance": balance,
                    "optional_material": optional_material,
                    "critical": critical,
                    "shared_critical": shared,
                    "affected_models": affected_models,
                }
            )

        current = 30 + int((processed_material_rows / material_row_count) * 30)
        if current != last_material_progress:
            progress(current, f"Analisando materiais · {processed_material_rows}/{material_row_count}")
            last_material_progress = current

    for model in model_states:
        model["at_risk"] = model["name"] in risk_model_names
        model.pop("column", None)

    active_models = sum(1 for model in model_states if model["active"])
    risk_models = len(risk_model_names)
    safe_models = max(active_models - risk_models, 0)
    material_coverage = (safe_models / active_models * 100.0) if active_models else 0.0
    pgd_exposed = sum(model["pgd"] for model in model_states if model["name"] in risk_model_names)
    changed_models = sum(1 for model in model_states if model["changed"])
    below_pgd_models = sum(1 for model in model_states if model["delta"] < -1e-9)
    above_pgd_models = sum(1 for model in model_states if model["delta"] > 1e-9)

    column_comparison = None
    if scenario is not None:
        column_comparison = _build_column_comparison_with_progress(
            rows=rows,
            headers=headers,
            header_row=header_row,
            material_col=material_col,
            origin_col=origin_col,
            check_col=check_col,
            scenario=scenario,
            analysis_id=analysis_id,
            progress=progress,
        )
    else:
        progress(98, "Finalizando resumo do DPP Final")

    progress(99, "Consolidando resultado da análise do DPP Final")
    return {
        "analysis_id": analysis_id,
        "filename": filename,
        "status": "DPP_FINAL",
        "models": model_states,
        "material_details": material_details,
        "critical_rule": critical_rule_metadata(),
        "critical_materials": sorted(critical_material_codes),
        "shared_critical_materials": sorted(shared_critical_material_codes),
        "column_comparison": column_comparison,
        "summary": {
            "pgd_total": pgd_total,
            "real_total": real_total,
            "model_count": len(model_states),
            "active_models": active_models,
            "changed_models": changed_models,
            "below_pgd_models": below_pgd_models,
            "above_pgd_models": above_pgd_models,
            "total_materials": total_materials,
            "critical_materials": len(critical_material_codes),
            "opc_count": opc_count,
            "risk_models": risk_models,
            "safe_models": safe_models,
            "material_coverage": material_coverage,
            "pgd_exposed": pgd_exposed,
            "shared_critical": len(shared_critical_material_codes),
        },
    }


def _run_final_analysis_job(job_id: str, filename: str, content: bytes) -> None:
    def progress(value: int, activity: str) -> None:
        _update_job(job_id, value, activity)

    try:
        analysis_id = uuid4().hex
        result = _summarize_final_dpp_content_with_progress(
            content=content,
            filename=filename,
            scenario=get_latest_monthly_scenario(),
            analysis_id=analysis_id,
            progress=progress,
        )
        _complete_job(job_id, result)
    except Exception as exc:
        _fail_job(job_id, exc)


def start_final_analysis_job(*, filename: str | None, content: bytes) -> dict:
    job_id = _create_job()
    _EXECUTOR.submit(_run_final_analysis_job, job_id, filename or "dpp.xlsx", content)
    return get_final_analysis_job(job_id) or {"job_id": job_id, "status": "queued", "progress": 0}


def get_final_analysis_job(job_id: str) -> dict | None:
    with _LOCK:
        job = _JOBS.get(job_id)
        if job is None:
            return None
        return _snapshot(job)
