from io import BytesIO

from openpyxl import Workbook

from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_final_analysis_progress_service import _summarize_final_dpp_content_with_progress


def _final_dpp_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD"
    sheet["E1"] = 10
    sheet["F1"] = 20
    sheet["A2"] = "REAL"
    sheet["E2"] = 12
    sheet["F2"] = 18

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODELO A",
        "MODELO B",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
        "Coments",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)

    rows = [
        ("MAT-A", "Material A", "UN", "LOCAL", 2, 1, "", "WIU", 42, 100, 10, "", 0, 110, 68),
        ("MAT-B", "Material B", "UN", "IMPORTADO", 1, 0, "", "WIU", 12, 5, 0, "", 0, 5, -7),
    ]
    for row_index, values in enumerate(rows, start=5):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_index, column, value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _scenario() -> dict:
    scenario_service._SCENARIOS.clear()
    return scenario_service.register_monthly_scenario(
        materials=[
            {
                "material": "MAT-A",
                "material_key": "MAT-A",
                "description": "Material A",
                "um": "UN",
                "group_origin": "LOCAL",
                "check": "",
                "in_current_wiu": True,
                "optional_material": None,
                "consumption_by_model": {"MODELO A": 2.0, "MODELO B": 1.0},
                "stock_sap_effective": 100.0,
                "explosion": 10.0,
                "stock_op": 0.0,
            },
            {
                "material": "MAT-B",
                "material_key": "MAT-B",
                "description": "Material B",
                "um": "UN",
                "group_origin": "IMPORTADO",
                "check": "",
                "in_current_wiu": True,
                "optional_material": None,
                "consumption_by_model": {"MODELO A": 1.0, "MODELO B": 0.0},
                "stock_sap_effective": 5.0,
                "explosion": 0.0,
                "stock_op": 0.0,
            },
        ],
        models=[
            {"name": "MODELO A", "kit_pgd": 10.0},
            {"name": "MODELO B", "kit_pgd": 20.0},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="teste progresso final",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )


def test_final_dpp_analysis_reports_real_monotonic_progress() -> None:
    events: list[tuple[int, str]] = []
    try:
        scenario = _scenario()
        scenario = scenario_service.recalculate_monthly_scenario(
            scenario["scenario_id"],
            {"MODELO A": 12.0, "MODELO B": 18.0},
        )

        result = _summarize_final_dpp_content_with_progress(
            content=_final_dpp_bytes(),
            filename="DPP FINAL JUL.xlsx",
            scenario=scenario,
            analysis_id="analysis-progress-test",
            progress=lambda value, activity: events.append((value, activity)),
        )

        assert result["summary"]["total_materials"] == 2
        assert result["column_comparison"] is not None
        assert events
        assert events[0][0] == 3
        assert events[-1][0] == 99
        assert all(left[0] <= right[0] for left, right in zip(events, events[1:]))
        assert any("Lendo aba DPP" in activity for _, activity in events)
        assert any("Analisando materiais" in activity for _, activity in events)
        assert any("Comparando colunas do DPP" in activity for _, activity in events)
        assert any(progress >= 90 for progress, activity in events if "Comparando colunas do DPP" in activity)
    finally:
        scenario_service._SCENARIOS.clear()
