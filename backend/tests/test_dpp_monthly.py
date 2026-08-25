from datetime import datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app
from app.services.dpp_monthly_service import _map_pgd_to_models
from app.services.dpp_test_service import _compare

MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _save(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_base_dpp() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"
    sheet["A3"] = "KIT Disponivel PGD (JULHO)"
    sheet["E3"] = 8
    sheet["F3"] = 3
    sheet["A5"] = "REAL"
    sheet["E5"] = 8
    sheet["F5"] = 3
    sheet["E8"] = "1000-01"
    sheet["F8"] = "2000-01"
    headers = [
        "Material", "Descrição", "UM", "Grupo Origem", "MODEL-A", "MODEL-B",
        "Check", "WIU", "NEC", "STK 01.07", "EXPLOSÃO 01.07", "OPC",
        "STK OP", "STK TTL", "SALDO", "Preço", "Amount", "Coments",
    ]
    for col, value in enumerate(headers, start=1):
        sheet.cell(9, col).value = value
    sheet.append(["MAT-001", "Material atual", "UN", "Importado", 2.5, 1, "", "", 0, 0, 0, "ALT-001", 0, 0, 0, 0, 0, ""])
    sheet.append(["OLD-001", "Material histórico", "UN", "Importado", 0, 0, "", "", 0, 0, 0, None, 0, 0, 0, 0, 0, ""])
    return _save(workbook)


def build_base_dpp_multiple_opc() -> bytes:
    workbook = load_workbook(BytesIO(build_base_dpp()))
    sheet = workbook["DPP"]
    sheet["L10"] = "ALT-001/ ALT-002"
    return _save(workbook)


def build_wiu() -> bytes:
    workbook = Workbook()
    power = workbook.active
    power.title = "POWER"
    power.append([
        "Código", "Modelo", "Nível", "Componente", "Descrição", "Uso BOM",
        "Unidade de Medida", "Centro", "Grupo Origem", "Total preço médio móvel", "LT alternativa",
    ])
    # A descrição muda no WIU para comprovar que materiais históricos preservam o texto do DPP anterior.
    power.append(["1000-01", "MODEL-A", 1, "MAT-001", "Descrição atualizada no WIU", 2.5, "UN", 1063, "Importado", 0, 1])
    power.append(["2000-01", "MODEL-B", 1, "MAT-001", "Descrição atualizada no WIU", 1, "UN", 1063, "Importado", 0, 1])
    power.append(["1000-01", "MODEL-A", 1, "NEW-001", "Material novo", 1, "UN", 1063, "Importado", 0, 1])
    matrix = workbook.create_sheet("WIU JULHO")
    matrix["E1"] = "1000-01"
    matrix["F1"] = "2000-01"
    matrix["E2"] = "MODEL-A"
    matrix["F2"] = "MODEL-B"
    return _save(workbook)


def build_explosion() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CONSOLIDADO"
    sheet.append(["Material", "Descrição", "UM", "EXPLOSÃO (01.07.2026)"])
    sheet.append(["MAT-001", "Material atual", "UN", 20])
    sheet.append(["NEW-001", "Material novo", "UN", 5])
    return _save(workbook)


def build_stock() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CONSOLIDADO"
    sheet.append(["Material", "Descrição", "UOM", "Controle qualidade", "Utilização livre", "STK TTL (01.07.2026)"])
    sheet.append(["MAT-001", "Material atual", "UN", 0, 100, 100])
    sheet.append(["ALT-001", "Alternativo", "UN", 0, 30, 30])
    sheet.append(["OLD-001", "Material histórico", "UN", 0, 50, 50])
    sheet.append(["NEW-001", "Material novo", "UN", 0, 10, 10])
    return _save(workbook)


def build_stock_multiple_opc() -> bytes:
    workbook = load_workbook(BytesIO(build_stock()))
    sheet = workbook["CONSOLIDADO"]
    sheet.append(["ALT-002", "Alternativo 2", "UN", 0, 20, 20])
    return _save(workbook)


def build_pgd() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PGD TESTE"
    sheet["B7"] = "PRODUCTION"
    sheet["C7"] = datetime(2026, 6, 1)
    sheet["D7"] = datetime(2026, 7, 1)
    sheet["A9"] = "1000-01"
    sheet["B9"] = "MODEL-A"
    sheet["B13"] = "KIT DISPONÍVEL"
    sheet["D13"] = 10
    sheet["A22"] = "2000-01"
    sheet["B22"] = "MODEL-B"
    sheet["B26"] = "KIT DISPONÍVEL"
    sheet["D26"] = 4
    return _save(workbook)


def build_open() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OPEN JULHO"
    sheet.append(["STATUS", "PO", "PI", "LOTE", "MODELO", "Material", "Texto breve", "Qtd.do pedido", "UOM"])
    sheet.append(["EM - Pendente", "PO-1", "PI-1", "43", "MODEL-A", "MAT-001", "Material atual", 40, "UN"])
    return _save(workbook)


def build_expected_dpp() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"
    sheet["A3"] = "KIT Disponivel PGD (JULHO)"
    sheet["E3"] = 10
    sheet["F3"] = 4
    sheet["A5"] = "REAL"
    sheet["E5"] = 8
    sheet["F5"] = 3
    sheet["E8"] = "1000-01"
    sheet["F8"] = "2000-01"
    headers = [
        "Material", "Descrição", "UM", "Grupo Origem", "MODEL-A", "MODEL-B",
        "Check", "WIU", "NEC", "STK 01.07", "EXPLOSÃO 01.07", "OPC",
        "STK OP", "STK TTL", "SALDO", "Preço", "Amount", "Coments",
    ]
    for col, value in enumerate(headers, start=1):
        sheet.cell(9, col).value = value
    sheet.append(["MAT-001", "Material atual", "UN", "Importado", 2.5, 1, "", "", 23, 100, 20, "ALT-001", 30, 150, 127, 0, 0, ""])
    sheet.append(["OLD-001", "Material histórico", "UN", "Importado", 0, 0, "", "", 0, 50, 0, None, 0, 50, 50, 0, 0, ""])
    sheet.append(["NEW-001", "Material novo", "UN", "Importado", 1, 0, "", "", 8, 10, 5, None, 0, 15, 7, 0, 0, ""])
    return _save(workbook)


def build_expected_dpp_with_human_opc() -> bytes:
    workbook = load_workbook(BytesIO(build_expected_dpp()))
    sheet = workbook["DPP"]
    # NEW-001 não existe no DPP anterior. O OPC é uma decisão adicionada durante o mês.
    sheet["L12"] = "ALT-001"
    sheet["M12"] = 30
    sheet["N12"] = 45
    sheet["O12"] = 37
    return _save(workbook)


def build_expected_dpp_with_missing_formulas() -> bytes:
    workbook = load_workbook(BytesIO(build_expected_dpp()))
    sheet = workbook["DPP"]
    # Simula material novo adicionado sem extensão das fórmulas de STK TTL e SALDO.
    sheet["N12"] = None
    sheet["O12"] = None
    return _save(workbook)


def test_monthly_generation_uses_history_pgd_and_opc_stock() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/monthly/generate",
            data={"reference_month": "2026-07"},
            files={
                "base_dpp": ("DPP_JUNHO.xlsx", build_base_dpp(), MIME),
                "wiu": ("WIU_JULHO.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO_JULHO.xlsm", build_explosion(), MIME),
                "stock": ("STK_JULHO.xlsx", build_stock(), MIME),
                "pgd": ("PGD.xlsx", build_pgd(), MIME),
                "open_orders": ("OPEN.xlsx", build_open(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["mode"] == "monthly_dpp"
    assert payload["summary"]["materials"] == 3
    assert payload["summary"]["historical_materials"] == 2
    assert payload["summary"]["historical_outside_wiu"] == 1
    assert payload["summary"]["new_materials_from_wiu"] == 1
    assert payload["summary"]["inherited_optional_materials"] == 1
    assert payload["summary"]["optional_materials_with_stock"] == 1
    assert payload["summary"]["pgd_unresolved_positive"] == 0

    model_by_name = {item["name"]: item for item in payload["models"]}
    assert model_by_name["MODEL-A"]["kit_pgd"] == 10
    assert model_by_name["MODEL-A"]["real"] == 10
    assert model_by_name["MODEL-B"]["kit_pgd"] == 4

    material_by_code = {item["material"]: item for item in payload["materials"]}
    current = material_by_code["MAT-001"]
    assert current["description"] == "Material atual"
    assert current["stock_sap"] == 100
    assert current["explosion"] == 20
    assert current["stock_op"] == 30
    assert current["stock_total"] == 150
    assert current["nec"] == 29
    assert current["balance"] == 121
    assert current["open_investigation"]["pending_records"] == 1

    new_material = material_by_code["NEW-001"]
    assert new_material["description"] == "Material novo"

    historical = material_by_code["OLD-001"]
    assert historical["from_history"] is True
    assert historical["in_current_wiu"] is False
    assert historical["nec"] == 0
    assert historical["balance"] == 50


def test_monthly_multiple_opc_sums_each_stock_entry() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/monthly/generate",
            data={"reference_month": "2026-07"},
            files={
                "base_dpp": ("DPP_JUNHO.xlsx", build_base_dpp_multiple_opc(), MIME),
                "wiu": ("WIU_JULHO.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO_JULHO.xlsm", build_explosion(), MIME),
                "stock": ("STK_JULHO.xlsx", build_stock_multiple_opc(), MIME),
                "pgd": ("PGD.xlsx", build_pgd(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    material = {item["material"]: item for item in payload["materials"]}["MAT-001"]
    assert material["optional_materials"] == ["ALT-001", "ALT-002"]
    assert material["optional_material_canonical"] == "ALT-001 / ALT-002"
    assert material["stock_op"] == 50
    assert material["stock_total"] == 170
    assert len(material["stock_op_sources"]) == 2


def test_monthly_pgd_rules_select_july_variants_and_suppress_mbx_kit() -> None:
    current_models = [
        {"name": "TV 32 LG", "code": "TV32-01"},
        {"name": "TV 32 BOE", "code": "TV32-01"},
        {"name": "TV 50 CHOT", "code": "TV50-01"},
        {"name": "TV 50 CSOT", "code": "TV50-01"},
        {"name": "MBX-01", "code": "MBX-01"},
    ]
    previous_models = [
        {"name": "TV 32 LG", "previous_kit_pgd": 100, "previous_real": 100},
        {"name": "TV 32 BOE", "previous_kit_pgd": 0, "previous_real": 0},
        {"name": "TV 50 CHOT", "previous_kit_pgd": 100, "previous_real": 100},
        {"name": "TV 50 CSOT", "previous_kit_pgd": 0, "previous_real": 0},
        {"name": "MBX-01", "previous_kit_pgd": 0, "previous_real": 0},
    ]
    pgd_models = [
        {"name": "TV 32", "code": "TV32-01", "kit_pgd": 76483, "kit_pgd_raw": 76483, "source": {}},
        {"name": "TV 50", "code": "TV50-01", "kit_pgd": 27100, "kit_pgd_raw": 27100, "source": {}},
        {"name": "MBX-01", "code": "MBX-01", "kit_pgd": 995, "kit_pgd_raw": 995, "source": {}},
    ]

    mapping = _map_pgd_to_models(pgd_models, current_models, previous_models, "2026-07")
    models = {item["name"]: item for item in mapping["models"]}
    assert models["TV 32 BOE"]["kit_pgd"] == 76483
    assert models["TV 32 LG"]["kit_pgd"] == 0
    assert models["TV 50 CSOT"]["kit_pgd"] == 27100
    assert models["TV 50 CHOT"]["kit_pgd"] == 0
    assert models["MBX-01"]["kit_pgd"] == 0
    assert mapping["variant_overrides_applied"] == 2
    assert mapping["kit_overrides_applied"] == 1
    assert mapping["unresolved_positive"] == []


def test_monthly_real_can_be_recalculated_without_reuploading_files() -> None:
    with TestClient(app) as client:
        generated = client.post(
            "/api/dpp/monthly/generate",
            data={"reference_month": "2026-07"},
            files={
                "base_dpp": ("DPP_JUNHO.xlsx", build_base_dpp(), MIME),
                "wiu": ("WIU_JULHO.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO_JULHO.xlsm", build_explosion(), MIME),
                "stock": ("STK_JULHO.xlsx", build_stock(), MIME),
                "pgd": ("PGD.xlsx", build_pgd(), MIME),
            },
        )
        assert generated.status_code == 200
        scenario_id = generated.json()["scenario_id"]

        recalculated = client.post(
            "/api/dpp/monthly/recalculate",
            json={
                "scenario_id": scenario_id,
                "real_by_model": {"MODEL-A": 50, "MODEL-B": 4},
            },
        )

    assert recalculated.status_code == 200
    payload = recalculated.json()
    assert payload["summary"]["models_above_kit"] == 1
    model_by_name = {item["name"]: item for item in payload["models"]}
    assert model_by_name["MODEL-A"]["real"] == 50

    material_by_code = {item["material"]: item for item in payload["materials"]}
    assert material_by_code["MAT-001"]["nec"] == 129
    assert material_by_code["MAT-001"]["balance"] == 21
    assert material_by_code["NEW-001"]["nec"] == 50
    assert material_by_code["NEW-001"]["balance"] == -35
    assert material_by_code["NEW-001"]["status"] == "INVESTIGAR"


def test_monthly_reconstruction_test_matches_known_consolidated_dpp() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/monthly/test",
            data={"reference_month": "2026-07"},
            files={
                "base_dpp": ("DPP_JUNHO.xlsx", build_base_dpp(), MIME),
                "expected_dpp": ("DPP_JULHO_CONSOLIDADO.xlsx", build_expected_dpp(), MIME),
                "wiu": ("WIU_JULHO.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO_JULHO.xlsm", build_explosion(), MIME),
                "stock": ("STK_JULHO.xlsx", build_stock(), MIME),
                "pgd": ("PGD.xlsx", build_pgd(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["mode"] == "monthly_dpp_reconstruction_test"
    assert payload["pass"] is True
    assert payload["status"] == "APROVADO"
    assert payload["summary"]["reference_real_models_applied"] == 2
    assert payload["summary"]["legacy_corrections_total"] == 0
    assert payload["summary"]["human_interventions_total"] == 0
    assert payload["checks"]["materials"]["mismatches"] == 0
    assert payload["checks"]["matrix"]["checked"] == 6
    assert payload["checks"]["matrix"]["mismatches"] == 0
    assert payload["checks"]["description"]["mismatches"] == 0
    assert payload["checks"]["kit_pgd"]["mismatches"] == 0
    assert payload["checks"]["stock_sap"]["mismatches"] == 0
    assert payload["checks"]["explosion"]["mismatches"] == 0
    assert payload["checks"]["stock_op"]["mismatches"] == 0
    assert payload["checks"]["nec"]["mismatches"] == 0
    assert payload["checks"]["balance"]["mismatches"] == 0
    assert payload["mismatches"] == []


def test_reconstruction_classifies_manual_opc_and_derivatives_as_human_intervention() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/monthly/test",
            data={"reference_month": "2026-07"},
            files={
                "base_dpp": ("DPP_JUNHO.xlsx", build_base_dpp(), MIME),
                "expected_dpp": ("DPP_JULHO_COM_OPC_MANUAL.xlsx", build_expected_dpp_with_human_opc(), MIME),
                "wiu": ("WIU_JULHO.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO_JULHO.xlsm", build_explosion(), MIME),
                "stock": ("STK_JULHO.xlsx", build_stock(), MIME),
                "pgd": ("PGD.xlsx", build_pgd(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["pass"] is True
    assert payload["status"] == "APROVADO_COM_INTERVENCOES_HUMANAS"
    assert payload["summary"]["orion_mismatches_total"] == 0
    assert payload["summary"]["human_interventions_total"] == 4
    assert payload["checks"]["optional_material"]["human_interventions"] == 1
    assert payload["checks"]["stock_op"]["human_interventions"] == 1
    assert payload["checks"]["stock_total"]["human_interventions"] == 1
    assert payload["checks"]["balance"]["human_interventions"] == 1
    assert len(payload["human_interventions"]) == 4
    assert payload["mismatches"] == []


def test_reconstruction_classifies_missing_stock_total_and_balance_formulas_as_legacy() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/monthly/test",
            data={"reference_month": "2026-07"},
            files={
                "base_dpp": ("DPP_JUNHO.xlsx", build_base_dpp(), MIME),
                "expected_dpp": ("DPP_JULHO_SEM_FORMULAS.xlsx", build_expected_dpp_with_missing_formulas(), MIME),
                "wiu": ("WIU_JULHO.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO_JULHO.xlsm", build_explosion(), MIME),
                "stock": ("STK_JULHO.xlsx", build_stock(), MIME),
                "pgd": ("PGD.xlsx", build_pgd(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["pass"] is True
    assert payload["status"] == "APROVADO_COM_CORRECOES_LEGADO"
    assert payload["summary"]["orion_mismatches_total"] == 0
    assert payload["summary"]["legacy_corrections_total"] == 2
    assert payload["checks"]["stock_total"]["legacy_corrections"] == 1
    assert payload["checks"]["balance"]["legacy_corrections"] == 1
    assert any("sem fórmula" in item["reason"] for item in payload["legacy_corrections"])
    assert payload["mismatches"] == []


def test_reconstruction_classifies_numeric_text_stock_difference_as_legacy_correction() -> None:
    material = "11219002000049"
    generated = {
        "models": [],
        "summary": {},
        "materials": [
            {
                "material": material,
                "description": "Material teste",
                "um": "UN",
                "group_origin": "Importado",
                "optional_material": None,
                "consumption_by_model": {},
                "stock_sap": 700,
                "stock_source": {"reference": "CONSOLIDADO!A2"},
                "explosion": 0,
                "stock_op": 0,
                "stock_total": 700,
                "nec": 0,
                "balance": 700,
            }
        ],
    }
    expected = {
        "models": {},
        "materials": {
            material: {
                "material": material,
                "material_was_numeric": True,
                "description": "Material teste",
                "um": "UN",
                "group_origin": "Importado",
                "optional_material": None,
                "consumption_by_model": {},
                "stock_sap": 0,
                "explosion": 0,
                "stock_op": 0,
                "stock_total": 0,
                "nec": 0,
                "balance": 0,
            }
        },
    }

    comparison = _compare(generated=generated, expected=expected)
    assert comparison["pass"] is True
    assert comparison["status"] == "APROVADO_COM_CORRECOES_LEGADO"
    assert comparison["checks"]["stock_sap"]["mismatches"] == 0
    assert comparison["checks"]["stock_sap"]["legacy_corrections"] == 1
    assert comparison["checks"]["stock_total"]["legacy_corrections"] == 1
    assert comparison["checks"]["balance"]["legacy_corrections"] == 1
    assert comparison["legacy_corrections_total"] == 3
    assert comparison["mismatches"] == []
