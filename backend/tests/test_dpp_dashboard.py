from io import BytesIO

from openpyxl import Workbook

from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_dashboard_service import summarize_final_dpp_content


def test_latest_monthly_scenario_returns_most_recent_registered_scenario() -> None:
    scenario_service._SCENARIOS.clear()
    try:
        first = scenario_service.register_monthly_scenario(
            materials=[],
            models=[],
            reference_month="2026-06",
            base_summary={},
            scope="Teste junho",
            capabilities={},
            sources=[],
            pending=[],
            diagnostics={},
            pgd_mapping={},
        )
        second = scenario_service.register_monthly_scenario(
            materials=[],
            models=[],
            reference_month="2026-07",
            base_summary={},
            scope="Teste julho",
            capabilities={},
            sources=[],
            pending=[],
            diagnostics={},
            pgd_mapping={},
        )

        latest = scenario_service.get_latest_monthly_scenario()
        assert latest is not None
        assert latest["scenario_id"] == second["scenario_id"]
        assert latest["scenario_id"] != first["scenario_id"]
        assert latest["reference_month"] == "2026-07"
    finally:
        scenario_service._SCENARIOS.clear()


def test_latest_monthly_scenario_returns_none_without_scenarios() -> None:
    scenario_service._SCENARIOS.clear()
    assert scenario_service.get_latest_monthly_scenario() is None


def _final_dpp_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD"
    sheet["E1"] = 100
    sheet["F1"] = 50
    sheet["A2"] = "REAL"
    sheet["E2"] = 120
    sheet["F2"] = 0

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODELO A",
        "MODELO B",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=value)

    rows = [
        ["MAT-A", "A", "UN", "Importado", 1, 0, None, None, 120, 100, 0, "ALT-A", 10, 110, -10, 1, -10],
        ["MAT-B", "B", "UN", "Importado", 1, 0, None, None, 120, 125, 0, None, 0, 125, 5, 1, 5],
        ["MAT-C", "C", "G", "Importado", 0, 1, None, None, 0, 0, 0, None, 0, 0, -100, 1, -100],
    ]
    for row_index, values in enumerate(rows, start=5):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_final_dashboard_summary_uses_consolidated_excel_state() -> None:
    result = summarize_final_dpp_content(_final_dpp_workbook_bytes(), "DPP JULHO.xlsx")
    summary = result["summary"]

    assert summary["pgd_total"] == 150
    assert summary["real_total"] == 120
    assert summary["model_count"] == 2
    assert summary["active_models"] == 1
    assert summary["changed_models"] == 2
    assert summary["below_pgd_models"] == 1
    assert summary["above_pgd_models"] == 1
    assert summary["total_materials"] == 3
    assert summary["critical_materials"] == 1
    assert summary["opc_count"] == 1
    assert summary["risk_models"] == 1
    assert summary["safe_models"] == 0
    assert summary["material_coverage"] == 0
    assert summary["pgd_exposed"] == 100
    assert summary["shared_critical"] == 0
    assert result["critical_materials"] == ["MAT-A"]
    assert result["shared_critical_materials"] == []


def test_final_dashboard_exposes_model_by_model_pgd_real_and_delta() -> None:
    result = summarize_final_dpp_content(_final_dpp_workbook_bytes(), "DPP JULHO.xlsx")

    assert result["models"] == [
        {
            "name": "MODELO A",
            "pgd": 100,
            "real": 120,
            "delta": 20,
            "active": True,
            "changed": True,
            "at_risk": True,
        },
        {
            "name": "MODELO B",
            "pgd": 50,
            "real": 0,
            "delta": -50,
            "active": False,
            "changed": True,
            "at_risk": False,
        },
    ]
