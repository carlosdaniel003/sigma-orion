from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_canonical_export_service import export_monthly_scenario_excel


def _template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD"
    sheet["E1"] = 999
    sheet["F1"] = 999
    sheet["A2"] = "REAL"
    sheet["E2"] = 999
    sheet["F2"] = 999

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODELO A",
        "MODELO B",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
        "Coments",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=value)

    for row, material in ((5, "MAT-A"), (6, "MAT-OLD")):
        sheet.cell(row=row, column=1, value=material)
        for column in range(2, len(headers) + 1):
            sheet.cell(row=row, column=column, value=777)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _scenario() -> dict:
    scenario_service._SCENARIOS.clear()
    scenario = scenario_service.register_monthly_scenario(
        materials=[
            {
                "material": "MAT-A",
                "material_key": "MAT-A",
                "description": "Material A",
                "um": "UN",
                "group_origin": "LOCAL",
                "check": "CHECK-A",
                "in_current_wiu": True,
                "optional_material": "OPC-A",
                "consumption_by_model": {"MODELO A": 2.0, "MODELO B": 1.0},
                "stock_sap_effective": 100.0,
                "explosion": 10.0,
                "stock_op": 5.0,
                "stock_op_sources": [{"material": "OPC-A", "value": 5.0}],
                "price": 2.5,
            },
            {
                "material": "MAT-B",
                "material_key": "MAT-B",
                "description": "Material B",
                "um": "UN",
                "group_origin": "IMPORTADO",
                "check": "",
                "in_current_wiu": False,
                "optional_material": None,
                "consumption_by_model": {"MODELO A": 1.0, "MODELO B": 0.0},
                "stock_sap_effective": 50.0,
                "explosion": 0.0,
                "stock_op": 0.0,
                "stock_op_sources": [],
                "price": 1.25,
            },
        ],
        models=[
            {"name": "MODELO A", "kit_pgd": 10.0},
            {"name": "MODELO B", "kit_pgd": 20.0},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="teste canônico",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )
    return scenario_service.recalculate_monthly_scenario(
        scenario["scenario_id"],
        {"MODELO A": 12.0, "MODELO B": 3.0},
    )


def test_excel_is_canonical_projection_of_dashboard_scenario() -> None:
    try:
        scenario = _scenario()
        content, filename, _media_type = export_monthly_scenario_excel(
            scenario_id=scenario["scenario_id"],
            template_content=_template_bytes(),
            template_filename="DPP JUN_2026.xlsx",
        )

        assert filename == "DPP_ORION_2026_07.xlsx"

        workbook = load_workbook(BytesIO(content), data_only=False)
        try:
            sheet = workbook["DPP"]

            assert sheet["E1"].value == 10
            assert sheet["F1"].value == 20
            assert sheet["E2"].value == 12
            assert sheet["F2"].value == 3
            assert sheet["E3"].value == "=E2-E1"
            assert sheet["F3"].value == "=F2-F1"

            assert sheet["A5"].value == "MAT-A"
            assert sheet["A6"].value == "MAT-B"
            assert sheet["A7"].value in (None, "")

            assert sheet["B5"].value == "Material A"
            assert sheet["C5"].value == "UN"
            assert sheet["D5"].value == "LOCAL"
            assert sheet["E5"].value == 2
            assert sheet["F5"].value == 1
            assert sheet["H5"].value == "WIU"
            assert sheet["J5"].value == 100
            assert sheet["K5"].value == 10
            assert sheet["L5"].value == "OPC-A"

            assert sheet["B6"].value == "Material B"
            assert sheet["H6"].value in (None, "")
            assert sheet["J6"].value == 50

            # Fórmulas equivalentes às fórmulas operacionais do DPP Final.
            assert sheet["G5"].value == '=TEXTJOIN("// ", TRUE, FILTER($E$4:$F$4, E5:F5>0, ""))'
            assert sheet["G6"].value == '=TEXTJOIN("// ", TRUE, FILTER($E$4:$F$4, E6:F6>0, ""))'
            assert sheet["I5"].value == "=SUMPRODUCT($E$2:$F$2,E5:F5)"
            assert sheet["M5"].value == "=VLOOKUP(L5,CONSOLIDADO!$A:$F,6,0)"
            assert sheet["N5"].value == "=J5+K5+M5"
            assert sheet["O5"].value == "=N5-I5"
            assert sheet["P5"].value == 2.5
            assert sheet["Q5"].value == "=P5*O5"

            assert sheet["I6"].value == "=SUMPRODUCT($E$2:$F$2,E6:F6)"
            assert sheet["M6"].value == 0
            assert sheet["N6"].value == "=J6+K6+M6"
            assert sheet["O6"].value == "=N6-I6"
            assert sheet["P6"].value == 1.25
            assert sheet["Q6"].value == "=P6*O6"

            assert sheet["R5"].value in (None, "")
            assert sheet["R6"].value in (None, "")

            support = workbook["CONSOLIDADO"]
            assert support.sheet_state == "hidden"
            assert support["A2"].value == "OPC-A"
            assert support["F2"].value == 5
        finally:
            workbook.close()
    finally:
        scenario_service._SCENARIOS.clear()


def test_excel_audit_reports_incremental_progress_after_serialization() -> None:
    updates: list[tuple[int, str]] = []
    try:
        scenario = _scenario()
        export_monthly_scenario_excel(
            scenario_id=scenario["scenario_id"],
            template_content=_template_bytes(),
            template_filename="DPP JUN_2026.xlsx",
            progress=lambda value, activity: updates.append((value, activity)),
        )

        audit_updates = [
            (value, activity)
            for value, activity in updates
            if activity.startswith("Auditando Excel ORION")
        ]
        assert audit_updates
        assert audit_updates[0][0] >= 96
        assert any("/2 materiais" in activity for _value, activity in audit_updates)
        assert max(value for value, _activity in audit_updates) >= 99
        assert updates[-1] == (99, "Excel ORION consistente com o Dashboard e fórmulas DPP")
    finally:
        scenario_service._SCENARIOS.clear()
