from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app

MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _save(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_wiu(material="MAT-001", um="UN") -> bytes:
    workbook = Workbook()
    power = workbook.active
    power.title = "POWER"
    power.append([
        "Código", "Modelo", "Nível", "Componente", "Descrição", "Uso BOM",
        "Unidade de Medida", "Centro", "Grupo Origem", "Total preço médio móvel", "LT alternativa",
    ])
    power.append(["1000-01", "MODEL-A", 1, material, "Material importado", 2, um, 1063, "Importado", 0, 1])
    power.append(["1000-01", "MODEL-A", 1, material, "Material importado", 0.5, um, 1063, "Importado", 0, 1])
    power.append(["2000-01", "MODEL-B", 1, material, "Material importado", 1, um, 1063, "Importado", 0, 1])
    power.append(["1000-01", "MODEL-A", 1, "MAT-NAC", "Material nacional", 5, "UN", 1063, "Nacional", 0, 1])

    matrix = workbook.create_sheet("WIU JULHO")
    matrix["E1"] = "1000-01"
    matrix["F1"] = "2000-01"
    matrix["E2"] = "MODEL-A"
    matrix["F2"] = "MODEL-B"
    return _save(workbook)


def build_explosion(material="MAT-001", um="UN", value=20) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CONSOLIDADO"
    sheet.append(["Material", "Descrição", "UM", "EXPLOSÃO (01.07.2026)"])
    sheet.append([material, "Material importado", um, value])
    sheet.append(["MAT-NAC", "Material nacional", "UN", 50])
    return _save(workbook)


def build_stock(material="MAT-001", uom="UN", value=100) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CONSOLIDADO"
    sheet.append(["Material", "Descrição", "UOM", "Controle qualidade", "Utilização livre", "STK TTL (01.07.2026)"])
    sheet.append([material, "Material importado", uom, 0, value, value])
    return _save(workbook)


def build_open(material="MAT-001") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OPEN JULHO"
    sheet.append(["STATUS", "PO", "PI", "LOTE", "MODELO", "Material", "Texto breve", "Qtd.do pedido", "UOM"])
    sheet.append(["EM - Pendente", "PO-1", "PI-1", "43", "MODEL-A", material, "Material importado", 40, "UN"])
    sheet.append(["EM - Pendente", "PO-2", "PI-2", "44", "MODEL-A", material, "Material importado", 60, "UN"])
    sheet.append(["EM - Concluído", "PO-3", "PI-3", "42", "MODEL-A", material, "Material importado", 500, "UN"])
    return _save(workbook)


def test_source_consolidation_requires_stock_snapshot() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/consolidate",
            files={
                "wiu": ("WIU.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO.xlsm", build_explosion(), MIME),
            },
        )

    assert response.status_code == 422


def test_source_consolidation_builds_physical_base_from_three_required_sources() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/consolidate",
            files={
                "wiu": ("WIU.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO.xlsm", build_explosion(), MIME),
                "stock": ("STK.xlsx", build_stock(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["mode"] == "source_consolidation"
    assert payload["status"] == "PARCIAL"
    assert payload["summary"]["materials"] == 1
    assert payload["summary"]["models"] == 2
    assert payload["summary"]["explosion_matches"] == 1
    assert payload["summary"]["stock_loaded"] is True
    assert payload["summary"]["stock_matches"] == 1
    assert payload["summary"]["open_loaded"] is False

    material = payload["materials"][0]
    assert material["material"] == "MAT-001"
    assert material["consumption_by_model"]["MODEL-A"] == 2.5
    assert material["consumption_by_model"]["MODEL-B"] == 1
    assert material["used_models"] == ["MODEL-A", "MODEL-B"]
    assert material["explosion"] == 20
    assert material["stock_sap"] == 100
    assert material["available_base"] == 120
    assert material["stock_source"]["cell"] == "F2"
    assert material["open_investigation"]["pending_records"] == 0
    assert payload["capabilities"]["material_code_normalization"] is True
    assert payload["capabilities"]["unit_conversion_enabled"] is False


def test_material_code_normalization_matches_numeric_wiu_with_text_stock() -> None:
    numeric_code = 14731008000020

    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/consolidate",
            files={
                "wiu": ("WIU.xlsx", build_wiu(material=numeric_code, um="G"), MIME),
                "explosion": ("EXPLOSAO.xlsm", build_explosion(material=str(numeric_code), um="G"), MIME),
                "stock": ("STK.xlsx", build_stock(material=str(numeric_code), uom="KG"), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["stock_matches"] == 1
    assert payload["summary"]["explosion_matches"] == 1
    assert payload["summary"]["unit_mismatches"] == 1
    assert payload["summary"]["convertible_unit_mismatches"] == 1

    material = payload["materials"][0]
    assert material["material"] == str(numeric_code)
    assert material["stock_sap"] == 100
    assert material["available_base"] == 120

    conversion = material["unit_conversion"]
    assert conversion["source_unit"] == "KG"
    assert conversion["target_unit"] == "G"
    assert conversion["factor"] == 1000
    assert conversion["supported"] is True
    assert conversion["enabled"] is False
    assert conversion["applied"] is False
    assert conversion["status"] == "CONVERSION_AVAILABLE_DISABLED"


def test_open_is_auxiliary_investigation_and_does_not_change_stock() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/consolidate",
            files={
                "wiu": ("WIU.xlsx", build_wiu(), MIME),
                "explosion": ("EXPLOSAO.xlsm", build_explosion(), MIME),
                "stock": ("STK.xlsx", build_stock(), MIME),
                "open_orders": ("OPEN.xlsx", build_open(), MIME),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["open_loaded"] is True
    assert payload["summary"]["open_pending_materials"] == 1
    assert payload["summary"]["open_pending_rows"] == 2
    assert payload["capabilities"]["open_investigation"] is True

    material = payload["materials"][0]
    assert material["stock_sap"] == 100
    assert material["explosion"] == 20
    assert material["available_base"] == 120

    investigation = material["open_investigation"]
    assert investigation["pending_records"] == 2
    assert investigation["pending_quantity"] == 100
    assert investigation["entries_total"] == 2
    assert {entry["pi"] for entry in investigation["entries"]} == {"PI-1", "PI-2"}
