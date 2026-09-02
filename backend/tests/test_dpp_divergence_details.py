from io import BytesIO

from openpyxl import Workbook

from app.services.dpp_dashboard_service import summarize_final_dpp_content
from app.services.dpp_projection_service import CRITICAL_BALANCE_TOLERANCE, is_critical_material


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD"
    sheet["E1"] = 100
    sheet["A2"] = "REAL"
    sheet["E2"] = 100

    headers = [
        "Material", "Descrição", "UM", "Grupo Origem", "MODELO A", "Check", "WIU", "NEC",
        "STK 01.07", "EXPLOSÃO 01.07", "OPC", "STK OP", "STK TTL", "SALDO", "Preço", "Amount", "Coments",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=header)

    rows = [
        ["MAT-CRIT", "Material crítico", "UN", "LOCAL", 1, None, "X", 120, 80, 10, None, 10, 100, -20, 1, -20, None],
        ["MAT-OK", "Material ok", "UN", "LOCAL", 1, None, "X", 90, 80, 10, None, 10, 100, 10, 1, 10, None],
        ["MAT-G", "Fora escopo", "G", "LOCAL", 1, None, "X", 120, 80, 10, None, 10, 100, -20, 1, -20, None],
    ]
    for row_index, values in enumerate(rows, start=5):
        for column, current in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=current)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_final_dashboard_exposes_material_evidence_for_divergence_explanation() -> None:
    result = summarize_final_dpp_content(_workbook_bytes(), "DPP JULHO.xlsx")
    details = {item["material"]: item for item in result["material_details"]}

    critical = details["MAT-CRIT"]
    assert critical["critical"] is True
    assert critical["nec"] == 120
    assert critical["stock"] == 80
    assert critical["explosion"] == 10
    assert critical["stock_op"] == 10
    assert critical["stock_total"] == 100
    assert critical["balance"] == -20
    assert critical["affected_models"] == ["MODELO A"]

    assert details["MAT-OK"]["critical"] is False
    assert details["MAT-G"]["critical"] is False
    assert result["critical_materials"] == ["MAT-CRIT"]

    rule = result["critical_rule"]
    assert rule["id"] == "un_negative_balance"
    assert rule["tolerance"] == CRITICAL_BALANCE_TOLERANCE
    assert "SALDO = STK TTL" in rule["formula"]
    assert "NEC" in rule["formula"]


def test_critical_rule_uses_one_operational_tolerance() -> None:
    assert is_critical_material("UN", -(CRITICAL_BALANCE_TOLERANCE * 2)) is True
    assert is_critical_material("UN", -(CRITICAL_BALANCE_TOLERANCE / 2)) is False
    assert is_critical_material("G", -100) is False
