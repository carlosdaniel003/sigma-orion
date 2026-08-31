from io import BytesIO

from openpyxl import Workbook

from app.services import dpp_dashboard_service as dashboard_service
from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_dashboard_service import get_column_divergences, summarize_final_dpp_content


def _column(result: dict, name: str) -> dict:
    return next(item for item in result["column_comparison"]["columns"] if item["name"] == name)


def _semantic_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"
    sheet["A1"] = "KIT Disponivel PGD (JULHO)"
    sheet["E1"] = 1
    sheet["F1"] = 1
    sheet["A2"] = "REAL"
    sheet["E2"] = 1
    sheet["F2"] = 1

    headers = [
        "Material", "Descrição", "UM", "Grupo Origem", "TV 50", "TV 50 INNOLUX",
        "Check", "WIU", "NEC", "STK 01.07", "EXPLOSÃO 01.07", "OPC", "STK OP",
        "STK TTL", "SALDO", "Preço", "Amount", "Coments",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column).value = header

    values = [
        "MAT-1", "Material 1", "UN", "LOCAL", 1, 1,
        "TV 50 / TV 50 INNOLUX", "WIU", 2, 10, 0, "OPC-NOVO", 0,
        10, 8, 1, 8, "Ajuste analisado em julho",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(5, column).value = value

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _semantic_scenario() -> dict:
    scenario_service._SCENARIOS.clear()
    scenario = scenario_service.register_monthly_scenario(
        materials=[{
            "material": "MAT-1",
            "description": "Material 1",
            "um": "UN",
            "group_origin": "LOCAL",
            "consumption_by_model": {"TV 50": 1, "TV 50 INNOLUX": 1},
            "check": "TV 50 INNOLUX // TV 50",
            "in_current_wiu": True,
            "optional_material": "OPC-ANTIGO",
            "stock_sap_effective": 10,
            "explosion": 0,
            "stock_op": 0,
            "stock_total": 10,
            "price": 1,
        }],
        models=[
            {"name": "TV 50", "kit_pgd": 1},
            {"name": "TV 50 INNOLUX", "kit_pgd": 1},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="Teste semântico",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )
    return scenario


def test_check_order_opc_updates_and_comments_are_not_divergences() -> None:
    try:
        result = summarize_final_dpp_content(
            _semantic_workbook(),
            "DPP JULHO.xlsx",
            scenario=_semantic_scenario(),
            analysis_id="semantic-comparison",
        )

        check = _column(result, "Check")
        assert check["final_total"] == 1
        assert check["orion_total"] == 1
        assert check["difference_count"] == 0
        assert check["delta"] == 0
        assert check["drilldown_available"] is False

        opc = _column(result, "OPC")
        assert opc["mode"] == "reference_final"
        assert opc["final_total"] == 1
        assert opc["orion_total"] == 1
        assert opc["difference_count"] == 0
        assert opc["delta"] is None
        assert opc["drilldown_available"] is False

        comments = _column(result, "Coments")
        assert comments["mode"] == "contextual"
        assert comments["final_total"] == 1
        assert comments["orion_total"] is None
        assert comments["difference_count"] == 0
        assert comments["drilldown_available"] is False

        assert result["column_comparison"]["reference_columns"] == 1
        assert result["column_comparison"]["contextual_columns"] == 1
    finally:
        scenario_service._SCENARIOS.clear()
        dashboard_service._FINAL_COLUMN_SNAPSHOTS.clear()


def _causal_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"
    sheet["A1"] = "KIT Disponivel PGD (JULHO)"
    kit = [27267, 0, 0, 0]
    final_real = [25000, 16500, 1650, 989]
    names = ["CM-220-N", "BBS-B", "BBS-LBL", "MBX-01"]
    for index, value in enumerate(kit, start=5):
        sheet.cell(1, index).value = value
    sheet["A2"] = "REAL"
    for index, value in enumerate(final_real, start=5):
        sheet.cell(2, index).value = value

    headers = [
        "Material", "Descrição", "UM", "Grupo Origem", *names,
        "Check", "WIU", "NEC", "STK 01.07", "EXPLOSÃO 01.07", "OPC", "STK OP",
        "STK TTL", "SALDO", "Preço", "Amount", "Coments",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column).value = header

    final_nec = sum(final_real)
    stock_total = 100000
    final_balance = stock_total - final_nec
    price = 2
    values = [
        "MAT-REAL", "Material afetado pelo REAL", "UN", "LOCAL", 1, 1, 1, 1,
        "CM-220-N // BBS-B // BBS-LBL // MBX-01", "WIU", final_nec,
        stock_total, 0, None, 0, stock_total, final_balance, price, price * final_balance, None,
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(5, column).value = value

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _causal_scenario() -> dict:
    scenario_service._SCENARIOS.clear()
    names = ["CM-220-N", "BBS-B", "BBS-LBL", "MBX-01"]
    kit = [27267, 0, 0, 0]
    return scenario_service.register_monthly_scenario(
        materials=[{
            "material": "MAT-REAL",
            "description": "Material afetado pelo REAL",
            "um": "UN",
            "group_origin": "LOCAL",
            "consumption_by_model": {name: 1 for name in names},
            "check": " // ".join(names),
            "in_current_wiu": True,
            "optional_material": None,
            "stock_sap_effective": 100000,
            "explosion": 0,
            "stock_op": 0,
            "stock_total": 100000,
            "price": 2,
        }],
        models=[{"name": name, "kit_pgd": value} for name, value in zip(names, kit, strict=True)],
        reference_month="2026-07",
        base_summary={},
        scope="Teste causal",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )


def test_real_changes_are_traced_through_nec_balance_and_amount() -> None:
    analysis_id = "causal-chain"
    try:
        result = summarize_final_dpp_content(
            _causal_workbook(),
            "DPP JULHO.xlsx",
            scenario=_causal_scenario(),
            analysis_id=analysis_id,
        )

        nec = _column(result, "NEC")
        balance = _column(result, "SALDO")
        amount = _column(result, "Amount")
        assert nec["difference_count"] == 1
        assert balance["difference_count"] == 1
        assert amount["difference_count"] == 1

        nec_detail = get_column_divergences(analysis_id, nec["column"])
        balance_detail = get_column_divergences(analysis_id, balance["column"])
        amount_detail = get_column_divergences(analysis_id, amount["column"])

        assert "alterações no REAL de 4 modelo(s)" in nec_detail["items"][0]["reason"]
        assert "CM-220-N" in nec_detail["items"][0]["reason"]
        assert "SALDO divergente porque o STK TTL permaneceu igual e o NEC mudou" in balance_detail["items"][0]["reason"]
        assert "alterações no REAL de 4 modelo(s)" in balance_detail["items"][0]["reason"]
        assert "Amount divergente porque o Preço permaneceu igual e o SALDO mudou" in amount_detail["items"][0]["reason"]
        assert "alterações no REAL de 4 modelo(s)" in amount_detail["items"][0]["reason"]
    finally:
        scenario_service._SCENARIOS.clear()
        dashboard_service._FINAL_COLUMN_SNAPSHOTS.clear()
