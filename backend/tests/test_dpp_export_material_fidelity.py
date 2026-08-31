from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from app.services.dpp_export_material_fidelity_service import finalize_and_audit_material_extent


HEADER_ROW = 9
TEMPLATE_MATERIALS = 4091
SCENARIO_MATERIALS = 4106
FIRST_MATERIAL_ROW = HEADER_ROW + 1
TEMPLATE_LAST_ROW = HEADER_ROW + TEMPLATE_MATERIALS
SCENARIO_LAST_ROW = HEADER_ROW + SCENARIO_MATERIALS


def _scenario(count: int = SCENARIO_MATERIALS) -> dict:
    return {
        "scenario_id": "scenario-material-fidelity",
        "reference_month": "2026-07",
        "materials": [
            {"material": f"MAT-{index:04d}"}
            for index in range(1, count + 1)
        ],
    }


def _workbook_bytes(*, physical_materials: int, table_materials: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    headers = ["Material", "Descrição", "UM", "Grupo Origem"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(HEADER_ROW, column).value = header

    for index in range(1, physical_materials + 1):
        row = HEADER_ROW + index
        sheet.cell(row, 1).value = f"MAT-{index:04d}"
        sheet.cell(row, 2).value = f"Material {index}"
        sheet.cell(row, 3).value = "UN"
        sheet.cell(row, 4).value = "LOCAL"

    table_last_row = HEADER_ROW + table_materials
    table = Table(displayName="DPPTable", ref=f"A{HEADER_ROW}:D{table_last_row}")
    sheet.add_table(table)
    sheet.auto_filter.ref = f"A{HEADER_ROW}:D{table_last_row}"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_export_expands_4091_template_extent_to_4106_dashboard_materials() -> None:
    content = _workbook_bytes(
        physical_materials=SCENARIO_MATERIALS,
        table_materials=TEMPLATE_MATERIALS,
    )

    finalized, audit = finalize_and_audit_material_extent(content, _scenario())

    assert audit["status"] == "DASHBOARD_EXCEL_MATERIALS_OK"
    assert audit["dashboard_materials"] == SCENARIO_MATERIALS
    assert audit["excel_materials"] == SCENARIO_MATERIALS
    assert audit["first_material_row"] == FIRST_MATERIAL_ROW
    assert audit["last_material_row"] == SCENARIO_LAST_ROW
    assert audit["auto_filter_ref"] == f"A{HEADER_ROW}:D{SCENARIO_LAST_ROW}"
    assert audit["table_refs"]["DPPTable"] == f"A{HEADER_ROW}:D{SCENARIO_LAST_ROW}"

    workbook = load_workbook(BytesIO(finalized), data_only=False)
    try:
        sheet = workbook["DPP"]
        assert sheet[f"A{TEMPLATE_LAST_ROW}"].value == f"MAT-{TEMPLATE_MATERIALS:04d}"
        assert sheet[f"A{TEMPLATE_LAST_ROW + 1}"].value == f"MAT-{TEMPLATE_MATERIALS + 1:04d}"
        assert sheet[f"A{SCENARIO_LAST_ROW}"].value == f"MAT-{SCENARIO_MATERIALS:04d}"
        assert sheet.auto_filter.ref == f"A{HEADER_ROW}:D{SCENARIO_LAST_ROW}"
        assert sheet.tables["DPPTable"].ref == f"A{HEADER_ROW}:D{SCENARIO_LAST_ROW}"
    finally:
        workbook.close()


def test_export_refuses_4091_physical_materials_when_dashboard_has_4106() -> None:
    content = _workbook_bytes(
        physical_materials=TEMPLATE_MATERIALS,
        table_materials=TEMPLATE_MATERIALS,
    )

    with pytest.raises(ValueError) as exc_info:
        finalize_and_audit_material_extent(content, _scenario())

    message = str(exc_info.value)
    assert "Dashboard=4106" in message
    assert "Excel=4091" in message
    assert "ausentes=15" in message
