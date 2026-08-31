from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


def build_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Estoque"
    sheet.append(["material", "quantidade"])
    sheet.append(["ABC001", 10])
    sheet.append(["ABC002", 20])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_dpp_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["E2"] = "OK"
    sheet["F2"] = "NG"

    sheet["D3"] = "KIT Disponivel PGD (JULHO)"
    sheet["E3"] = 100
    sheet["F3"] = 50

    sheet["E4"] = "MODEL-A"
    sheet["F4"] = "MODEL-B"

    sheet["D5"] = "REAL"
    sheet["E5"] = 80
    sheet["F5"] = 50

    sheet["E8"] = "1000-01"
    sheet["F8"] = "2000-01"

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODEL-A",
        "MODEL-B",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
        "Coments",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(9, column, value)

    values = [
        "MAT-001",
        "Material de teste",
        "UN",
        "Local",
        2,
        1,
        "MODEL-A// MODEL-B",
        "MODEL-A// MODEL-B",
        210,
        100,
        20,
        None,
        0,
        120,
        -90,
        1.5,
        -135,
        "Investigar divergência",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(10, column, value)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_health() -> None:
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_agent_status_uses_offline_rag_by_default() -> None:
    with TestClient(app) as client:
        response = client.get("/api/agent/status")

    assert response.status_code == 200
    payload = response.json()
    assert payload["provider"] == "mock"
    assert payload["model"] == "mock-local"
    assert payload["mode"] == "offline-rag"
    assert "RAG" in payload["message"]


def test_knowledge_status_loads_versioned_files() -> None:
    with TestClient(app) as client:
        response = client.get("/api/knowledge/status")

    assert response.status_code == 200
    payload = response.json()
    assert payload["mode"] == "lexical-local"
    assert payload["embedding_enabled"] is False
    assert payload["document_count"] >= 4
    assert payload["chunk_count"] >= 4
    assert "guardrails.md" in payload["files"]
    assert "motor-deterministico.md" in payload["files"]


def test_agent_demo_is_explicitly_fake() -> None:
    with TestClient(app) as client:
        response = client.get("/api/agent/demo")

    assert response.status_code == 200
    payload = response.json()
    assert payload["provider"] == "mock"
    assert payload["is_demo"] is True
    assert "Demonstra" in payload["demo_notice"]
    assert payload["risks"]
    assert payload["recommendations"]


def test_agent_demo_chat() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/agent/chat-demo",
            json={"question": "Por que MAT-DEMO-001 está crítico?"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["provider"] == "mock"
    assert payload["is_demo"] is True
    assert "-220" in payload["answer"]


def test_agent_chat_uses_local_rag_without_external_llm() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/agent/chat",
            json={"question": "Qual a fórmula para calcular NEC?"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["provider"] == "local-rag"
    assert payload["model"] == "lexical-local"
    assert payload["is_demo"] is False
    assert "NEC" in payload["answer"]
    assert "REAL" in payload["answer"]
    assert "consumo" in payload["answer"].lower()
    assert payload["knowledge_sources"]
    assert any(
        source in {"motor-deterministico.md", "regras-globais.md"}
        for source in payload["knowledge_sources"]
    )


def test_structured_mock_analysis_is_saved_to_history() -> None:
    request_payload = {
        "metrics": {
            "total_materials": 2,
            "critical": 1,
            "attention": 0,
            "ok": 1,
        },
        "facts": {
            "scenario": "teste automatizado",
            "materials": [
                {"material": "TEST-001", "gap": -10},
                {"material": "TEST-002", "gap": 5},
            ],
        },
        "objective": "Validar o contrato estruturado sem usar uma LLM externa.",
    }

    with TestClient(app) as client:
        analysis_response = client.post("/api/agent/analyze", json=request_payload)
        assert analysis_response.status_code == 200
        analysis = analysis_response.json()
        assert analysis["provider"] == "mock"
        assert analysis["metrics"] == request_payload["metrics"]
        assert analysis["risks"] == []
        assert analysis["recommendations"] == []

        history_response = client.get("/api/analyses/history")
        assert history_response.status_code == 200
        history = history_response.json()
        assert any(item["analysis_id"] == analysis["analysis_id"] for item in history)

        detail_response = client.get(f"/api/analyses/{analysis['analysis_id']}")
        assert detail_response.status_code == 200
        assert detail_response.json()["payload"]["analysis_id"] == analysis["analysis_id"]


def test_save_feedback() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/feedback",
            json={
                "analysis_id": "demo-test",
                "recommendation_id": "rec-test",
                "decision": "approved",
                "comment": None,
            },
        )

    assert response.status_code == 200
    assert response.json()["saved"] is True
    assert isinstance(response.json()["id"], int)


def test_inspect_xlsx() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/files/inspect",
            files={
                "files": (
                    "estoque.xlsx",
                    build_xlsx(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["files_received"] == 1
    assert payload["files"][0]["filename"] == "estoque.xlsx"
    assert payload["files"][0]["sheets"][0]["name"] == "Estoque"
    assert payload["files"][0]["sheets"][0]["rows"] == 2
    assert payload["files"][0]["sheets"][0]["columns"] == ["material", "quantidade"]


def test_dpp_analysis_recalculates_deterministic_fields() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/analyze?divergence_limit=20",
            files={
                "file": (
                    "DPP_TESTE.xlsx",
                    build_dpp_xlsx(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["sheet"] == "DPP"
    assert payload["structure"]["model_count"] == 2
    assert payload["summary"]["total_materials"] == 1
    assert payload["summary"]["materials_to_investigate"] == 1
    assert payload["summary"]["validation"]["nec"]["mismatches"] == 0
    assert payload["summary"]["validation"]["stk_total"]["mismatches"] == 0
    assert payload["summary"]["validation"]["saldo"]["mismatches"] == 0
    assert payload["summary"]["validation"]["amount"]["mismatches"] == 0

    divergence = payload["divergences"][0]
    assert divergence["material"] == "MAT-001"
    assert divergence["python"]["nec"] == 210
    assert divergence["python"]["stock_total"] == 120
    assert divergence["python"]["balance"] == -90
    assert divergence["python"]["amount"] == -135
    assert divergence["python"]["status"] == "INVESTIGAR"