from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_export_service import export_monthly_scenario_excel


def _template_with_stale_rows_and_formulas() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD"
    sheet["E1"] = 10
    sheet["F1"] = 20
    sheet["A2"] = "REAL"
    sheet["E2"] = 10
    sheet["F2"] = 20

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODELO A",
        "MODELO B",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
        "Coments",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)

    for row, material in enumerate(("1001", "1002", "1003"), start=5):
        sheet.cell(row, 1, material)
        sheet.cell(row, 2, f"Histórico {material}")
        sheet.cell(row, 3, "UN")
        sheet.cell(row, 4, "ANTERIOR")
        sheet.cell(row, 5, 9)
        sheet.cell(row, 6, 9)
        sheet.cell(row, 7, f'=IF(H{row}<>"","OK","")')
        sheet.cell(row, 8, f'=IF(E{row}+F{row}>0,"WIU","")')
        sheet.cell(row, 9, f"=$E$2*E{row}+$F$2*F{row}")
        sheet.cell(row, 10, 999)
        sheet.cell(row, 11, 888)
        sheet.cell(row, 12, "OPC-ANT")
        sheet.cell(row, 13, 777)
        sheet.cell(row, 14, f"=J{row}+K{row}+M{row}")
        sheet.cell(row, 15, f"=N{row}-I{row}")
        sheet.cell(row, 16, 123)
        sheet.cell(row, 17, f"=O{row}*P{row}")
        sheet.cell(row, 18, "comentário anterior")

    sheet["B8"] = "Conteúdo abaixo da base deve continuar existindo"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _register_scenario() -> str:
    scenario = scenario_service.register_monthly_scenario(
        materials=[
            {
                "material": "1001",
                "material_key": "1001",
                "description": "Material atual 1001",
                "um": "UN",
                "group_origin": "LOCAL",
                "optional_material": "OPC-01",
                "check": "",
                "in_current_wiu": True,
                "consumption_by_model": {"MODELO A": 2, "MODELO B": 0.5},
                "stock_sap": 40,
                "stock_sap_effective": 40,
                "explosion": 5,
                "stock_op": 5,
                "stock_total": 50,
            },
            {
                "material": "1004",
                "material_key": "1004",
                "description": "Material novo 1004",
                "um": "UN",
                "group_origin": "IMPORTADO",
                "optional_material": None,
                "check": "",
                "in_current_wiu": True,
                "consumption_by_model": {"MODELO A": 1, "MODELO B": 0},
                "stock_sap": 20,
                "stock_sap_effective": 20,
                "explosion": 3,
                "stock_op": 0,
                "stock_total": 23,
            },
        ],
        models=[
            {"name": "MODELO A", "code": "A-01", "kit_pgd": 10, "real": 10},
            {"name": "MODELO B", "code": "B-01", "kit_pgd": 20, "real": 20},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="teste consistência",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )
    return scenario["scenario_id"]


def test_export_material_rows_are_exactly_the_dashboard_scenario_and_stale_rows_are_removed():
    scenario_service._SCENARIOS.clear()
    try:
        content, _filename, _media_type = export_monthly_scenario_excel(
            scenario_id=_register_scenario(),
            template_content=_template_with_stale_rows_and_formulas(),
            template_filename="DPP JUN_2026.xlsx",
        )

        workbook = load_workbook(BytesIO(content), data_only=False)
        try:
            sheet = workbook["DPP"]
            assert [sheet["A5"].value, sheet["A6"].value] == ["1001", "1004"]
            assert sheet["A7"].value is None
            assert sheet["B7"].value == "Conteúdo abaixo da base deve continuar existindo"

            # Cabeçalhos datados precisam receber os valores atuais, não os valores do DPP anterior.
            assert sheet["J5"].value == 40
            assert sheet["J6"].value == 20
            assert sheet["K5"].value == 5
            assert sheet["K6"].value == 3

            # Campos não calculados pelo ORION não podem vazar do mês anterior.
            assert sheet["P5"].value is None
            assert sheet["Q5"].value is None
            assert sheet["R5"].value is None
        finally:
            workbook.close()
    finally:
        scenario_service._SCENARIOS.clear()


def test_export_propagates_structural_formulas_to_every_orion_material_row():
    scenario_service._SCENARIOS.clear()
    try:
        content, _filename, _media_type = export_monthly_scenario_excel(
            scenario_id=_register_scenario(),
            template_content=_template_with_stale_rows_and_formulas(),
            template_filename="DPP JUN_2026.xlsx",
        )

        workbook = load_workbook(BytesIO(content), data_only=False)
        try:
            sheet = workbook["DPP"]
            for row in (5, 6):
                for column in (7, 8, 9, 14, 15):
                    value = sheet.cell(row, column).value
                    assert isinstance(value, str)
                    assert value.startswith("=")

            assert "E6" in sheet["I6"].value
            assert "F6" in sheet["I6"].value
            assert "J6" in sheet["N6"].value
            assert "N6" in sheet["O6"].value
        finally:
            workbook.close()
    finally:
        scenario_service._SCENARIOS.clear()
