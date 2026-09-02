from io import BytesIO
from time import monotonic, sleep

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from app.services import dpp_export_progress_service as progress_service
from app.services.dpp_export_formula_fidelity_service import finalize_and_audit_dpp_formulas
from app.services.dpp_scenario_service import register_monthly_scenario, recalculate_monthly_scenario


def _formula_text(value):
    if isinstance(value, ArrayFormula):
        return value.text
    return value


def _workbook_without_formulas() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["D3"] = "KIT Disponivel PGD (JUNHO)"
    sheet["E3"] = 0
    sheet["F3"] = 0
    sheet["G3"] = 27267
    sheet["E4"] = "CM-150 L1"
    sheet["F4"] = "CM-150 L2"
    sheet["G4"] = "CM-220-N"
    sheet["D5"] = "REAL"
    sheet["E5"] = 0
    sheet["F5"] = 0
    sheet["G5"] = 25000
    # G6 propositalmente vazio: regressão observada no DPP_ORION_2026_07 (6).

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "CM-150 L1",
        "CM-150 L2",
        "CM-220-N",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(9, column).value = header

    sheet["A10"] = "MAT-A"
    sheet["B10"] = "Material A"
    sheet["C10"] = "UN"
    sheet["D10"] = "LOCAL"
    sheet["E10"] = 0
    sheet["F10"] = 0
    sheet["G10"] = 1
    sheet["I10"] = "WIU"
    sheet["K10"] = 100
    sheet["L10"] = 10
    sheet["M10"] = "OPC-A"
    sheet["N10"] = 5
    sheet["Q10"] = 2.5

    sheet["A11"] = "MAT-B"
    sheet["B11"] = "Material B"
    sheet["C11"] = "UN"
    sheet["D11"] = "IMPORTADO"
    sheet["E11"] = 1
    sheet["F11"] = 0
    sheet["G11"] = 0
    sheet["K11"] = 50
    sheet["L11"] = 0
    sheet["N11"] = 0
    sheet["Q11"] = 1.25

    consolidado = workbook.create_sheet("CONSOLIDADO")
    consolidado.sheet_state = "hidden"
    consolidado["A1"] = "Material"
    consolidado["F1"] = "STK OP"
    consolidado["A2"] = "OPC-A"
    consolidado["F2"] = 5

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _scenario() -> dict:
    scenario = register_monthly_scenario(
        materials=[
            {
                "material": "MAT-A",
                "material_key": "MAT-A",
                "description": "Material A",
                "um": "UN",
                "group_origin": "LOCAL",
                "optional_material": "OPC-A",
                "consumption_by_model": {"CM-150 L1": 0, "CM-150 L2": 0, "CM-220-N": 1},
                "stock_sap_effective": 100,
                "explosion": 10,
                "stock_op": 5,
                "stock_op_sources": [{"material": "OPC-A", "value": 5}],
                "price": 2.5,
            },
            {
                "material": "MAT-B",
                "material_key": "MAT-B",
                "description": "Material B",
                "um": "UN",
                "group_origin": "IMPORTADO",
                "optional_material": None,
                "consumption_by_model": {"CM-150 L1": 1, "CM-150 L2": 0, "CM-220-N": 0},
                "stock_sap_effective": 50,
                "explosion": 0,
                "stock_op": 0,
                "stock_op_sources": [],
                "price": 1.25,
            },
        ],
        models=[
            {"name": "CM-150 L1", "kit_pgd": 0},
            {"name": "CM-150 L2", "kit_pgd": 0},
            {"name": "CM-220-N", "kit_pgd": 27267},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="teste de fidelidade de fórmulas",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )
    return recalculate_monthly_scenario(
        scenario["scenario_id"],
        {"CM-150 L1": 0, "CM-150 L2": 0, "CM-220-N": 25000},
    )


def test_final_formula_barrier_restores_every_deterministic_dpp_formula() -> None:
    scenario = _scenario()
    content = finalize_and_audit_dpp_formulas(_workbook_without_formulas(), scenario)

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        sheet = workbook["DPP"]
        assert sheet["D3"].value == "KIT Disponivel PGD (JULHO)"
        assert sheet["G6"].value == "=G5-G3"

        assert _formula_text(sheet["H10"].value) == (
            '=_xlfn.TEXTJOIN("// ", TRUE, _xlfn._xlws.FILTER($E$9:$G$9, E10:G10>0, ""))'
        )
        assert sheet["J10"].value == "=SUMPRODUCT($E$5:$G$5,E10:G10)"
        assert sheet["N10"].value == "=VLOOKUP(M10,CONSOLIDADO!$A:$F,6,0)"
        assert sheet["O10"].value == "=K10+L10+N10"
        assert sheet["P10"].value == "=O10-J10"
        assert sheet["R10"].value == "=Q10*P10"

        assert _formula_text(sheet["H11"].value) == (
            '=_xlfn.TEXTJOIN("// ", TRUE, _xlfn._xlws.FILTER($E$9:$G$9, E11:G11>0, ""))'
        )
        assert sheet["J11"].value == "=SUMPRODUCT($E$5:$G$5,E11:G11)"
        assert sheet["N11"].value == 0
        assert sheet["O11"].value == "=K11+L11+N11"
        assert sheet["P11"].value == "=O11-J11"
        assert sheet["R11"].value == "=Q11*P11"
    finally:
        workbook.close()


def test_export_job_applies_formula_barrier_to_exact_download_bytes(monkeypatch) -> None:
    scenario = _scenario()
    raw = _workbook_without_formulas()

    def fake_canonical_export(**_kwargs):
        return raw, "DPP_ORION_2026_07.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    monkeypatch.setattr(progress_service, "export_monthly_scenario_excel", fake_canonical_export)

    job = progress_service.start_export_job(
        scenario_id=scenario["scenario_id"],
        base_filename="DPP JUN_2026.xlsx",
        base_content=b"template",
    )
    job_id = job["job_id"]
    deadline = monotonic() + 5
    snapshot = job
    while snapshot["status"] not in {"completed", "failed"} and monotonic() < deadline:
        sleep(0.02)
        snapshot = progress_service.get_export_job(job_id)
        assert snapshot is not None

    assert snapshot["status"] == "completed", snapshot.get("error")
    assert snapshot["result"]["formula_audit"] == "DPP_FINAL_RULES_OK"

    download = progress_service.get_export_download(job_id)
    assert download is not None
    content, _filename, _media_type = download
    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        sheet = workbook["DPP"]
        assert sheet["G6"].value == "=G5-G3"
        assert _formula_text(sheet["H10"].value).startswith("=_xlfn.TEXTJOIN")
        assert sheet["J10"].value == "=SUMPRODUCT($E$5:$G$5,E10:G10)"
        assert sheet["O10"].value == "=K10+L10+N10"
        assert sheet["P10"].value == "=O10-J10"
        assert sheet["R10"].value == "=Q10*P10"
    finally:
        workbook.close()
