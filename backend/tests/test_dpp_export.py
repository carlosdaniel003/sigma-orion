from io import BytesIO
from time import monotonic, sleep

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.services.dpp_export_progress_service import get_export_download, get_export_job, start_export_job
from app.services.dpp_export_service import export_monthly_scenario_excel
from app.services.dpp_scenario_service import register_monthly_scenario


def _previous_dpp_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD (JUNHO)"
    sheet["E1"] = 111
    sheet["F1"] = 222
    sheet["A2"] = "REAL"
    sheet["E2"] = 111
    sheet["F2"] = 222
    sheet.append([])
    sheet.append([
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODELO A",
        "MODELO B",
        "Check",
        "OPC",
        "STK SAP",
        "Explosão",
        "STK OP",
        "STK TTL",
        "NEC",
        "SALDO",
    ])
    sheet.append([
        "1001",
        "Descrição histórica",
        "UN",
        "ANTERIOR",
        99,
        99,
        "ANTERIOR",
        "OPC-ANTERIOR",
        999,
        999,
        999,
        999,
        999,
        999,
    ])

    sheet["A4"].fill = PatternFill(fill_type="solid", fgColor="123456")
    sheet["A5"].fill = PatternFill(fill_type="solid", fgColor="654321")
    sheet.column_dimensions["A"].width = 18
    workbook.create_sheet("Resumo de Análise")["A1"] = "Layout preservado"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _scenario_id() -> str:
    scenario = register_monthly_scenario(
        materials=[
            {
                "material": "1001",
                "material_key": "1001",
                "description": "Material ORION",
                "um": "UN",
                "group_origin": "LOCAL",
                "optional_material": "OPC-01",
                "check": "OK",
                "consumption_by_model": {"MODELO A": 2, "MODELO B": 0.5},
                "stock_sap": 40,
                "stock_sap_effective": 40,
                "explosion": 5,
                "stock_op": 5,
                "stock_total": 50,
            },
            {
                "material": "1002",
                "material_key": "1002",
                "description": "Material novo ORION",
                "um": "UN",
                "group_origin": "IMPORTADO",
                "optional_material": None,
                "check": "OK",
                "consumption_by_model": {"MODELO A": 1},
                "stock_sap": 20,
                "stock_sap_effective": 20,
                "explosion": 0,
                "stock_op": 0,
                "stock_total": 20,
            },
        ],
        models=[
            {"name": "MODELO A", "code": "A-01", "kit_pgd": 10, "real": 10},
            {"name": "MODELO B", "code": "B-01", "kit_pgd": 20, "real": 20},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="teste",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )
    return scenario["scenario_id"]


def test_export_uses_previous_dpp_only_as_layout_and_writes_orion_values():
    content, filename, media_type = export_monthly_scenario_excel(
        scenario_id=_scenario_id(),
        template_content=_previous_dpp_workbook(),
        template_filename="DPP JUN_2026.xlsx",
    )

    assert filename == "DPP_ORION_2026_07.xlsx"
    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        assert workbook.sheetnames == ["DPP", "Resumo de Análise"]
        sheet = workbook["DPP"]

        # O cabeçalho/estrutura visual é o do arquivo anterior, mas KIT e REAL são do ORION.
        assert sheet["E1"].value == 10
        assert sheet["F1"].value == 20
        assert sheet["E2"].value == 10
        assert sheet["F2"].value == 20

        # Material já existente: todos os valores operacionais antigos foram substituídos.
        assert sheet["B5"].value == "Material ORION"
        assert sheet["D5"].value == "LOCAL"
        assert sheet["E5"].value == 2
        assert sheet["F5"].value == 0.5
        assert sheet["H5"].value == "OPC-01"
        assert sheet["I5"].value == 40
        assert sheet["J5"].value == 5
        assert sheet["K5"].value == 5
        assert sheet["L5"].value == 50
        assert sheet["M5"].value == 30
        assert sheet["N5"].value == 20

        # Material novo do mês é incluído copiando apenas o layout da linha anterior.
        assert sheet["A6"].value == "1002"
        assert sheet["B6"].value == "Material novo ORION"
        assert sheet["D6"].value == "IMPORTADO"
        assert sheet["E6"].value == 1
        assert sheet["M6"].value == 10
        assert sheet["N6"].value == 10
        assert sheet["A6"].fill.fgColor.rgb.endswith("654321")

        assert sheet["A4"].fill.fgColor.rgb.endswith("123456")
        assert sheet.column_dimensions["A"].width == 18
        assert workbook["Resumo de Análise"]["A1"].value == "Layout preservado"
    finally:
        workbook.close()


def test_export_reports_monotonic_progress_from_real_processing_stages():
    updates: list[tuple[int, str]] = []

    export_monthly_scenario_excel(
        scenario_id=_scenario_id(),
        template_content=_previous_dpp_workbook(),
        template_filename="DPP JUN_2026.xlsx",
        progress=lambda value, activity: updates.append((value, activity)),
    )

    values = [value for value, _activity in updates]
    assert values
    assert values == sorted(values)
    assert values[0] == 4
    assert values[-1] == 99
    assert 90 in values
    assert 94 in values
    assert any("Preenchendo materiais" in activity for _value, activity in updates)


def test_export_job_finishes_at_100_and_keeps_generated_file_for_download():
    job = start_export_job(
        scenario_id=_scenario_id(),
        base_filename="DPP JUN_2026.xlsx",
        base_content=_previous_dpp_workbook(),
    )
    job_id = job["job_id"]
    deadline = monotonic() + 5
    snapshot = job

    while snapshot["status"] not in {"completed", "failed"} and monotonic() < deadline:
        sleep(0.02)
        snapshot = get_export_job(job_id)
        assert snapshot is not None

    assert snapshot["status"] == "completed"
    assert snapshot["progress"] == 100
    assert snapshot["result"]["filename"] == "DPP_ORION_2026_07.xlsx"

    download = get_export_download(job_id)
    assert download is not None
    content, filename, media_type = download
    assert content
    assert filename == "DPP_ORION_2026_07.xlsx"
    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
