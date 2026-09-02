from io import BytesIO

from openpyxl import Workbook

from app.services import dpp_dashboard_service as dashboard_service
from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_dashboard_service import get_column_divergences, summarize_final_dpp_content


def _final_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["D1"] = "KIT Disponivel PGD (JULHO)"
    kit = [27267, 0, 0, 0]
    final_real = [25000, 16500, 1650, 989]
    models = ["CM-220-N", "BBS-B", "BBS-LBL", "MBX-01"]
    for index, value in enumerate(kit, start=5):
        sheet.cell(1, index).value = value
    for index, name in enumerate(models, start=5):
        sheet.cell(2, index).value = name
    sheet["D3"] = "REAL"
    for index, value in enumerate(final_real, start=5):
        sheet.cell(3, index).value = value

    headers = [
        "Material", "Descrição", "UM", "Grupo Origem",
        *models,
        "Check", "WIU", "NEC", "STK 01.07", "EXPLOSÃO 01.07",
        "OPC", "STK OP", "STK TTL", "SALDO", "Preço", "Amount", "Coments",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(5, column).value = header

    final_nec = sum(final_real)
    values = [
        "MAT-RESINA", "Resina teste", "G", "LOCAL",
        1, 1, 1, 1,
        None, "WIU", final_nec, 0, 0, None, 0, 0, -final_nec, 1, -final_nec, None,
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(6, column).value = value

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _scenario() -> dict:
    scenario_service._SCENARIOS.clear()
    return scenario_service.register_monthly_scenario(
        materials=[{
            "material": "MAT-RESINA",
            "description": "Resina teste",
            "um": "G",
            "group_origin": "LOCAL",
            "consumption_by_model": {
                "CM-220-N": 1,
                "BBS-B": 1,
                "BBS-LBL": 1,
                "MBX-01": 1,
            },
            "check": "",
            "in_current_wiu": True,
            "optional_material": None,
            "stock_sap_effective": 0,
            "explosion": 0,
            "stock_op": 0,
            "stock_total": 0,
            "price": 1,
        }],
        models=[
            {"name": "CM-220-N", "kit_pgd": 27267},
            {"name": "BBS-B", "kit_pgd": 0},
            {"name": "BBS-LBL", "kit_pgd": 0},
            {"name": "MBX-01", "kit_pgd": 0},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="Teste explicação NEC",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )


def test_nec_divergence_explains_four_real_changes_instead_of_only_delta() -> None:
    analysis_id = "nec-real-cause-test"
    try:
        scenario = _scenario()
        result = summarize_final_dpp_content(
            _final_workbook(),
            "DPP JULHO.xlsx",
            scenario=scenario,
            analysis_id=analysis_id,
        )
        nec = next(item for item in result["column_comparison"]["columns"] if item["name"] == "NEC")
        assert nec["difference_count"] == 1

        details = get_column_divergences(analysis_id, nec["column"], offset=0, limit=10)
        assert details is not None
        assert details["total"] == 1
        reason = details["items"][0]["reason"]

        assert "alterações no REAL de 4 modelo(s)" in reason
        assert "CM-220-N (27.267 → 25.000)" in reason
        assert "BBS-B (0 → 16.500)" in reason
        assert "BBS-LBL (0 → 1.650)" in reason
        assert "MBX-01 (0 → 989)" in reason
        assert "explicada por essas alterações no REAL" in reason
        assert "difere do consolidado final em" not in reason
    finally:
        scenario_service._SCENARIOS.clear()
        dashboard_service._FINAL_COLUMN_SNAPSHOTS.clear()
