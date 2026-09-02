from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.services.dpp_export_postprocess_service import enforce_final_dpp_header_and_gap


def _orion_like_export() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    # Estrutura real observada no DPP de julho.
    sheet["D3"] = "KIT Disponivel PGD (JUNHO)"
    sheet["E3"] = 0
    sheet["F3"] = 0
    sheet["G3"] = 27267

    sheet["E4"] = "CM-150 L1"
    sheet["F4"] = "CM-150 L2"
    sheet["G4"] = "CM-220-N"

    sheet["D5"] = "REAL"
    sheet["E5"] = 0
    sheet["F5"] = 0
    sheet["G5"] = 25000

    # A versão problemática do ORION chegava aqui sem a fórmula de diferença.
    sheet["E6"] = None
    sheet["F6"] = None
    sheet["G6"] = None

    headers = ["Material", "Descrição", "UM", "Grupo Origem", "CM-150 L1", "CM-150 L2", "CM-220-N", "Check"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=9, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_postprocess_uses_serialized_kit_and_real_for_exact_gap_formula() -> None:
    content = enforce_final_dpp_header_and_gap(_orion_like_export(), "2026-07")

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        sheet = workbook["DPP"]
        assert sheet["D3"].value == "KIT Disponivel PGD (JULHO)"
        assert sheet["E6"].value in (None, "")
        assert sheet["F6"].value in (None, "")
        assert sheet["G6"].value == "=G5-G3"
    finally:
        workbook.close()


def test_postprocess_removes_gap_when_real_equals_kit() -> None:
    workbook = load_workbook(BytesIO(_orion_like_export()), data_only=False)
    sheet = workbook["DPP"]
    sheet["G5"] = 27267
    sheet["G6"] = "=G5-G3"
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    content = enforce_final_dpp_header_and_gap(output.getvalue(), "2026-07")
    audited = load_workbook(BytesIO(content), data_only=False)
    try:
        assert audited["DPP"]["G6"].value in (None, "")
    finally:
        audited.close()
