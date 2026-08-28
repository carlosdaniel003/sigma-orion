from io import BytesIO

from openpyxl import Workbook

from app.services import dpp_dashboard_service as dashboard_service
from app.services import dpp_scenario_service as scenario_service
from app.services.dpp_dashboard_service import get_column_divergences, summarize_final_dpp_content


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["A1"] = "KIT Disponivel PGD"
    sheet["E1"] = 100
    sheet["F1"] = 50
    sheet["A2"] = "REAL"
    sheet["E2"] = 120
    sheet["F2"] = 0

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODELO A",
        "MODELO B",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
        "Coments",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=value)

    rows = [
        ["MAT-A", "A", "UN", "Importado", 1, 0, "INVESTIGAR", "X", 120, 100, 0, "ALT-A", 10, 110, -10, 1, -10, "Revisar"],
        ["MAT-B", "B", "UN", "Importado", 1, 0, "OK", "X", 120, 125, 0, None, 0, 125, 5, 1, 5, None],
        ["MAT-C", "C", "G", "Importado", 0, 1, "FORA_ESCOPO_UM", None, 0, 0, 0, None, 0, 0, 0, 1, 0, None],
    ]
    for row_index, values in enumerate(rows, start=5):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _scenario() -> dict:
    scenario_service._SCENARIOS.clear()
    scenario = scenario_service.register_monthly_scenario(
        materials=[
            {
                "material": "MAT-A",
                "description": "A",
                "um": "UN",
                "group_origin": "Importado",
                "consumption_by_model": {"MODELO A": 1, "MODELO B": 0},
                "check": "",
                "in_current_wiu": True,
                "optional_material": "ALT-A",
                "stock_sap_effective": 100,
                "explosion": 0,
                "stock_op": 10,
                "stock_total": 110,
            },
            {
                "material": "MAT-B",
                "description": "B ORION",
                "um": "UN",
                "group_origin": "Importado",
                "consumption_by_model": {"MODELO A": 1, "MODELO B": 0},
                "check": "",
                "in_current_wiu": True,
                "optional_material": None,
                "stock_sap_effective": 125,
                "explosion": 0,
                "stock_op": 0,
                "stock_total": 125,
            },
            {
                "material": "MAT-C",
                "description": "C",
                "um": "G",
                "group_origin": "Importado",
                "consumption_by_model": {"MODELO A": 0, "MODELO B": 1},
                "check": "",
                "in_current_wiu": False,
                "optional_material": None,
                "stock_sap_effective": 0,
                "explosion": 0,
                "stock_op": 0,
                "stock_total": 0,
            },
        ],
        models=[
            {"name": "MODELO A", "kit_pgd": 100},
            {"name": "MODELO B", "kit_pgd": 50},
        ],
        reference_month="2026-07",
        base_summary={},
        scope="Teste",
        capabilities={},
        sources=[],
        pending=[],
        diagnostics={},
        pgd_mapping={},
    )
    return scenario_service.recalculate_monthly_scenario(
        scenario["scenario_id"],
        {"MODELO A": 120, "MODELO B": 0},
    )


def _column(result: dict, name: str) -> dict:
    return next(item for item in result["column_comparison"]["columns"] if item["name"] == name)


def test_column_comparison_reports_totals_and_cell_divergences_without_listing_items() -> None:
    try:
        result = summarize_final_dpp_content(
            _workbook_bytes(),
            "DPP JULHO.xlsx",
            scenario=_scenario(),
        )
        comparison = result["column_comparison"]

        assert comparison["final_materials"] == 3
        assert comparison["orion_materials"] == 3
        assert comparison["columns_total"] == 18
        assert comparison["unsupported_columns"] == 3

        material = _column(result, "Material")
        assert material["final_total"] == 3
        assert material["orion_total"] == 3
        assert material["delta"] == 0
        assert material["difference_count"] == 0

        description = _column(result, "Descrição")
        assert description["final_total"] == 3
        assert description["orion_total"] == 3
        assert description["delta"] == 0
        assert description["difference_count"] == 1
        assert description["drilldown_available"] is False

        model_a = _column(result, "MODELO A")
        assert model_a["final_total"] == 2
        assert model_a["orion_total"] == 2
        assert model_a["delta"] == 0
        assert model_a["difference_count"] == 0

        wiu = _column(result, "WIU")
        assert wiu["final_total"] == 2
        assert wiu["orion_total"] == 2
        assert wiu["difference_count"] == 0

        nec = _column(result, "NEC")
        assert nec["final_total"] == 240
        assert nec["orion_total"] == 240
        assert nec["difference_count"] == 0

        price = _column(result, "Preço")
        assert price["final_total"] == 3
        assert price["orion_total"] is None
        assert price["delta"] is None
        assert price["difference_count"] is None
        assert price["supported"] is False

        comments = _column(result, "Coments")
        assert comments["final_total"] == 1
        assert comments["orion_total"] is None
        assert comments["supported"] is False
    finally:
        scenario_service._SCENARIOS.clear()


def test_column_divergence_drilldown_is_paginated_and_explains_the_rule() -> None:
    analysis_id = "column-pagination-test"
    try:
        scenario = _scenario()
        stored = scenario_service._SCENARIOS[scenario["scenario_id"]]
        for material in stored["materials"]:
            material["description"] = f"{material['description']} ORION"
        scenario = scenario_service.get_monthly_scenario(scenario["scenario_id"])

        result = summarize_final_dpp_content(
            _workbook_bytes(),
            "DPP JULHO.xlsx",
            scenario=scenario,
            analysis_id=analysis_id,
        )
        description = _column(result, "Descrição")
        assert description["difference_count"] == 3
        assert description["drilldown_available"] is True

        first_page = get_column_divergences(analysis_id, description["column"], offset=0, limit=1)
        assert first_page is not None
        assert first_page["total"] == 3
        assert first_page["returned"] == 1
        assert first_page["has_previous"] is False
        assert first_page["has_next"] is True
        assert len(first_page["items"]) == 1
        assert "valores normalizados" in first_page["rule"]["criterion"]
        assert "diferentes" in first_page["items"][0]["reason"].lower()

        second_page = get_column_divergences(analysis_id, description["column"], offset=1, limit=1)
        assert second_page is not None
        assert second_page["total"] == 3
        assert second_page["returned"] == 1
        assert second_page["has_previous"] is True
        assert second_page["has_next"] is True
        assert second_page["items"][0]["material"] != first_page["items"][0]["material"]
    finally:
        scenario_service._SCENARIOS.clear()
        dashboard_service._FINAL_COLUMN_SNAPSHOTS.clear()
