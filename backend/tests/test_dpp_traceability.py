from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


def build_traceable_dpp() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP"

    sheet["E2"] = "NG"
    sheet["D3"] = "KIT Disponivel PGD (JULHO)"
    sheet["E3"] = 100
    sheet["D5"] = "REAL"
    sheet["E5"] = 100
    sheet["E8"] = "1000-01"

    headers = [
        "Material",
        "Descrição",
        "UM",
        "Grupo Origem",
        "MODEL-A",
        "Check",
        "WIU",
        "NEC",
        "STK 01.07",
        "EXPLOSÃO 01.07",
        "OPC",
        "STK OP",
        "STK TTL",
        "SALDO",
        "Preço",
        "Amount",
        "Coments",
    ]
    for column, value in enumerate(headers, start=1):
        sheet.cell(9, column, value)

    values = [
        "MAT-TRACE-001",
        "Material rastreável",
        "UN",
        "Importado",
        2,
        "MODEL-A",
        "MODEL-A",
        200,
        50,
        0,
        None,
        0,
        50,
        -150,
        1,
        -150,
        "Teste de rastreabilidade",
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(10, column, value)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_dpp_analysis_returns_original_excel_location() -> None:
    with TestClient(app) as client:
        response = client.post(
            "/api/dpp/analyze?divergence_limit=20",
            files={
                "file": (
                    "DPP_TRACE.xlsx",
                    build_traceable_dpp(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    payload = response.json()

    expected_source = {
        "sheet": "DPP",
        "row": 10,
        "material_cell": "A10",
        "reference": "DPP!A10",
    }
    assert payload["materials"][0]["source"] == expected_source
    assert payload["divergences"][0]["source"] == expected_source
